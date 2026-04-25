from flask import Flask, render_template, request, jsonify, redirect, url_for, session, send_file, make_response
from flask_sqlalchemy import SQLAlchemy
from flask_login import LoginManager, UserMixin, login_user, logout_user, login_required, current_user
from flask_socketio import SocketIO, emit, join_room, leave_room
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
from datetime import datetime, timedelta, timezone
import json
import os
import re
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from io import BytesIO
from itsdangerous import URLSafeTimedSerializer, BadSignature, SignatureExpired
from sqlalchemy import func
from sqlalchemy.orm import joinedload
from sqlalchemy.schema import UniqueConstraint
import smtplib
import ssl
from email.mime.text import MIMEText

# Project directory first, then load .env from project (not dependent on cwd)
_APP_DIR = os.path.dirname(os.path.abspath(__file__))
try:
    from dotenv import load_dotenv
    load_dotenv(os.path.join(_APP_DIR, '.env'))
except ImportError:
    pass  # python-dotenv not installed; fall back to OS env vars

# Always load templates from this project folder (avoids stale/wrong UI when cwd differs)
app = Flask(__name__, template_folder=os.path.join(_APP_DIR, 'templates'))

# Production is opt-in via PRODUCTION=1 or FLASK_ENV=production. In that mode we
# refuse to start without a real SECRET_KEY and without an explicit CORS allow-list.
# Dev / test runs (plain `python app.py`, pytest) keep the old permissive defaults so
# the app still boots out of the box.
_IS_TESTING = os.environ.get('TESTING') == '1'
_IS_PRODUCTION = (
    os.environ.get('PRODUCTION') == '1'
    or os.environ.get('FLASK_ENV', '').lower() == 'production'
)
_IS_DEBUG = (os.environ.get('FLASK_DEBUG') == '1') or (not _IS_PRODUCTION and not _IS_TESTING)

_SECRET_KEY = os.environ.get('SECRET_KEY')
if not _SECRET_KEY:
    if _IS_PRODUCTION:
        raise RuntimeError(
            'SECRET_KEY environment variable is required when PRODUCTION=1. '
            'Set SECRET_KEY in .env or the process environment before starting.'
        )
    _SECRET_KEY = 'dev-only-insecure-key'
    print(
        '[WARN] SECRET_KEY not set — using insecure dev default. '
        'Set SECRET_KEY in .env for any real deployment.',
        flush=True,
    )
app.config['SECRET_KEY'] = _SECRET_KEY
app.config['TEMPLATES_AUTO_RELOAD'] = _IS_DEBUG
app.config['SQLALCHEMY_DATABASE_URI'] = os.environ.get('SQLALCHEMY_DATABASE_URI', 'sqlite:///classroom_app.db')
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db = SQLAlchemy(app)
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'

# Socket.IO CORS: env list wins, otherwise permissive in dev, strict (same-origin) in prod.
_raw_origins = (os.environ.get('SOCKETIO_ALLOWED_ORIGINS') or '').strip()
if _raw_origins:
    _cors_allowed = [o.strip() for o in _raw_origins.split(',') if o.strip()]
elif _IS_PRODUCTION:
    _cors_allowed = []
else:
    _cors_allowed = '*'
socketio = SocketIO(app, cors_allowed_origins=_cors_allowed, async_mode="threading")

STUDENT_TOKEN_SALT = 'student-session'
STUDENT_TOKEN_MAX_AGE = 7 * 24 * 60 * 60  # 7 days

PASSWORD_RESET_SALT = 'prof-password-reset'
PASSWORD_RESET_MAX_AGE = 60 * 60  # 1 hour


def _student_token_serializer():
    return URLSafeTimedSerializer(app.secret_key, salt=STUDENT_TOKEN_SALT)


def issue_student_token(student_id, needs_password=False):
    return _student_token_serializer().dumps({'id': int(student_id), 'np': bool(needs_password)})


def verify_student_token(token, max_age=STUDENT_TOKEN_MAX_AGE):
    if not token or not isinstance(token, str):
        return None, None
    try:
        data = _student_token_serializer().loads(token, max_age=max_age)
        sid = data.get('id')
        if sid is None:
            return None, None
        return int(sid), bool(data.get('np', False))
    except (BadSignature, SignatureExpired, TypeError, ValueError, KeyError):
        return None, None


def _bearer_token_from_request():
    auth = request.headers.get('Authorization', '') or ''
    parts = auth.split(None, 1)
    if len(parts) == 2 and parts[0].lower() == 'bearer':
        t = parts[1].strip()
        return t or None
    return None


def _authenticated_student_id():
    token = _bearer_token_from_request()
    student_id, _ = verify_student_token(token)
    return student_id


def _student_id_from_socket_token(data):
    if not isinstance(data, dict):
        return None
    token = data.get('token')
    student_id, _ = verify_student_token(token)
    return student_id


def _looks_like_email(s):
    if not s or not isinstance(s, str):
        return False
    s = s.strip()
    if len(s) > 254 or '@' not in s:
        return False
    return bool(re.match(r'^[^@\s]+@[^@\s]+\.[^@\s]+$', s))


def _password_reset_serializer():
    return URLSafeTimedSerializer(app.secret_key, salt=PASSWORD_RESET_SALT)


def make_professor_password_reset_token(professor_id):
    return _password_reset_serializer().dumps({'id': int(professor_id)})


def verify_professor_password_reset_token(token, max_age=PASSWORD_RESET_MAX_AGE):
    if not token or not isinstance(token, str):
        return None
    try:
        data = _password_reset_serializer().loads(token, max_age=max_age)
        pid = data.get('id')
        if pid is None:
            return None
        return int(pid)
    except (BadSignature, SignatureExpired, TypeError, ValueError, KeyError):
        return None


def _send_professor_password_reset_email(professor, reset_url, delivery_email=None):
    """Send password reset instructions. PASSWORD_RESET_EMAIL_TO overrides recipient when set."""
    override = (os.environ.get('PASSWORD_RESET_EMAIL_TO') or '').strip()
    if override:
        to_addr = override
    else:
        cand = (delivery_email or '').strip()
        to_addr = cand if cand else (professor.email or '').strip()
    if not to_addr:
        app.logger.error('[password reset] No recipient email for professor id=%s', professor.id)
        return False
    subject = 'Comet — Password reset request'
    body = (
        f"A password reset was requested for this Comet professor account:\n\n"
        f"  Username: {professor.username}\n"
        f"  Email on file: {professor.email}\n"
        f"  Sending this message to: {to_addr}\n\n"
        f"Use this link to set a new password (expires in 1 hour):\n{reset_url}\n\n"
        f"If you did not request this, you can ignore this message.\n"
    )
    mail_server = (os.environ.get('MAIL_SERVER') or '').strip()
    from_addr = (os.environ.get('MAIL_FROM') or os.environ.get('MAIL_USERNAME') or 'noreply@localhost').strip()

    if not mail_server:
        app.logger.warning(
            '[password reset] No MAIL_SERVER in .env — email is NOT sent. '
            'Link for %s (intended recipient %s) written to password_reset_last_link.txt',
            professor.username,
            to_addr,
        )
        print(
            f"\n[password reset] NO EMAIL SENT (configure MAIL_SERVER in .env). "
            f"Reset link for {professor.username} (would go to {to_addr}):\n{reset_url}\n",
            flush=True,
        )
        try:
            out_path = os.path.join(_APP_DIR, 'password_reset_last_link.txt')
            with open(out_path, 'w', encoding='utf-8') as f:
                f.write(
                    'Password reset (no SMTP configured — copy link below)\n\n'
                    f'To: {to_addr}\n'
                    f'Account: {professor.username}\n\n'
                    f'{reset_url}\n'
                )
            app.logger.info('[password reset] Link saved to %s', out_path)
        except OSError as e:
            app.logger.warning('Could not write password_reset_last_link.txt: %s', e)
        return True

    port = int(os.environ.get('MAIL_PORT', '587'))
    user = (os.environ.get('MAIL_USERNAME') or '').strip()
    password = (os.environ.get('MAIL_PASSWORD') or '').strip()
    msg = MIMEText(body, 'plain', 'utf-8')
    msg['Subject'] = subject
    msg['From'] = from_addr
    msg['To'] = to_addr

    try:
        if port == 465:
            context = ssl.create_default_context()
            with smtplib.SMTP_SSL(mail_server, port, context=context) as smtp:
                if user:
                    smtp.login(user, password)
                smtp.sendmail(from_addr, [to_addr], msg.as_string())
        else:
            context = ssl.create_default_context()
            with smtplib.SMTP(mail_server, port) as smtp:
                smtp.ehlo()
                smtp.starttls(context=context)
                smtp.ehlo()
                if user:
                    smtp.login(user, password)
                smtp.sendmail(from_addr, [to_addr], msg.as_string())
        app.logger.info('[password reset] Email sent via %s to %s', mail_server, to_addr)
        return True
    except Exception as e:
        app.logger.exception('Failed to send password reset email: %s', e)
        print(f"\n[password reset] Email send failed; reset link for {professor.username}:\n{reset_url}\n", flush=True)
        try:
            out_path = os.path.join(_APP_DIR, 'password_reset_last_link.txt')
            with open(out_path, 'w', encoding='utf-8') as f:
                f.write(
                    'Password reset (SMTP failed — use link below)\n\n'
                    f'To: {to_addr}\n'
                    f'Error: {e!s}\n\n'
                    f'{reset_url}\n'
                )
        except OSError:
            pass
        return False


# Database Models
class Professor(UserMixin, db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False)
    email = db.Column(db.String(120), unique=True, nullable=False)
    password_hash = db.Column(db.String(255), nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

class Student(UserMixin, db.Model):
    id = db.Column(db.Integer, primary_key=True)
    student_number = db.Column(db.String(50), unique=True, nullable=False)
    first_name = db.Column(db.String(100), nullable=False)
    preferred_name = db.Column(db.String(100), nullable=True)
    last_name = db.Column(db.String(100), nullable=False)
    email = db.Column(db.String(120), nullable=False)
    rfid_card_id = db.Column(db.String(100), unique=True, nullable=True)
    password_hash = db.Column(db.String(255), nullable=True)  # Nullable for first-time setup
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

class Class(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    professor_id = db.Column(db.Integer, db.ForeignKey('professor.id'), nullable=False)
    name = db.Column(db.String(200), nullable=False)
    class_code = db.Column(db.String(20), unique=True, nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    is_active = db.Column(db.Boolean, default=False)
    professor = db.relationship('Professor', backref=db.backref('classes', lazy=True))

class Enrollment(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    class_id = db.Column(db.Integer, db.ForeignKey('class.id'), nullable=False)
    student_id = db.Column(db.Integer, db.ForeignKey('student.id'), nullable=False)
    enrolled_at = db.Column(db.DateTime, default=datetime.utcnow)
    is_active = db.Column(db.Boolean, default=True, nullable=False)
    class_obj = db.relationship('Class', backref=db.backref('enrollments', lazy=True))
    student = db.relationship('Student', backref=db.backref('enrollments', lazy=True))

class Attendance(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    class_id = db.Column(db.Integer, db.ForeignKey('class.id'), nullable=False)
    student_id = db.Column(db.Integer, db.ForeignKey('student.id'), nullable=False)
    class_session_id = db.Column(db.Integer, db.ForeignKey('class_session.id'), nullable=True)
    date = db.Column(db.Date, nullable=False)
    present = db.Column(db.Boolean, default=True)
    join_time = db.Column(db.DateTime, nullable=True)  # Time when student joined
    leave_time = db.Column(db.DateTime, nullable=True)  # Time when student left (early logout or class end)
    timestamp = db.Column(db.DateTime, default=datetime.utcnow)
    class_obj = db.relationship('Class', backref=db.backref('attendances', lazy=True))
    student = db.relationship('Student', backref=db.backref('attendances', lazy=True))
    class_session = db.relationship('ClassSession', backref=db.backref('attendances', lazy=True))

class Participation(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    class_id = db.Column(db.Integer, db.ForeignKey('class.id'), nullable=False)
    student_id = db.Column(db.Integer, db.ForeignKey('student.id'), nullable=False)
    date = db.Column(db.Date, nullable=False)
    peer_grade = db.Column(db.Float, default=0.0)
    instructor_grade = db.Column(db.Float, default=0.0)
    hand_raises = db.Column(db.Integer, default=0)
    thumbs_up = db.Column(db.Integer, default=0)
    thumbs_down = db.Column(db.Integer, default=0)
    class_obj = db.relationship('Class', backref=db.backref('participations', lazy=True))
    student = db.relationship('Student', backref=db.backref('participations', lazy=True))

class Poll(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    class_id = db.Column(db.Integer, db.ForeignKey('class.id'), nullable=False)
    question = db.Column(db.String(500), nullable=False)
    options = db.Column(db.Text, nullable=False)  # JSON string
    correct_answer = db.Column(db.Integer, nullable=True)
    is_graded = db.Column(db.Boolean, default=False)  # Whether this poll counts toward grade
    is_anonymous = db.Column(db.Boolean, default=False)
    show_results_when_stopped = db.Column(
        db.Boolean, default=True, nullable=False
    )  # If True, students see aggregate results (and correct answer if set) when poll ends
    is_active = db.Column(db.Boolean, default=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    class_obj = db.relationship('Class', backref=db.backref('polls', lazy=True))


class PollBankSet(db.Model):
    """Named grouping for pre-uploaded poll-bank questions."""
    __tablename__ = 'poll_bank_set'
    __table_args__ = (UniqueConstraint('class_id', 'name', name='uq_pollbankset_class_name'),)
    id = db.Column(db.Integer, primary_key=True)
    class_id = db.Column(db.Integer, db.ForeignKey('class.id'), nullable=False)
    name = db.Column(db.String(120), nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow, nullable=False)
    class_obj = db.relationship('Class', backref=db.backref('poll_bank_sets', lazy=True))


class PollBankQuestion(db.Model):
    """Reusable pre-uploaded poll question (not a live poll row)."""
    __tablename__ = 'poll_bank_question'
    __table_args__ = (UniqueConstraint('set_id', 'poll_index', name='uq_pollbank_set_index'),)
    id = db.Column(db.Integer, primary_key=True)
    class_id = db.Column(db.Integer, db.ForeignKey('class.id'), nullable=False)
    set_id = db.Column(db.Integer, db.ForeignKey('poll_bank_set.id'), nullable=False)
    poll_index = db.Column(db.Integer, nullable=False)
    title = db.Column(db.String(200), nullable=False, default='Poll')
    question = db.Column(db.String(500), nullable=False)
    options = db.Column(db.Text, nullable=False)  # JSON string
    correct_answer = db.Column(db.Integer, nullable=True)
    source_filename = db.Column(db.String(255), nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow, nullable=False)
    class_obj = db.relationship('Class', backref=db.backref('poll_bank_questions', lazy=True))
    poll_set = db.relationship('PollBankSet', backref=db.backref('questions', lazy=True))


class PollResponse(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    poll_id = db.Column(db.Integer, db.ForeignKey('poll.id'), nullable=False)
    student_id = db.Column(db.Integer, db.ForeignKey('student.id'), nullable=False)
    answer = db.Column(db.Integer, nullable=False)
    is_correct = db.Column(db.Boolean, default=False)
    timestamp = db.Column(db.DateTime, default=datetime.utcnow)
    poll = db.relationship('Poll', backref=db.backref('responses', lazy=True))
    student = db.relationship('Student', backref=db.backref('poll_responses', lazy=True))

class ClassSettings(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    class_id = db.Column(db.Integer, db.ForeignKey('class.id'), nullable=False, unique=True)
    show_first_name_only = db.Column(db.Boolean, default=False)
    quiet_mode = db.Column(db.Boolean, default=False)
    class_obj = db.relationship('Class', backref=db.backref('settings', uselist=False))

class ProfessorPreferences(db.Model):
    """Global default preferences per professor, used as defaults when creating new classes."""
    id = db.Column(db.Integer, primary_key=True)
    professor_id = db.Column(db.Integer, db.ForeignKey('professor.id'), nullable=False, unique=True)
    default_show_first_name_only = db.Column(db.Boolean, default=False)
    default_quiet_mode = db.Column(db.Boolean, default=False)
    dark_mode = db.Column(db.Boolean, default=False)
    professor = db.relationship('Professor', backref=db.backref('preferences', uselist=False))

class ClassSession(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    class_id = db.Column(db.Integer, db.ForeignKey('class.id'), nullable=False)
    start_time = db.Column(db.DateTime, nullable=False, default=datetime.utcnow)
    end_time = db.Column(db.DateTime, nullable=True)
    exclude_from_grading = db.Column(db.Boolean, default=False)  # If True, this session doesn't count toward attendance grades
    class_obj = db.relationship('Class', backref=db.backref('sessions', lazy=True))


class AbsenceExemption(db.Model):
    """Instructor marks a student excused for a session; that session is omitted from their attendance average."""
    __tablename__ = 'absence_exemption'
    __table_args__ = (
        UniqueConstraint('class_session_id', 'student_id', name='uq_absence_exemption_session_student'),
    )
    id = db.Column(db.Integer, primary_key=True)
    class_id = db.Column(db.Integer, db.ForeignKey('class.id'), nullable=False)
    class_session_id = db.Column(db.Integer, db.ForeignKey('class_session.id'), nullable=False)
    student_id = db.Column(db.Integer, db.ForeignKey('student.id'), nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)
    class_obj = db.relationship('Class', backref=db.backref('absence_exemptions', lazy=True))
    class_session = db.relationship('ClassSession', backref=db.backref('absence_exemptions', lazy=True))
    student = db.relationship('Student', backref=db.backref('absence_exemptions', lazy=True))


class GradingWeights(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    class_id = db.Column(db.Integer, db.ForeignKey('class.id'), nullable=False, unique=True)
    attendance_weight = db.Column(db.Float, default=25.0, nullable=False)
    participation_weight = db.Column(db.Float, default=50.0, nullable=False)
    participation_instructor_share = db.Column(db.Float, default=50.0, nullable=False)
    poll_weight = db.Column(db.Float, default=25.0, nullable=False)
    quiz_weight = db.Column(db.Float, default=0.0, nullable=False)
    quiz_count_target = db.Column(db.Integer, default=0, nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    class_obj = db.relationship('Class', backref=db.backref('grading_weights', uselist=False))


class Quiz(db.Model):
    """In-class quiz definition (content); `quiz_index` maps to grading slot 1..N."""
    __tablename__ = 'quiz'
    __table_args__ = (UniqueConstraint('class_id', 'quiz_index', name='uq_quiz_class_index'),)
    id = db.Column(db.Integer, primary_key=True)
    class_id = db.Column(db.Integer, db.ForeignKey('class.id'), nullable=False)
    title = db.Column(db.String(200), nullable=False, default='Quiz')
    time_limit_seconds = db.Column(db.Integer, nullable=False, default=300)
    quiz_index = db.Column(db.Integer, nullable=False)
    source_filename = db.Column(db.String(255), nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)
    class_obj = db.relationship('Class', backref=db.backref('quizzes', lazy=True))


class QuizQuestion(db.Model):
    __tablename__ = 'quiz_question'
    id = db.Column(db.Integer, primary_key=True)
    quiz_id = db.Column(db.Integer, db.ForeignKey('quiz.id'), nullable=False)
    order = db.Column(db.Integer, nullable=False)
    prompt = db.Column(db.String(1000), nullable=False)
    options = db.Column(db.Text, nullable=False)  # JSON list of strings
    correct_index = db.Column(db.Integer, nullable=False)
    quiz = db.relationship('Quiz', backref=db.backref('questions', lazy=True, order_by='QuizQuestion.order'))


class QuizRun(db.Model):
    __tablename__ = 'quiz_run'
    id = db.Column(db.Integer, primary_key=True)
    quiz_id = db.Column(db.Integer, db.ForeignKey('quiz.id'), nullable=False)
    class_id = db.Column(db.Integer, db.ForeignKey('class.id'), nullable=False)
    started_at = db.Column(db.DateTime, nullable=False, default=datetime.utcnow)
    deadline_at = db.Column(db.DateTime, nullable=False)
    ended_at = db.Column(db.DateTime, nullable=True)
    is_active = db.Column(db.Boolean, default=False, nullable=False)
    quiz = db.relationship('Quiz', backref=db.backref('runs', lazy=True))
    class_obj = db.relationship('Class', backref=db.backref('quiz_runs', lazy=True))


class QuizAnswer(db.Model):
    __tablename__ = 'quiz_answer'
    __table_args__ = (
        UniqueConstraint('quiz_run_id', 'student_id', 'question_id', name='uq_quiz_answer_run_student_q'),
    )
    id = db.Column(db.Integer, primary_key=True)
    quiz_run_id = db.Column(db.Integer, db.ForeignKey('quiz_run.id'), nullable=False)
    student_id = db.Column(db.Integer, db.ForeignKey('student.id'), nullable=False)
    question_id = db.Column(db.Integer, db.ForeignKey('quiz_question.id'), nullable=False)
    selected_index = db.Column(db.Integer, nullable=False)
    is_correct = db.Column(db.Boolean, default=False, nullable=False)
    submitted_at = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)
    quiz_run = db.relationship('QuizRun', backref=db.backref('answers', lazy=True))
    student = db.relationship('Student', backref=db.backref('quiz_answers', lazy=True))
    question = db.relationship('QuizQuestion', backref=db.backref('answers', lazy=True))


class HandRaise(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    class_id = db.Column(db.Integer, db.ForeignKey('class.id'), nullable=False)
    student_id = db.Column(db.Integer, db.ForeignKey('student.id'), nullable=False)
    timestamp = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)
    cleared = db.Column(db.Boolean, default=False, nullable=False)
    class_obj = db.relationship('Class', backref=db.backref('hand_raises', lazy=True))
    student = db.relationship('Student', backref=db.backref('hand_raises', lazy=True))


class ParticipationGradeRound(db.Model):
    """Professor-triggered grading round for a subject student (raised-hand participation)."""
    __tablename__ = 'participation_grade_round'
    id = db.Column(db.Integer, primary_key=True)
    class_id = db.Column(db.Integer, db.ForeignKey('class.id'), nullable=False)
    subject_student_id = db.Column(db.Integer, db.ForeignKey('student.id'), nullable=False)
    date = db.Column(db.Date, nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)
    hand_raise_id = db.Column(db.Integer, db.ForeignKey('hand_raise.id'), nullable=True)
    exclude_from_grading = db.Column(db.Boolean, default=False, nullable=False)
    class_session_id = db.Column(db.Integer, db.ForeignKey('class_session.id'), nullable=True)
    class_obj = db.relationship('Class', backref=db.backref('participation_grade_rounds', lazy=True))
    class_session = db.relationship(
        'ClassSession',
        foreign_keys=[class_session_id],
        backref=db.backref('participation_grade_rounds', lazy=True),
    )
    subject = db.relationship(
        'Student',
        foreign_keys=[subject_student_id],
        backref=db.backref('participation_grade_rounds_as_subject', lazy=True),
    )


class InstructorParticipationGrade(db.Model):
    __tablename__ = 'instructor_participation_grade'
    id = db.Column(db.Integer, primary_key=True)
    round_id = db.Column(db.Integer, db.ForeignKey('participation_grade_round.id'), nullable=False, unique=True)
    score = db.Column(db.Integer, nullable=False)  # 1–100
    round = db.relationship('ParticipationGradeRound', backref=db.backref('instructor_grade', uselist=False))


class PeerParticipationGrade(db.Model):
    __tablename__ = 'peer_participation_grade'
    __table_args__ = (
        UniqueConstraint('round_id', 'grader_student_id', name='uq_peer_participation_round_grader'),
    )
    id = db.Column(db.Integer, primary_key=True)
    round_id = db.Column(db.Integer, db.ForeignKey('participation_grade_round.id'), nullable=False)
    grader_student_id = db.Column(db.Integer, db.ForeignKey('student.id'), nullable=False)
    rating = db.Column(db.Integer, nullable=False)  # 0–4
    score_percent = db.Column(db.Float, nullable=False)
    round = db.relationship('ParticipationGradeRound', backref=db.backref('peer_grades', lazy=True))
    grader = db.relationship(
        'Student',
        foreign_keys=[grader_student_id],
        backref=db.backref('peer_grades_given', lazy=True),
    )


# Composite / secondary indexes for hot filter paths.
#
# db.create_all() creates each of these with SQLite's default IF NOT EXISTS semantics
# on fresh databases. For existing deployments, migrate_database() re-runs create_all
# which is a no-op for indexes that already exist.
db.Index('ix_enrollment_class_active', Enrollment.class_id, Enrollment.is_active)
db.Index('ix_enrollment_student_active', Enrollment.student_id, Enrollment.is_active)
db.Index('ix_attendance_class_session', Attendance.class_id, Attendance.class_session_id)
db.Index('ix_attendance_class_student', Attendance.class_id, Attendance.student_id)
db.Index('ix_attendance_session_student', Attendance.class_session_id, Attendance.student_id)
db.Index('ix_attendance_class_date', Attendance.class_id, Attendance.date)
db.Index('ix_participation_class_date', Participation.class_id, Participation.date)
db.Index('ix_participation_class_student_date', Participation.class_id, Participation.student_id, Participation.date)
db.Index('ix_pollresponse_poll', PollResponse.poll_id)
db.Index('ix_pollresponse_poll_student', PollResponse.poll_id, PollResponse.student_id)
db.Index('ix_poll_class_active', Poll.class_id, Poll.is_active)
db.Index('ix_poll_class_created', Poll.class_id, Poll.created_at)
db.Index('ix_pollbank_set_class_name', PollBankSet.class_id, PollBankSet.name)
db.Index('ix_pollbank_class_set_index', PollBankQuestion.class_id, PollBankQuestion.set_id, PollBankQuestion.poll_index)
db.Index('ix_classsession_class_end', ClassSession.class_id, ClassSession.end_time)
db.Index('ix_classsession_class_start', ClassSession.class_id, ClassSession.start_time)
db.Index('ix_handraise_class_student_cleared', HandRaise.class_id, HandRaise.student_id, HandRaise.cleared)
db.Index('ix_handraise_class_timestamp', HandRaise.class_id, HandRaise.timestamp)
db.Index('ix_absence_class_student', AbsenceExemption.class_id, AbsenceExemption.student_id)
db.Index('ix_pgr_class_subject', ParticipationGradeRound.class_id, ParticipationGradeRound.subject_student_id)
db.Index('ix_pgr_class_date', ParticipationGradeRound.class_id, ParticipationGradeRound.date)
db.Index('ix_pgr_session', ParticipationGradeRound.class_session_id)
db.Index('ix_quiz_class_index', Quiz.class_id, Quiz.quiz_index)
db.Index('ix_quizrun_class_active', QuizRun.class_id, QuizRun.is_active)
db.Index('ix_quizrun_quiz_started', QuizRun.quiz_id, QuizRun.started_at)


def _peer_rating_to_percent(rating):
    m = {0: 0.0, 1: 25.0, 2: 50.0, 3: 75.0, 4: 100.0}
    if rating not in m:
        raise ValueError('invalid peer rating')
    return m[rating]


def _recompute_subject_participation_grades(class_id, subject_student_id, grade_date):
    """Set today's Participation.instructor_grade and .peer_grade from daily rounds (transparency / legacy).

    The course gradebook participation column uses course_mean_session_participation (per-session max scores),
    not these daily aggregates.
    """
    rounds = (
        ParticipationGradeRound.query.filter_by(
            class_id=class_id,
            subject_student_id=subject_student_id,
            date=grade_date,
        )
        .filter(ParticipationGradeRound.exclude_from_grading.is_(False))
        .all()
    )
    rids = [r.id for r in rounds]
    inst_val = 0.0
    peer_val = 0.0
    if rids:
        inst_avg = (
            db.session.query(func.avg(InstructorParticipationGrade.score))
            .filter(InstructorParticipationGrade.round_id.in_(rids))
            .scalar()
        )
        peer_avg = (
            db.session.query(func.avg(PeerParticipationGrade.score_percent))
            .filter(PeerParticipationGrade.round_id.in_(rids))
            .scalar()
        )
        if inst_avg is not None:
            inst_val = float(inst_avg)
        if peer_avg is not None:
            peer_val = float(peer_avg)

    participation = Participation.query.filter_by(
        class_id=class_id,
        student_id=subject_student_id,
        date=grade_date,
    ).first()
    if not participation:
        participation = Participation(
            class_id=class_id,
            student_id=subject_student_id,
            date=grade_date,
        )
        db.session.add(participation)
    participation.instructor_grade = inst_val
    participation.peer_grade = peer_val


def _subject_display_name_for_participation_grade(class_id, student):
    """Label for Socket.IO payload (aligned with first-name-only / full name display)."""
    settings = ClassSettings.query.filter_by(class_id=class_id).first()
    show_fn = bool(settings and settings.show_first_name_only)
    if show_fn:
        labels = _first_name_only_labels_for_class(class_id)
        return labels.get(
            student.id,
            (student.preferred_name or student.first_name or '').strip() or '?',
        )
    return f'{student.preferred_name or student.first_name} {student.last_name}'.strip()


def get_active_class_session(class_id):
    """Open ClassSession for this class, or None."""
    return ClassSession.query.filter_by(
        class_id=class_id,
        end_time=None,
    ).order_by(ClassSession.start_time.desc()).first()


def _first_name_only_labels_for_students(students):
    """
    For first-name-only display: map each student id -> label.
    If two or more students share the same visible short name (preferred_name, else first_name),
    disambiguate with first letter of last name: e.g. "Alex M."
    """
    from collections import defaultdict

    if not students:
        return {}
    key_to_students = defaultdict(list)
    for s in students:
        short = (s.preferred_name or s.first_name or '').strip()
        if not short:
            short = (s.first_name or '').strip()
        if not short:
            short = '?'
        key_to_students[short.lower()].append(s)
    out = {}
    for group in key_to_students.values():
        if len(group) == 1:
            s = group[0]
            short = (s.preferred_name or s.first_name or '').strip() or '?'
            out[s.id] = short
        else:
            for s in group:
                short = (s.preferred_name or s.first_name or '').strip() or '?'
                last = (s.last_name or '').strip()
                if last:
                    out[s.id] = f'{short} {last[0].upper()}.'
                else:
                    out[s.id] = short
    return out


def _first_name_only_labels_for_class(class_id):
    """Labels for all actively enrolled students in this class."""
    students = (
        db.session.query(Student)
        .join(Enrollment)
        .filter(
            Enrollment.class_id == class_id,
            Enrollment.is_active == True,
        )
        .all()
    )
    return _first_name_only_labels_for_students(students)


def count_graded_attendance_for_student(class_id, student_id):
    """
    Count how many graded (non-excluded) sessions the student attended.
    Session-scoped rows use class_session_id; legacy rows (NULL session) match by date
    only when there is a single graded session on that calendar day.
    Sessions with an absence exemption for this student are omitted from both numerator and denominator.
    Returns (attended_count, countable_graded_sessions).
    """
    sessions = ClassSession.query.filter_by(class_id=class_id).all()
    graded = [s for s in sessions if not s.exclude_from_grading]
    if not graded:
        return 0, 0
    exempt_ids = {
        row.class_session_id
        for row in AbsenceExemption.query.filter_by(class_id=class_id, student_id=student_id).all()
    }
    all_att = Attendance.query.filter_by(class_id=class_id, student_id=student_id).all()
    by_sid = {a.class_session_id: a for a in all_att if a.class_session_id is not None}
    legacy = [a for a in all_att if a.class_session_id is None]
    count = 0
    denom = 0
    for s in graded:
        if s.id in exempt_ids:
            continue
        denom += 1
        a = by_sid.get(s.id)
        if a is not None:
            if a.join_time is not None and a.present:
                count += 1
            continue
        same_date_graded = [x for x in graded if x.start_time.date() == s.start_time.date()]
        leg = next((a for a in legacy if a.date == s.start_time.date()), None)
        if leg and leg.join_time is not None and leg.present and len(same_date_graded) == 1:
            count += 1
    return count, denom


def gradebook_poll_responses_by_student(class_id):
    """Map student_id -> PollResponse rows that count toward the gradebook (graded + session window + not exclude_from_grading)."""
    sessions = ClassSession.query.filter_by(class_id=class_id).all()
    graded_windows = [s for s in sessions if not s.exclude_from_grading]
    now = datetime.utcnow()
    responses = (
        PollResponse.query.join(Poll)
        .filter(
            Poll.class_id == class_id,
            Poll.is_graded == True,
        )
        .all()
    )
    by_student = {}
    for pr in responses:
        poll = pr.poll
        t = poll.created_at
        matched = False
        for s in graded_windows:
            end = s.end_time or now
            if s.start_time <= t <= end:
                matched = True
                break
        if not matched:
            continue
        by_student.setdefault(pr.student_id, []).append(pr)
    return by_student


def deactivate_active_polls_for_class(class_id):
    """Set all active polls for this class to inactive. Caller must commit. Returns poll ids that were active."""
    active = Poll.query.filter_by(class_id=class_id, is_active=True).all()
    ids = [p.id for p in active]
    for p in active:
        p.is_active = False
    return ids


def deactivate_active_quiz_runs_for_class(class_id):
    """End all active quiz runs for this class. Caller must commit. Returns list of run ids that were active."""
    active = QuizRun.query.filter_by(class_id=class_id, is_active=True).all()
    ids = [r.id for r in active]
    now = datetime.utcnow()
    for r in active:
        r.is_active = False
        if r.ended_at is None:
            r.ended_at = now
    return ids


def _quiz_run_completed(run, now):
    if run.ended_at is not None:
        return True
    if run.deadline_at and now >= run.deadline_at:
        return True
    return False


def poll_option_counts(poll, responses=None):
    """Parse a Poll's option list and count responses per option.

    Pass `responses` when callers have already loaded the rows (avoids a second query).
    Returns (options_list, option_counts_dict, total_responses).
    """
    try:
        options = json.loads(poll.options) if poll.options else []
    except (json.JSONDecodeError, TypeError):
        options = []
    n = len(options)
    if responses is None:
        responses = PollResponse.query.filter_by(poll_id=poll.id).all()
    option_counts = {i: 0 for i in range(n)}
    for r in responses:
        try:
            ai = int(r.answer)
        except (TypeError, ValueError):
            continue
        if 0 <= ai < n:
            option_counts[ai] += 1
    total = sum(option_counts.values())
    return options, option_counts, total


def poll_results_payload(poll_id):
    """Aggregate counts for Socket.IO clients when a poll ends (student results flash)."""
    poll = Poll.query.get(poll_id)
    if not poll:
        return None
    if not bool(poll.show_results_when_stopped):
        return None
    options, option_counts, total = poll_option_counts(poll)
    counts = [option_counts[i] for i in range(len(options))]
    return {
        'question': poll.question,
        'options': options,
        'counts': counts,
        'total_responses': total,
        'is_graded': bool(poll.is_graded),
        'correct_answer': poll.correct_answer,
        'is_anonymous': bool(poll.is_anonymous),
    }


def emit_poll_stopped_events(class_id, poll_ids):
    for pid in poll_ids:
        socketio.emit(
            'poll_stopped',
            {
                'poll_id': pid,
                'class_id': class_id,
                'results': poll_results_payload(pid),
            },
            room=f'class_{class_id}',
        )


def _delete_quiz_and_related(quiz_obj):
    """Remove quiz runs, answers, questions, and the quiz row."""
    runs = QuizRun.query.filter_by(quiz_id=quiz_obj.id).all()
    for run in runs:
        QuizAnswer.query.filter_by(quiz_run_id=run.id).delete()
    QuizRun.query.filter_by(quiz_id=quiz_obj.id).delete()
    QuizQuestion.query.filter_by(quiz_id=quiz_obj.id).delete()
    db.session.delete(quiz_obj)


def _quiz_questions_public_list(quiz):
    """Questions for students (no correct answer)."""
    out = []
    for q in sorted(quiz.questions, key=lambda x: x.order):
        try:
            opts = json.loads(q.options or '[]')
        except (json.JSONDecodeError, TypeError):
            opts = []
        out.append({'id': q.id, 'prompt': q.prompt, 'options': opts})
    return out


QUIZ_TEMPLATE_XLSX = os.path.join(_APP_DIR, 'static', 'quiz_template.xlsx')


def _detect_quiz_sheet_format(headers):
    """Return 'letter' (Option A/B/…) + Question Description layout, else 'legacy' (Option 1…)."""
    if not headers:
        return 'legacy'
    hl = []
    for h in headers:
        hl.append(str(h).strip().lower() if h is not None else '')
    if not any(h and 'question description' in h for h in hl):
        return 'legacy'
    for h in hl:
        if h and re.match(r'^option\s+a\s*$', h.strip()):
            return 'letter'
    return 'legacy'


def _parse_quiz_rows_letter_headers(ws, headers):
    """Parse workbook with columns: Question Description, # of Options, Option A…, Correct Answer."""
    questions_to_add = []
    errors = []

    def col_question_desc():
        for i, h in enumerate(headers):
            if h and 'question description' in str(h).strip().lower():
                return i
        return None

    def col_num_options():
        for i, h in enumerate(headers):
            if not h:
                continue
            s = str(h).strip().lower()
            if '# of options' in s or s.replace(' ', '') in ('numberofoptions', '#ofoptions'):
                return i
            if s == 'number of options':
                return i
        return None

    def col_correct():
        for i, h in enumerate(headers):
            if not h:
                continue
            if 'correct' in str(h).strip().lower():
                return i
        return None

    idx_desc = col_question_desc()
    idx_n = col_num_options()
    idx_correct = col_correct()
    if idx_desc is None or idx_n is None or idx_correct is None:
        return [], [
            'Missing columns: need "Question Description", "# of Options", and "Correct Answer" '
            '(see the downloadable Quiz Template).',
        ]
    option_indices = list(range(idx_n + 1, idx_correct))
    if len(option_indices) < 2:
        return [], ['Template needs at least two option columns (Option A, Option B, …) before Correct Answer.']

    max_slots = len(option_indices)

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        row = list(row)
        need = max(idx_desc, idx_n, idx_correct) + 1
        while len(row) < need:
            row.append(None)
        key_cells = (row[idx_desc], row[idx_n])
        if all(
            c is None or (isinstance(c, str) and not str(c).strip())
            for c in key_cells
        ):
            continue
        try:
            prompt = str(row[idx_desc]).strip() if row[idx_desc] is not None else ''
            n_opts = int(row[idx_n])
        except (TypeError, ValueError, IndexError):
            errors.append(f'Row {row_idx}: Invalid question text or # of options.')
            continue
        if not prompt:
            errors.append(f'Row {row_idx}: Question description is required.')
            continue
        if n_opts < 2 or n_opts > 8:
            errors.append(f'Row {row_idx}: # of options must be between 2 and 8.')
            continue
        if n_opts > max_slots:
            errors.append(
                f'Row {row_idx}: # of options ({n_opts}) exceeds the number of option columns in the sheet ({max_slots}).'
            )
            continue
        opts = []
        ok = True
        for j in range(n_opts):
            oi = option_indices[j]
            cell = row[oi] if oi < len(row) else None
            t = str(cell).strip() if cell is not None else ''
            if not t:
                errors.append(f'Row {row_idx}: Option {j + 1} is empty.')
                ok = False
                break
            opts.append(t)
        if not ok:
            continue
        correct_cell = row[idx_correct] if idx_correct < len(row) else None
        correct_one_based = _parse_quiz_excel_correct_answer(correct_cell, n_opts)
        if correct_one_based is None:
            letter_hi = chr(ord('A') + n_opts - 1)
            errors.append(
                f'Row {row_idx}: Correct answer must be option number 1–{n_opts} '
                f'or letter A–{letter_hi} (e.g. 3 or C for the third option).'
            )
            continue
        questions_to_add.append(
            {
                'order': len(questions_to_add) + 1,
                'prompt': prompt,
                'options': opts,
                'correct_index': correct_one_based - 1,
            }
        )

    return questions_to_add, errors


def _parse_quiz_rows_legacy_fixed_columns(ws):
    """Original template: Question text (col B), Number of options, Option 1…N, then correct (next col)."""
    questions_to_add = []
    errors = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or all(
            (c is None or (isinstance(c, str) and not str(c).strip())) for c in (row[:3] if row else [])
        ):
            continue
        try:
            prompt = str(row[1]).strip() if row[1] is not None else ''
            n_opts = int(row[2])
        except (TypeError, ValueError, IndexError):
            errors.append(f'Row {row_idx}: Invalid question text or number of options.')
            continue
        if not prompt:
            errors.append(f'Row {row_idx}: Question text is required.')
            continue
        if n_opts < 2 or n_opts > 8:
            errors.append(f'Row {row_idx}: Number of options must be between 2 and 8.')
            continue
        need_len = 3 + n_opts + 1
        if not row or len(row) < need_len:
            errors.append(f'Row {row_idx}: Row too short for {n_opts} options and correct answer column.')
            continue
        opts = []
        ok = True
        for j in range(n_opts):
            cell = row[3 + j]
            t = str(cell).strip() if cell is not None else ''
            if not t:
                errors.append(f'Row {row_idx}: Option {j + 1} is empty.')
                ok = False
                break
            opts.append(t)
        if not ok:
            continue
        correct_cell = row[3 + n_opts]
        correct_one_based = _parse_quiz_excel_correct_answer(correct_cell, n_opts)
        if correct_one_based is None:
            letter_hi = chr(ord('A') + n_opts - 1)
            errors.append(
                f'Row {row_idx}: Correct answer must be option number 1–{n_opts} '
                f'or letter A–{letter_hi} (e.g. 3 or C for the third option).'
            )
            continue
        questions_to_add.append(
            {
                'order': len(questions_to_add) + 1,
                'prompt': prompt,
                'options': opts,
                'correct_index': correct_one_based - 1,
            }
        )
    return questions_to_add, errors


def _parse_quiz_workbook(ws):
    first = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not first:
        return [], ['The spreadsheet must have a header row in row 1.']
    headers = list(first)
    if _detect_quiz_sheet_format(headers) == 'letter':
        return _parse_quiz_rows_letter_headers(ws, headers)
    return _parse_quiz_rows_legacy_fixed_columns(ws)


def _parse_poll_bank_workbook(ws):
    """Parse poll-bank workbook rows using quiz-compatible sheet headers."""
    questions, errors = _parse_quiz_workbook(ws)
    out = []
    for i, qd in enumerate(questions):
        idx = qd.get('order')
        if idx is None:
            idx = i + 1
        try:
            idx = int(idx)
        except (TypeError, ValueError):
            idx = i + 1
        if idx < 1:
            idx = i + 1
        out.append({
            'poll_index': idx,
            'question': qd.get('prompt', '').strip(),
            'options': qd.get('options') or [],
            'correct_answer': qd.get('correct_index'),
        })
    return out, errors


def _parse_quiz_excel_correct_answer(cell, n_opts):
    """
    Map an Excel cell to a 1-based correct option index, or None if invalid.

    Accepts integers or whole floats (1..n_opts), numeric strings, or a single
    letter A..Z mapping to 1..26 (only positions 1..n_opts are valid).
    """
    if cell is None:
        return None
    if isinstance(cell, bool):
        return None
    if isinstance(cell, (int, float)):
        if isinstance(cell, float):
            if cell != cell or abs(cell - round(cell)) > 1e-9:  # NaN or non-whole
                return None
            v = int(round(cell))
        else:
            v = int(cell)
        if 1 <= v <= n_opts:
            return v
        return None
    s = str(cell).strip()
    if not s:
        return None
    s_clean = s.rstrip('.)').strip()
    if len(s_clean) == 1:
        ch = s_clean.upper()
        if 'A' <= ch <= 'Z':
            pos = ord(ch) - ord('A') + 1
            if 1 <= pos <= n_opts:
                return pos
            return None
    try:
        v = int(round(float(s_clean)))
    except (TypeError, ValueError):
        return None
    if 1 <= v <= n_opts:
        return v
    return None


def _isoformat_utc_for_js(dt):
    """Emit an instant JavaScript can parse as UTC. Naive datetimes in this app are stored as UTC."""
    if dt is None:
        return None
    if dt.tzinfo is not None:
        return dt.astimezone(timezone.utc).isoformat().replace('+00:00', 'Z')
    return dt.isoformat() + 'Z'


def quiz_run_public_payload(quiz_run):
    """Payload for Socket.IO `quiz_started` and student `active_quiz` (no solutions)."""
    quiz = quiz_run.quiz
    return {
        'quiz_run_id': quiz_run.id,
        'class_id': quiz_run.class_id,
        'quiz_id': quiz.id,
        'title': quiz.title,
        'time_limit_seconds': quiz.time_limit_seconds,
        'started_at': _isoformat_utc_for_js(quiz_run.started_at),
        'deadline_at': _isoformat_utc_for_js(quiz_run.deadline_at),
        'questions': _quiz_questions_public_list(quiz),
    }


def effective_attendance_and_poll_weights(class_id, grading_weights):
    """
    Redistribute poll_weight toward attendance for graded sessions with no poll in the session window.
    Returns (effective_attendance_weight, effective_poll_weight).
    """
    sessions = ClassSession.query.filter_by(class_id=class_id).all()
    graded = [s for s in sessions if not s.exclude_from_grading]
    n = len(graded)
    aw = float(grading_weights.attendance_weight)
    pw = float(grading_weights.poll_weight)
    if n == 0:
        return aw, pw
    now = datetime.utcnow()
    sessions_with_poll = 0
    for s in graded:
        end = s.end_time or now
        if (
            Poll.query.filter(
                Poll.class_id == class_id,
                Poll.created_at >= s.start_time,
                Poll.created_at <= end,
            ).first()
            is not None
        ):
            sessions_with_poll += 1
    if sessions_with_poll == 0:
        return aw + pw, 0.0
    eff_poll = pw * (sessions_with_poll / n)
    eff_att = aw + pw * ((n - sessions_with_poll) / n)
    return eff_att, eff_poll


def blended_participation_grade(avg_instructor_grade, avg_peer_grade, participation_instructor_share):
    """0–100 blended score inside the participation bucket (peer share = 100 − instructor share)."""
    si = float(participation_instructor_share) / 100.0
    sp = 1.0 - si
    return float(avg_instructor_grade) * si + float(avg_peer_grade) * sp


def student_attended_class_session(class_id, session, student_id):
    """Whether the student counts as present for this session (aligned with count_graded_attendance_for_student)."""
    by_sid = Attendance.query.filter_by(
        class_id=class_id,
        student_id=student_id,
        class_session_id=session.id,
    ).first()
    if by_sid is not None:
        return bool(by_sid.join_time is not None and by_sid.present)
    sessions = ClassSession.query.filter_by(class_id=class_id).all()
    graded = [s for s in sessions if not s.exclude_from_grading]
    same_date_graded = [x for x in graded if x.start_time.date() == session.start_time.date()]
    if len(same_date_graded) != 1:
        return False
    leg = Attendance.query.filter_by(
        class_id=class_id,
        student_id=student_id,
        class_session_id=None,
        date=session.start_time.date(),
    ).first()
    return bool(leg and leg.join_time is not None and leg.present)


def session_participation_score(class_id, session, student_id, participation_instructor_share):
    """
    Single-session participation 0–100: max(25 if attended, 40 if hand raised in window,
    blended scores for each graded round in session). None if student did not attend this session
    or session is excluded from grading.
    """
    if session.exclude_from_grading:
        return None
    if not student_attended_class_session(class_id, session, student_id):
        return None

    now = datetime.utcnow()
    session_start = session.start_time
    session_end = session.end_time if session.end_time else now

    candidates = [25.0]

    hr_exists = (
        HandRaise.query.filter(
            HandRaise.class_id == class_id,
            HandRaise.student_id == student_id,
            HandRaise.timestamp >= session_start,
            HandRaise.timestamp <= session_end,
        ).first()
        is not None
    )
    if hr_exists:
        candidates.append(40.0)

    rounds = ParticipationGradeRound.query.filter_by(
        class_id=class_id,
        subject_student_id=student_id,
    ).filter(ParticipationGradeRound.exclude_from_grading.is_(False)).all()

    for rnd in rounds:
        in_session = False
        if rnd.class_session_id == session.id:
            in_session = True
        elif rnd.class_session_id is None and rnd.created_at:
            if session_start <= rnd.created_at <= session_end:
                in_session = True
        if not in_session:
            continue
        inst = InstructorParticipationGrade.query.filter_by(round_id=rnd.id).first()
        if not inst:
            continue
        peer_avg = (
            db.session.query(func.avg(PeerParticipationGrade.score_percent))
            .filter(PeerParticipationGrade.round_id == rnd.id)
            .scalar()
        )
        peer_val = float(peer_avg) if peer_avg is not None else 0.0
        blended = blended_participation_grade(
            float(inst.score), peer_val, participation_instructor_share
        )
        candidates.append(blended)

    return max(candidates)


def course_mean_session_participation(class_id, student_id, participation_instructor_share):
    """Mean of per-session participation scores over graded sessions the student attended."""
    sessions = ClassSession.query.filter_by(class_id=class_id).all()
    graded = [s for s in sessions if not s.exclude_from_grading]
    scores = []
    for s in graded:
        sc = session_participation_score(class_id, s, student_id, participation_instructor_share)
        if sc is not None:
            scores.append(sc)
    return sum(scores) / len(scores) if scores else 0.0


def _compute_gradebook_rows(class_id, grading_weights):
    """Per-class bulk gradebook compute used by both `get_gradebook` and `export_gradebook`.

    Preserves the exact math of `count_graded_attendance_for_student`,
    `session_participation_score`, `course_mean_session_participation`,
    `effective_attendance_and_poll_weights`, and `gradebook_poll_responses_by_student`
    while replacing their per-student / per-session / per-round query fan-out with
    a fixed, small number of batched queries.

    Returns a list of dicts keyed by `student`, `attendance_grade`, `participation_grade`,
    `avg_peer_grade`, `avg_instructor_grade`, `poll_grade`, `quiz_grade`, `quiz_scores_by_index`,
    `overall_grade`.
    """
    inst_share = float(grading_weights.participation_instructor_share)
    part_weight = float(grading_weights.participation_weight)
    aw = float(grading_weights.attendance_weight)
    pw = float(grading_weights.poll_weight)
    qw = float(getattr(grading_weights, 'quiz_weight', 0.0) or 0.0)
    quiz_n_target = int(getattr(grading_weights, 'quiz_count_target', 0) or 0)
    now = datetime.utcnow()

    students = (
        db.session.query(Student)
        .join(Enrollment)
        .filter(Enrollment.class_id == class_id, Enrollment.is_active == True)
        .all()
    )

    all_sessions = ClassSession.query.filter_by(class_id=class_id).all()
    graded_sessions = [s for s in all_sessions if not s.exclude_from_grading]
    graded_by_date_count = {}
    for s in graded_sessions:
        d = s.start_time.date()
        graded_by_date_count[d] = graded_by_date_count.get(d, 0) + 1

    attend_by_sid = {}
    legacy_by_date = {}
    for a in Attendance.query.filter_by(class_id=class_id).all():
        if a.class_session_id is not None:
            attend_by_sid[(a.student_id, a.class_session_id)] = a
        else:
            legacy_by_date[(a.student_id, a.date)] = a

    exempt_map = {}
    for row in AbsenceExemption.query.filter_by(class_id=class_id).all():
        exempt_map.setdefault(row.student_id, set()).add(row.class_session_id)

    parts_by_student = {}
    for p in Participation.query.filter_by(class_id=class_id).all():
        parts_by_student.setdefault(p.student_id, []).append(p)

    hr_by_student = {}
    for hr in HandRaise.query.filter_by(class_id=class_id).all():
        hr_by_student.setdefault(hr.student_id, []).append(hr.timestamp)

    rounds = (
        ParticipationGradeRound.query.filter_by(class_id=class_id)
        .filter(ParticipationGradeRound.exclude_from_grading.is_(False))
        .all()
    )
    rounds_by_subject = {}
    for rnd in rounds:
        rounds_by_subject.setdefault(rnd.subject_student_id, []).append(rnd)

    round_ids = [r.id for r in rounds]
    instr_by_round = {}
    peer_avg_by_round = {}
    if round_ids:
        for ig in InstructorParticipationGrade.query.filter(
            InstructorParticipationGrade.round_id.in_(round_ids)
        ).all():
            instr_by_round[ig.round_id] = float(ig.score)
        for rid, avg_val in (
            db.session.query(
                PeerParticipationGrade.round_id,
                func.avg(PeerParticipationGrade.score_percent),
            )
            .filter(PeerParticipationGrade.round_id.in_(round_ids))
            .group_by(PeerParticipationGrade.round_id)
            .all()
        ):
            peer_avg_by_round[rid] = float(avg_val) if avg_val is not None else 0.0

    polls = Poll.query.filter_by(class_id=class_id).all()
    sessions_with_poll = 0
    for s in graded_sessions:
        end = s.end_time or now
        for poll in polls:
            t = poll.created_at
            if t is not None and s.start_time <= t <= end:
                sessions_with_poll += 1
                break

    class_quiz_runs = QuizRun.query.filter_by(class_id=class_id).all()
    sessions_with_quiz = 0
    for s in graded_sessions:
        end = s.end_time or now
        for run in class_quiz_runs:
            t = run.started_at
            if t is not None and s.start_time <= t <= end:
                sessions_with_quiz += 1
                break

    n_graded = len(graded_sessions)
    if n_graded == 0:
        eff_att_w, eff_poll_w, eff_quiz_w = aw, pw, qw
    else:
        sp = float(sessions_with_poll)
        sq = float(sessions_with_quiz)
        eff_poll_w = pw * (sp / n_graded)
        eff_quiz_w = qw * (sq / n_graded)
        eff_att_w = aw + pw * ((n_graded - sp) / n_graded) + qw * ((n_graded - sq) / n_graded)

    graded_poll_ids = [p.id for p in polls if p.is_graded]
    poll_in_graded_window = {}
    poll_by_id = {p.id: p for p in polls}
    for pid in graded_poll_ids:
        t = poll_by_id[pid].created_at
        inside = False
        if t is not None:
            for s in graded_sessions:
                end = s.end_time or now
                if s.start_time <= t <= end:
                    inside = True
                    break
        poll_in_graded_window[pid] = inside

    poll_map = {}
    if graded_poll_ids:
        for pr in PollResponse.query.filter(PollResponse.poll_id.in_(graded_poll_ids)).all():
            if poll_in_graded_window.get(pr.poll_id):
                poll_map.setdefault(pr.student_id, []).append(pr)

    def _run_in_graded_session_window(run):
        t = run.started_at
        if t is None:
            return False
        for s in graded_sessions:
            end = s.end_time or now
            if s.start_time <= t <= end:
                return True
        return False

    quizzes_in_class = Quiz.query.filter_by(class_id=class_id).all()
    quizzes_by_index = {q.quiz_index: q for q in quizzes_in_class}
    question_counts = {}
    if quizzes_in_class:
        qids = [q.id for q in quizzes_in_class]
        for cnt, qz_id in (
            db.session.query(func.count(QuizQuestion.id), QuizQuestion.quiz_id)
            .filter(QuizQuestion.quiz_id.in_(qids))
            .group_by(QuizQuestion.quiz_id)
            .all()
        ):
            question_counts[qz_id] = int(cnt)

    latest_graded_run_by_quiz_id = {}
    for qz in quizzes_in_class:
        runs_desc = (
            QuizRun.query.filter_by(class_id=class_id, quiz_id=qz.id)
            .order_by(QuizRun.started_at.desc())
            .all()
        )
        chosen = None
        for run in runs_desc:
            if _quiz_run_completed(run, now) and _run_in_graded_session_window(run):
                chosen = run
                break
        latest_graded_run_by_quiz_id[qz.id] = chosen

    run_ids_for_answers = [r.id for r in latest_graded_run_by_quiz_id.values() if r is not None]
    answers_by_run_student = {}
    if run_ids_for_answers:
        for ans in QuizAnswer.query.filter(QuizAnswer.quiz_run_id.in_(run_ids_for_answers)).all():
            answers_by_run_student.setdefault((ans.quiz_run_id, ans.student_id), []).append(ans)

    def _attended(student_id, session):
        a = attend_by_sid.get((student_id, session.id))
        if a is not None:
            return bool(a.join_time is not None and a.present)
        d = session.start_time.date()
        if graded_by_date_count.get(d, 0) != 1:
            return False
        leg = legacy_by_date.get((student_id, d))
        return bool(leg and leg.join_time is not None and leg.present)

    def _session_score(student_id, session, student_hrs, student_rounds):
        if not _attended(student_id, session):
            return None
        session_start = session.start_time
        session_end = session.end_time or now
        candidates = [25.0]
        for ts in student_hrs:
            if ts is not None and session_start <= ts <= session_end:
                candidates.append(40.0)
                break
        for rnd in student_rounds:
            in_session = False
            if rnd.class_session_id == session.id:
                in_session = True
            elif rnd.class_session_id is None and rnd.created_at:
                if session_start <= rnd.created_at <= session_end:
                    in_session = True
            if not in_session:
                continue
            inst_score = instr_by_round.get(rnd.id)
            if inst_score is None:
                continue
            peer_val = peer_avg_by_round.get(rnd.id, 0.0)
            candidates.append(
                blended_participation_grade(inst_score, peer_val, inst_share)
            )
        return max(candidates)

    rows = []
    for stu in students:
        student_hrs = hr_by_student.get(stu.id, [])
        student_rounds = rounds_by_subject.get(stu.id, [])
        exempts = exempt_map.get(stu.id, set())

        count = 0
        denom = 0
        scores = []
        for s in graded_sessions:
            if s.id in exempts:
                pass
            else:
                denom += 1
                if _attended(stu.id, s):
                    count += 1
            sc = _session_score(stu.id, s, student_hrs, student_rounds)
            if sc is not None:
                scores.append(sc)
        attendance_grade = (count / denom) * 100 if denom > 0 else 100.0
        participation_grade = sum(scores) / len(scores) if scores else 0.0

        parts = parts_by_student.get(stu.id, [])
        if parts:
            avg_peer = sum(p.peer_grade for p in parts) / len(parts)
            avg_inst = sum(p.instructor_grade for p in parts) / len(parts)
        else:
            avg_peer = 0
            avg_inst = 0

        prs = poll_map.get(stu.id, [])
        if prs:
            poll_grade = (sum(1 for pr in prs if pr.is_correct) / len(prs)) * 100
        else:
            poll_grade = 0

        quiz_scores_by_index = {}
        if qw > 0 and quiz_n_target >= 1:
            slot_scores = []
            for idx in range(1, quiz_n_target + 1):
                qdef = quizzes_by_index.get(idx)
                if not qdef:
                    quiz_scores_by_index[idx] = 0.0
                    slot_scores.append(0.0)
                    continue
                run = latest_graded_run_by_quiz_id.get(qdef.id)
                nq = question_counts.get(qdef.id, 0)
                if not run or nq <= 0:
                    quiz_scores_by_index[idx] = 0.0
                    slot_scores.append(0.0)
                    continue
                ans_list = answers_by_run_student.get((run.id, stu.id), [])
                correct = sum(1 for a in ans_list if a.is_correct)
                sc = (correct / nq) * 100.0
                quiz_scores_by_index[idx] = sc
                slot_scores.append(sc)
            quiz_grade = sum(slot_scores) / len(slot_scores) if slot_scores else 0.0
        else:
            quiz_grade = 0.0
            quiz_scores_by_index = {}

        overall_grade = (
            (attendance_grade * eff_att_w / 100)
            + (participation_grade * part_weight / 100)
            + (poll_grade * eff_poll_w / 100)
            + (quiz_grade * eff_quiz_w / 100)
        )

        rows.append({
            'student': stu,
            'attendance_grade': attendance_grade,
            'participation_grade': participation_grade,
            'avg_peer_grade': avg_peer,
            'avg_instructor_grade': avg_inst,
            'poll_grade': poll_grade,
            'quiz_grade': quiz_grade,
            'quiz_scores_by_index': quiz_scores_by_index,
            'overall_grade': overall_grade,
        })

    return rows


@login_manager.user_loader
def load_user(user_id):
    return Professor.query.get(int(user_id))

# Routes
@app.route('/')
def index():
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    return redirect(url_for('login'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        user_type = request.form.get('user_type', 'professor')
        
        if user_type == 'professor':
            ident = (username or '').strip()
            professor = Professor.query.filter_by(username=ident).first() if ident else None
            if professor is None and ident:
                professor = Professor.query.filter(
                    func.lower(Professor.email) == ident.lower()
                ).first()
            if professor and check_password_hash(professor.password_hash, password):
                login_user(professor)
                return jsonify({'success': True, 'redirect': url_for('dashboard')})
            return jsonify({'success': False, 'error': 'Invalid credentials'})
        else:
            # Student login will be handled differently
            return jsonify({'success': False, 'error': 'Use student interface'})
    
    return render_template('login.html')

@app.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'POST':
        data = request.get_json() if request.is_json else request.form.to_dict()
        username = data.get('username')
        email = data.get('email')
        password = data.get('password')
        
        if not username or not email or not password:
            return jsonify({'success': False, 'error': 'All fields are required'})
        
        # Check if username already exists
        if Professor.query.filter_by(username=username).first():
            return jsonify({'success': False, 'error': 'Username already exists'})
        
        # Check if email already exists
        if Professor.query.filter_by(email=email).first():
            return jsonify({'success': False, 'error': 'Email already exists'})
        
        # Create new professor
        professor = Professor(
            username=username,
            email=email,
            password_hash=generate_password_hash(password)
        )
        db.session.add(professor)
        db.session.commit()
        
        # Auto-login the new professor
        login_user(professor)
        
        return jsonify({'success': True, 'redirect': url_for('dashboard')})
    
    return render_template('register.html')


@app.route('/forgot-password', methods=['GET', 'POST'])
def forgot_password():
    """Professor password reset request; optional delivery email distinct from account email."""
    generic_ok = (
        'If an account exists with that email, password reset instructions have been sent.'
    )
    if request.method == 'POST':
        raw = (request.form.get('account_email') or request.form.get('email') or '').strip()
        send_to_raw = (request.form.get('send_to_email') or '').strip()
        if send_to_raw and not _looks_like_email(send_to_raw):
            return jsonify({
                'success': False,
                'error': 'Please enter a valid email address for where to send the reset link.',
            })
        email = raw.lower()
        if email:
            professor = Professor.query.filter(func.lower(Professor.email) == email).first()
            if professor:
                token = make_professor_password_reset_token(professor.id)
                reset_url = url_for('reset_password', token=token, _external=True)
                delivery = send_to_raw if send_to_raw else None
                _send_professor_password_reset_email(professor, reset_url, delivery_email=delivery)
        return jsonify({'success': True, 'message': generic_ok})
    return render_template('forgot_password.html')


@app.route('/reset-password/<token>', methods=['GET', 'POST'])
def reset_password(token):
    professor_id = verify_professor_password_reset_token(token)
    if professor_id is None:
        if request.method == 'POST':
            return jsonify({'success': False, 'error': 'This reset link is invalid or has expired.'}), 400
        return render_template('reset_password.html', invalid_token=True)

    professor = Professor.query.get(professor_id)
    if not professor:
        if request.method == 'POST':
            return jsonify({'success': False, 'error': 'This reset link is invalid or has expired.'}), 400
        return render_template('reset_password.html', invalid_token=True)

    if request.method == 'POST':
        data = request.get_json() if request.is_json else request.form.to_dict()
        pwd = (data.get('password') or '').strip()
        confirm = (data.get('confirmPassword') or data.get('confirm_password') or '').strip()
        if len(pwd) < 8:
            return jsonify({'success': False, 'error': 'Password must be at least 8 characters.'})
        if pwd != confirm:
            return jsonify({'success': False, 'error': 'Passwords do not match.'})
        professor.password_hash = generate_password_hash(pwd)
        db.session.commit()
        return jsonify({'success': True, 'redirect': url_for('login')})

    return render_template('reset_password.html', invalid_token=False, token=token)


@app.route('/logout')
@login_required
def logout():
    # Stop all active classes for this professor before logging out
    active_classes = Class.query.filter_by(professor_id=current_user.id, is_active=True).all()
    end_time = datetime.utcnow()
    poll_stops_after_commit = []
    quiz_stops_after_commit = []
    for class_obj in active_classes:
        class_obj.is_active = False
        # Close open session
        active_session = ClassSession.query.filter_by(
            class_id=class_obj.id, end_time=None
        ).order_by(ClassSession.start_time.desc()).first()
        if active_session:
            active_session.end_time = end_time
        # Mark leave time for students still in the active session
        if active_session:
            for att in Attendance.query.filter_by(
                class_id=class_obj.id,
                class_session_id=active_session.id,
                present=True,
            ).all():
                if not att.leave_time:
                    att.leave_time = end_time
        for pid in deactivate_active_polls_for_class(class_obj.id):
            poll_stops_after_commit.append((class_obj.id, pid))
        for qrid in deactivate_active_quiz_runs_for_class(class_obj.id):
            quiz_stops_after_commit.append((class_obj.id, qrid))
        socketio.emit('class_stopped', {'class_id': class_obj.id})
    if active_classes:
        db.session.commit()
        for cid, pid in poll_stops_after_commit:
            socketio.emit(
                'poll_stopped',
                {
                    'poll_id': pid,
                    'class_id': cid,
                    'results': poll_results_payload(pid),
                },
                room=f'class_{cid}',
            )
        for cid, qrid in quiz_stops_after_commit:
            socketio.emit(
                'quiz_stopped',
                {'quiz_run_id': qrid, 'class_id': cid},
                room=f'class_{cid}',
            )
    logout_user()
    return redirect(url_for('login'))

@app.route('/dashboard')
@login_required
def dashboard():
    classes = Class.query.filter_by(professor_id=current_user.id).all()
    return render_template('dashboard.html', classes=classes)

@app.route('/preferences')
@login_required
def preferences():
    return render_template('preferences.html')


@app.route('/preferences/account')
@login_required
def account_settings():
    return render_template('account_settings.html')


@app.route('/api/account', methods=['GET'])
@login_required
def get_account():
    """Return current professor username and email (no password)."""
    p = Professor.query.get(current_user.id)
    if not p:
        return jsonify({'success': False, 'error': 'Not found.'}), 404
    return jsonify({
        'success': True,
        'username': p.username,
        'email': p.email,
    })


@app.route('/api/account', methods=['POST'])
@login_required
def update_account():
    """Update professor username, email, and/or password. Requires current password."""
    data = request.get_json() or {}
    current_pw = (data.get('current_password') or '').strip()
    if not current_pw:
        return jsonify({'success': False, 'error': 'Enter your current password to save changes.'}), 400

    prof = Professor.query.get(current_user.id)
    if not prof or not check_password_hash(prof.password_hash, current_pw):
        return jsonify({'success': False, 'error': 'Current password is incorrect.'}), 400

    username = (data.get('username') or '').strip()
    email = (data.get('email') or '').strip()
    new_pw = (data.get('new_password') or '').strip()
    confirm_pw = (data.get('confirm_password') or '').strip()

    if not username:
        return jsonify({'success': False, 'error': 'Username cannot be empty.'}), 400
    if not email or not _looks_like_email(email):
        return jsonify({'success': False, 'error': 'Please enter a valid email address.'}), 400

    changed = False

    if username != prof.username:
        taken = Professor.query.filter_by(username=username).first()
        if taken and taken.id != prof.id:
            return jsonify({'success': False, 'error': 'That Username is already taken.'}), 400
        prof.username = username
        changed = True

    if email.lower() != prof.email.lower():
        taken = Professor.query.filter(func.lower(Professor.email) == email.lower()).first()
        if taken and taken.id != prof.id:
            return jsonify({'success': False, 'error': 'That email is already in use.'}), 400
        prof.email = email
        changed = True

    if new_pw or confirm_pw:
        if len(new_pw) < 8:
            return jsonify({'success': False, 'error': 'New password must be at least 8 characters.'}), 400
        if new_pw != confirm_pw:
            return jsonify({'success': False, 'error': 'New password and confirmation do not match.'}), 400
        prof.password_hash = generate_password_hash(new_pw)
        changed = True

    if not changed:
        return jsonify({'success': False, 'error': 'No changes to save.'}), 400

    db.session.commit()
    return jsonify({'success': True})


@app.route('/api/preferences', methods=['GET'])
@login_required
def get_preferences():
    """Get global professor preferences."""
    prefs = ProfessorPreferences.query.filter_by(professor_id=current_user.id).first()
    if not prefs:
        prefs = ProfessorPreferences(professor_id=current_user.id)
        db.session.add(prefs)
        db.session.commit()
    return jsonify({
        'success': True,
        'show_first_name_only': prefs.default_show_first_name_only,
        'quiet_mode': prefs.default_quiet_mode,
        'dark_mode': bool(getattr(prefs, 'dark_mode', False)),
    })

@app.route('/api/preferences', methods=['POST'])
@login_required
def save_preferences():
    """Save global professor preferences. Only keys present in JSON are updated."""
    data = request.get_json() or {}
    prefs = ProfessorPreferences.query.filter_by(professor_id=current_user.id).first()
    if not prefs:
        prefs = ProfessorPreferences(professor_id=current_user.id)
        db.session.add(prefs)
    if 'show_first_name_only' in data:
        prefs.default_show_first_name_only = bool(data['show_first_name_only'])
    if 'quiet_mode' in data:
        prefs.default_quiet_mode = bool(data['quiet_mode'])
    if 'dark_mode' in data:
        prefs.dark_mode = bool(data['dark_mode'])
    db.session.commit()
    return jsonify({
        'success': True,
        'dark_mode': bool(getattr(prefs, 'dark_mode', False)),
    })


@app.context_processor
def inject_ui_theme():
    """Expose dark mode to templates for logged-in professors."""
    try:
        if current_user.is_authenticated:
            prefs = ProfessorPreferences.query.filter_by(professor_id=current_user.id).first()
            dark = bool(prefs and getattr(prefs, 'dark_mode', False))
            return {'ui_dark_mode': dark}
    except Exception:
        pass
    return {'ui_dark_mode': False}


@app.route('/classroom/<int:class_id>')
@login_required
def classroom(class_id):
    class_obj = Class.query.get_or_404(class_id)
    if class_obj.professor_id != current_user.id:
        return redirect(url_for('dashboard'))
    
    students = db.session.query(Student).join(Enrollment).filter(
        Enrollment.class_id == class_id
    ).all()
    
    settings = ClassSettings.query.filter_by(class_id=class_id).first()
    if not settings:
        settings = ClassSettings(class_id=class_id)
        db.session.add(settings)
        db.session.commit()
    
    return render_template('classroom.html', class_obj=class_obj, students=students, settings=settings)

@app.route('/classroom/<int:class_id>/students')
@login_required
def students_list(class_id):
    class_obj = Class.query.get_or_404(class_id)
    if class_obj.professor_id != current_user.id:
        return redirect(url_for('dashboard'))
    
    # Get active students
    active_students = db.session.query(Student).join(Enrollment).filter(
        Enrollment.class_id == class_id,
        Enrollment.is_active == True
    ).order_by(Student.last_name, Student.first_name).all()
    
    # Get inactive students
    inactive_students = db.session.query(Student).join(Enrollment).filter(
        Enrollment.class_id == class_id,
        Enrollment.is_active == False
    ).order_by(Student.last_name, Student.first_name).all()
    
    return render_template('students_list.html', class_obj=class_obj, active_students=active_students, inactive_students=inactive_students)

@app.route('/classroom/<int:class_id>/class_data')
@login_required
def class_data(class_id):
    class_obj = Class.query.get_or_404(class_id)
    if class_obj.professor_id != current_user.id:
        return redirect(url_for('dashboard'))
    
    return render_template('class_data.html', class_obj=class_obj)

@app.route('/api/start_class/<int:class_id>', methods=['POST'])
@login_required
def start_class(class_id):
    class_obj = Class.query.get_or_404(class_id)
    if class_obj.professor_id != current_user.id:
        return jsonify({'success': False, 'error': 'Unauthorized'})
    
    class_obj.is_active = True
    
    # Get exclude_from_grading from request if provided
    data = request.get_json() or {}
    exclude_from_grading = data.get('exclude_from_grading', False)
    
    # Create a new class session
    session_record = ClassSession(
        class_id=class_id,
        start_time=datetime.utcnow(),
        exclude_from_grading=exclude_from_grading
    )
    db.session.add(session_record)

    # Reset interaction counters so each session starts at zero
    today = datetime.utcnow().date()
    for p in Participation.query.filter_by(class_id=class_id, date=today).all():
        p.hand_raises = 0
        p.thumbs_up = 0
        p.thumbs_down = 0
    # Clear any lingering active hand raises from a previous session
    for hr in HandRaise.query.filter_by(class_id=class_id, cleared=False).all():
        hr.cleared = True

    db.session.commit()

    # Broadcast so every connected client sees it immediately (room-only can miss students on My Classes)
    started_payload = {'class_id': class_id, 'class_code': class_obj.class_code}
    socketio.emit('class_started', started_payload)

    return jsonify({'success': True, 'redirect': url_for('faculty_dashboard', class_id=class_id)})

@app.route('/api/stop_class/<int:class_id>', methods=['POST'])
@login_required
def stop_class(class_id):
    class_obj = Class.query.get_or_404(class_id)
    if class_obj.professor_id != current_user.id:
        return jsonify({'success': False, 'error': 'Unauthorized'})
    
    class_obj.is_active = False
    
    # Update the active session with end time
    active_session = ClassSession.query.filter_by(
        class_id=class_id,
        end_time=None
    ).order_by(ClassSession.start_time.desc()).first()
    
    end_time = datetime.utcnow()
    if active_session:
        active_session.end_time = end_time
    
    # Log leave_time for students still in this session (auto-logout at class end)
    if active_session:
        attendances = Attendance.query.filter_by(
            class_id=class_id,
            class_session_id=active_session.id,
            present=True,
        ).all()
        for attendance in attendances:
            if not attendance.leave_time:
                attendance.leave_time = end_time

    poll_ids_stopped = deactivate_active_polls_for_class(class_id)
    quiz_run_ids_stopped = deactivate_active_quiz_runs_for_class(class_id)

    db.session.commit()

    emit_poll_stopped_events(class_id, poll_ids_stopped)
    for qrid in quiz_run_ids_stopped:
        socketio.emit(
            'quiz_stopped',
            {'quiz_run_id': qrid, 'class_id': class_id},
            room=f'class_{class_id}',
        )

    # Notify students immediately — must run before update_gradebook (can take many seconds)
    stopped_payload = {'class_id': class_id}
    socketio.emit('class_stopped', stopped_payload)

    # Update gradebook with participation data (heavy; do not block real-time events above)
    update_gradebook(class_id)

    return jsonify({'success': True})

def update_gradebook(class_id):
    today = datetime.utcnow().date()
    students = db.session.query(Student).join(Enrollment).filter(
        Enrollment.class_id == class_id
    ).all()
    poll_map = gradebook_poll_responses_by_student(class_id)

    for student in students:
        participation = Participation.query.filter_by(
            class_id=class_id,
            student_id=student.id,
            date=today
        ).first()

        if not participation:
            participation = Participation(
                class_id=class_id,
                student_id=student.id,
                date=today
            )
            db.session.add(participation)

        poll_responses = poll_map.get(student.id, [])
        poll_grade = 0.0
        if poll_responses:
            correct_count = sum(1 for pr in poll_responses if pr.is_correct)
            poll_grade = (correct_count / len(poll_responses)) * 100

    db.session.commit()

@app.route('/faculty_dashboard/<int:class_id>')
@login_required
def faculty_dashboard(class_id):
    class_obj = Class.query.get_or_404(class_id)
    if class_obj.professor_id != current_user.id:
        return redirect(url_for('dashboard'))
    
    students = db.session.query(Student).join(Enrollment).filter(
        Enrollment.class_id == class_id
    ).all()
    
    active_poll = Poll.query.filter_by(class_id=class_id, is_active=True).first()
    
    return render_template('faculty_dashboard.html', class_obj=class_obj, students=students, active_poll=active_poll)

@app.route('/api/gradebook/<int:class_id>')
@login_required
def get_gradebook(class_id):
    class_obj = Class.query.get_or_404(class_id)
    if class_obj.professor_id != current_user.id:
        return jsonify({'error': 'Unauthorized'}), 403

    grading_weights = GradingWeights.query.filter_by(class_id=class_id).first()
    if not grading_weights:
        grading_weights = GradingWeights(
            class_id=class_id,
            attendance_weight=25.0,
            participation_weight=50.0,
            participation_instructor_share=50.0,
            poll_weight=25.0,
            quiz_weight=0.0,
            quiz_count_target=0,
        )
        db.session.add(grading_weights)
        db.session.commit()

    rows = _compute_gradebook_rows(class_id, grading_weights)
    gradebook_data = [
        {
            'student_id': r['student'].id,
            'student_number': r['student'].student_number,
            'name': f"{r['student'].first_name} {r['student'].last_name}",
            'attendance_grade': round(r['attendance_grade'], 2),
            'participation_grade': round(r['participation_grade'], 2),
            'peer_participation': round(r['avg_peer_grade'], 2),
            'instructor_participation': round(r['avg_instructor_grade'], 2),
            'poll_grade': round(r['poll_grade'], 2),
            'quiz_grade': round(r['quiz_grade'], 2),
            'quiz_scores_by_index': {str(k): round(v, 2) for k, v in (r.get('quiz_scores_by_index') or {}).items()},
            'overall_grade': round(r['overall_grade'], 2),
        }
        for r in rows
    ]

    return jsonify(gradebook_data)

@app.route('/api/export_gradebook/<int:class_id>')
@login_required
def export_gradebook(class_id):
    class_obj = Class.query.get_or_404(class_id)
    if class_obj.professor_id != current_user.id:
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403

    grading_weights = GradingWeights.query.filter_by(class_id=class_id).first()
    if not grading_weights:
        grading_weights = GradingWeights(
            class_id=class_id,
            attendance_weight=25.0,
            participation_weight=50.0,
            participation_instructor_share=50.0,
            poll_weight=25.0,
            quiz_weight=0.0,
            quiz_count_target=0,
        )

    wb = Workbook()
    ws = wb.active
    ws.title = "Gradebook"

    headers = ['Student Name', 'Student Number', 'Email', 'Attendance Grade (%)',
               'Participation (%)', 'Instructor Participation', 'Peer Participation', 'Poll Grade (%)',
               'Quiz Grade (%)', 'Overall Grade (%)']
    ws.append(headers)

    header_fill = PatternFill(start_color="2A1A40", end_color="2A1A40", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    for r in _compute_gradebook_rows(class_id, grading_weights):
        student = r['student']
        row = [
            f"{student.first_name} {student.last_name}",
            student.student_number,
            student.email if hasattr(student, 'email') else '',
            round(r['attendance_grade'], 2),
            round(r['participation_grade'], 2),
            round(r['avg_instructor_grade'], 2),
            round(r['avg_peer_grade'], 2),
            round(r['poll_grade'], 2),
            round(r['quiz_grade'], 2),
            round(r['overall_grade'], 2),
        ]
        ws.append(row)
    
    # Auto-fit column widths
    from openpyxl.utils import get_column_letter
    for col_idx, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_idx)
        max_length = len(str(header))
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max(max_length * 1.15 + 2, 12), 50)
    
    # Create BytesIO object to store the Excel file in memory
    excel_buffer = BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    return send_file(
        excel_buffer,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'gradebook_{class_obj.name.replace(" ", "_")}_{datetime.now().strftime("%Y%m%d")}.xlsx'
    )

@app.route('/api/class_metrics/<int:class_id>')
@login_required
def get_class_metrics(class_id):
    class_obj = Class.query.get_or_404(class_id)
    if class_obj.professor_id != current_user.id:
        return jsonify({'error': 'Unauthorized'}), 403

    sessions = ClassSession.query.filter_by(class_id=class_id).order_by(ClassSession.start_time.desc()).all()

    enrolled_students = db.session.query(Student).join(Enrollment).filter(
        Enrollment.class_id == class_id,
        Enrollment.is_active == True
    ).all()

    # Hoist per-class data out of the per-session loop. All of these queries
    # previously ran once per session, multiplying with the number of rounds.
    settings_for_class = ClassSettings.query.filter_by(class_id=class_id).first()
    show_first_only = bool(settings_for_class and settings_for_class.show_first_name_only)
    fn_labels = _first_name_only_labels_for_students(enrolled_students) if show_first_only else {}

    def _subject_label(student):
        if show_first_only:
            return fn_labels.get(
                student.id,
                (student.preferred_name or student.first_name or '').strip() or '?',
            )
        return f'{student.preferred_name or student.first_name} {student.last_name}'.strip()

    attendance_by_session = {}
    attendance_legacy_by_date = {}
    for a in Attendance.query.filter_by(class_id=class_id).all():
        if a.class_session_id is not None:
            attendance_by_session.setdefault(a.class_session_id, []).append(a)
        else:
            attendance_legacy_by_date.setdefault(a.date, []).append(a)

    all_hrs = HandRaise.query.filter_by(class_id=class_id).all()
    all_polls = Poll.query.filter_by(class_id=class_id).all()

    responses_by_poll = {}
    poll_ids = [p.id for p in all_polls]
    if poll_ids:
        for pr in PollResponse.query.filter(PollResponse.poll_id.in_(poll_ids)).all():
            responses_by_poll.setdefault(pr.poll_id, []).append(pr)

    exempt_by_session = {}
    for row in AbsenceExemption.query.filter_by(class_id=class_id).all():
        exempt_by_session.setdefault(row.class_session_id, set()).add(row.student_id)

    all_rounds = (
        ParticipationGradeRound.query.filter_by(class_id=class_id)
        .order_by(ParticipationGradeRound.created_at.desc())
        .all()
    )
    round_ids = [r.id for r in all_rounds]
    instructor_score_by_round = {}
    peer_scores_by_round = {}
    if round_ids:
        for ig in InstructorParticipationGrade.query.filter(
            InstructorParticipationGrade.round_id.in_(round_ids)
        ).all():
            instructor_score_by_round[ig.round_id] = ig.score
        for pg in PeerParticipationGrade.query.filter(
            PeerParticipationGrade.round_id.in_(round_ids)
        ).all():
            peer_scores_by_round.setdefault(pg.round_id, []).append(pg.score_percent)

    # Some round subjects may no longer be actively enrolled; fetch them in one IN query.
    enrolled_by_id = {s.id: s for s in enrolled_students}
    missing_subject_ids = {
        r.subject_student_id for r in all_rounds if r.subject_student_id is not None
    } - set(enrolled_by_id)
    subjects_by_id = dict(enrolled_by_id)
    if missing_subject_ids:
        for s in Student.query.filter(Student.id.in_(missing_subject_ids)).all():
            subjects_by_id[s.id] = s

    all_quiz_runs = (
        QuizRun.query.filter_by(class_id=class_id)
        .options(joinedload(QuizRun.quiz).joinedload(Quiz.questions))
        .all()
    )
    quiz_run_ids = [r.id for r in all_quiz_runs]
    submitted_quiz_pairs = set()
    correct_quiz_by_pair = {}
    answers_by_run_student = {}
    if quiz_run_ids:
        from collections import defaultdict

        correct_quiz_by_pair = defaultdict(int)
        answers_by_run_student = defaultdict(dict)
        for a in QuizAnswer.query.filter(QuizAnswer.quiz_run_id.in_(quiz_run_ids)).all():
            submitted_quiz_pairs.add((a.quiz_run_id, a.student_id))
            if a.is_correct:
                correct_quiz_by_pair[(a.quiz_run_id, a.student_id)] += 1
            answers_by_run_student[(a.quiz_run_id, a.student_id)][a.question_id] = a

    same_day_count = {}
    for s in sessions:
        d = s.start_time.date()
        same_day_count[d] = same_day_count.get(d, 0) + 1

    total_enrolled = len(enrolled_students)
    sessions_data = []
    session_number = len(sessions)

    for session in sessions:
        session_date = session.start_time.date()
        session_start = session.start_time
        session_end = session.end_time if session.end_time else datetime.utcnow()

        attendance_map = {att.student_id: att for att in attendance_by_session.get(session.id, [])}
        if same_day_count.get(session_date, 0) == 1:
            for att in attendance_legacy_by_date.get(session_date, []):
                if att.student_id not in attendance_map:
                    attendance_map[att.student_id] = att

        attendance_count = sum(
            1 for st in enrolled_students
            if attendance_map.get(st.id) and attendance_map[st.id].join_time is not None
        )
        attendance_percentage = (attendance_count / total_enrolled * 100) if total_enrolled > 0 else 0

        unique_hands_raised = len({
            hr.student_id for hr in all_hrs
            if hr.timestamp is not None and session_start <= hr.timestamp <= session_end
        })

        polls = [
            p for p in all_polls
            if p.created_at is not None and session_start <= p.created_at <= session_end
        ]

        poll_results = []
        total_poll_responses = 0
        for poll in polls:
            responses = responses_by_poll.get(poll.id, [])
            total_poll_responses += len(responses)

            options = json.loads(poll.options)
            option_counts = {}
            for i in range(len(options)):
                option_counts[i] = sum(1 for r in responses if r.answer == i)

            poll_results.append({
                'poll_id': poll.id,
                'question': poll.question,
                'options': options,
                'option_counts': option_counts,
                'total_responses': len(responses),
                'is_graded': poll.is_graded
            })

        # Calculate overall poll vote percentage for the session as:
        # total submitted responses / total possible responses across all polls.
        # If no polls were conducted, return None so UI can render "N/A".
        poll_vote_percentage = None
        if polls:
            total_possible_poll_responses = attendance_count * len(polls)
            if total_possible_poll_responses > 0:
                poll_vote_percentage = min(100.0, (total_poll_responses / total_possible_poll_responses) * 100)
            else:
                poll_vote_percentage = 0.0

        exempt_student_ids = exempt_by_session.get(session.id, set())

        attendance_list = []
        for student in enrolled_students:
            att = attendance_map.get(student.id)

            if att:
                sign_in_time = att.join_time if att.join_time else att.timestamp
                sign_out_time = None
                if att.leave_time:
                    if session.end_time is None:
                        sign_out_time = att.leave_time
                    elif att.leave_time <= session.end_time:
                        # Early logout (before end) or exit stamped when teacher ends class (same timestamp)
                        sign_out_time = att.leave_time
            else:
                sign_in_time = None
                sign_out_time = None

            attended_session = bool(att and att.join_time is not None)
            attendance_list.append({
                'student_id': student.id,
                'student_number': student.student_number,
                'student_name': f"{student.preferred_name or student.first_name} {student.last_name}",
                'sign_in_time': sign_in_time.isoformat() if sign_in_time else None,
                'sign_out_time': sign_out_time.isoformat() if sign_out_time else None,
                'present': attended_session,
                'absence_exempt': student.id in exempt_student_ids,
            })

        grade_rounds = []
        for rnd in all_rounds:
            if rnd.class_session_id == session.id:
                grade_rounds.append(rnd)
            elif rnd.class_session_id is None and rnd.date == session_date:
                if rnd.created_at and session_start <= rnd.created_at <= session_end:
                    grade_rounds.append(rnd)

        participation_rounds = []
        for rnd in grade_rounds:
            subject = subjects_by_id.get(rnd.subject_student_id)
            if not subject:
                continue
            inst_score = instructor_score_by_round.get(rnd.id)
            peers = peer_scores_by_round.get(rnd.id, [])
            peer_avg = (sum(peers) / len(peers)) if peers else None
            participation_rounds.append({
                'round_id': rnd.id,
                'subject_student_id': rnd.subject_student_id,
                'subject_name': _subject_label(subject),
                'instructor_score': int(inst_score) if inst_score is not None else None,
                'peer_score_percent': round(peer_avg, 2) if peer_avg is not None else None,
                'exclude_from_grading': bool(rnd.exclude_from_grading),
                'created_at': rnd.created_at.isoformat() if rnd.created_at else None,
            })

        quiz_runs_in_session = [
            r for r in all_quiz_runs
            if r.started_at is not None and session_start <= r.started_at <= session_end
        ]
        quiz_runs_in_session.sort(key=lambda r: r.started_at or datetime.min)
        quiz_results = []
        for run in quiz_runs_in_session:
            qz = run.quiz
            quiz_title = (qz.title if qz else '') or 'Quiz'
            nq = len(qz.questions) if qz and qz.questions else 0
            q_rows_ordered = sorted(qz.questions, key=lambda q: q.order) if qz and qz.questions else []
            student_scores = []
            for student in enrolled_students:
                pair = (run.id, student.id)
                if pair not in submitted_quiz_pairs:
                    student_scores.append({
                        'student_id': student.id,
                        'student_number': student.student_number,
                        'student_name': _subject_label(student),
                        'correct_count': None,
                        'total_questions': nq,
                        'percent': None,
                        'question_breakdown': None,
                    })
                else:
                    cc = correct_quiz_by_pair[pair]
                    pct = round(100.0 * cc / nq, 1) if nq > 0 else None
                    ans_map = answers_by_run_student.get(pair, {})
                    question_breakdown = []
                    for qq in q_rows_ordered:
                        ans = ans_map.get(qq.id)
                        if not ans:
                            continue
                        try:
                            opts = json.loads(qq.options or '[]')
                        except (json.JSONDecodeError, TypeError):
                            opts = []

                        def _opt_label(idx):
                            if idx is None or not isinstance(idx, int) or idx < 0 or idx >= len(opts):
                                return '—'
                            return str(opts[idx])

                        question_breakdown.append({
                            'question_id': qq.id,
                            'prompt': qq.prompt,
                            'selected_index': ans.selected_index,
                            'correct_index': qq.correct_index,
                            'selected_text': _opt_label(ans.selected_index),
                            'correct_text': _opt_label(qq.correct_index),
                            'is_correct': bool(ans.is_correct),
                        })
                    student_scores.append({
                        'student_id': student.id,
                        'student_number': student.student_number,
                        'student_name': _subject_label(student),
                        'correct_count': cc,
                        'total_questions': nq,
                        'percent': pct,
                        'question_breakdown': question_breakdown,
                    })
            quiz_results.append({
                'quiz_run_id': run.id,
                'quiz_title': quiz_title,
                'started_at': run.started_at.isoformat() if run.started_at else None,
                'question_count': nq,
                'student_scores': student_scores,
            })

        sessions_data.append({
            'session_id': session.id,
            'session_number': session_number,
            'start_time': session.start_time.isoformat(),
            'end_time': session.end_time.isoformat() if session.end_time else None,
            'exclude_from_grading': session.exclude_from_grading,
            'engagement_metrics': {
                'attendance': round(attendance_percentage, 1),
                'unique_hands_raised': unique_hands_raised,
                'poll_vote_percentage': (round(poll_vote_percentage, 1) if poll_vote_percentage is not None else None)
            },
            'poll_results': poll_results,
            'quiz_results': quiz_results,
            'participation_rounds': participation_rounds,
            'attendance_list': attendance_list
        })
        
        session_number -= 1  # Decrement for next session
    
    return jsonify(sessions_data)

@app.route('/api/update_poll_grading/<int:poll_id>', methods=['POST'])
@login_required
def update_poll_grading(poll_id):
    poll = Poll.query.get_or_404(poll_id)
    class_obj = Class.query.get_or_404(poll.class_id)
    
    if class_obj.professor_id != current_user.id:
        return jsonify({'error': 'Unauthorized'}), 403
    
    data = request.get_json()
    is_graded = data.get('is_graded', False)
    
    poll.is_graded = is_graded
    db.session.commit()
    
    return jsonify({'success': True, 'is_graded': poll.is_graded})

@app.route('/api/update_session_grading/<int:session_id>', methods=['POST'])
@login_required
def update_session_grading(session_id):
    session = ClassSession.query.get_or_404(session_id)
    class_obj = Class.query.get_or_404(session.class_id)
    
    if class_obj.professor_id != current_user.id:
        return jsonify({'error': 'Unauthorized'}), 403
    
    data = request.get_json()
    exclude_from_grading = data.get('exclude_from_grading', False)
    
    session.exclude_from_grading = exclude_from_grading
    db.session.commit()
    
    return jsonify({'success': True, 'exclude_from_grading': session.exclude_from_grading})

@app.route('/api/delete_session/<int:session_id>', methods=['DELETE'])
@login_required
def delete_session(session_id):
    session = ClassSession.query.get_or_404(session_id)
    class_obj = Class.query.get_or_404(session.class_id)
    
    if class_obj.professor_id != current_user.id:
        return jsonify({'error': 'Unauthorized'}), 403
    
    AbsenceExemption.query.filter_by(class_session_id=session_id).delete()
    # Delete the session
    db.session.delete(session)
    db.session.commit()
    
    return jsonify({'success': True})


@app.route('/api/class_session/<int:session_id>/absence_exemption', methods=['POST'])
@login_required
def set_absence_exemption(session_id):
    """Set or clear absence exemption for a student in a class session (Class Metrics)."""
    class_session = ClassSession.query.get_or_404(session_id)
    class_obj = Class.query.get_or_404(class_session.class_id)
    if class_obj.professor_id != current_user.id:
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403

    data = request.get_json() or {}
    student_id = data.get('student_id')
    try:
        student_id = int(student_id)
    except (TypeError, ValueError):
        return jsonify({'success': False, 'error': 'student_id required'}), 400

    exempt = bool(data.get('exempt', True))
    enroll = Enrollment.query.filter_by(
        class_id=class_obj.id,
        student_id=student_id,
        is_active=True,
    ).first()
    if not enroll:
        return jsonify({'success': False, 'error': 'Student is not enrolled in this class.'}), 400

    existing = AbsenceExemption.query.filter_by(
        class_session_id=session_id,
        student_id=student_id,
    ).first()
    if exempt:
        if not existing:
            db.session.add(
                AbsenceExemption(
                    class_id=class_obj.id,
                    class_session_id=session_id,
                    student_id=student_id,
                )
            )
    else:
        if existing:
            db.session.delete(existing)
    db.session.commit()
    row = AbsenceExemption.query.filter_by(
        class_session_id=session_id,
        student_id=student_id,
    ).first()
    return jsonify({'success': True, 'absence_exempt': row is not None})


@app.route('/api/participation_grade_round/<int:round_id>/exclude_from_grading', methods=['POST'])
@login_required
def participation_grade_round_exclude_from_grading(round_id):
    """Professor excludes or re-includes a participation grading round in averages."""
    rnd = ParticipationGradeRound.query.get_or_404(round_id)
    class_obj = Class.query.get_or_404(rnd.class_id)
    if class_obj.professor_id != current_user.id:
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403
    data = request.get_json() or {}
    exclude = bool(data.get('exclude_from_grading', False))
    rnd.exclude_from_grading = exclude
    _recompute_subject_participation_grades(rnd.class_id, rnd.subject_student_id, rnd.date)
    db.session.commit()
    return jsonify({'success': True, 'exclude_from_grading': rnd.exclude_from_grading})


@app.route('/api/participation_grade_round/<int:round_id>', methods=['DELETE'])
@login_required
def delete_participation_grade_round(round_id):
    """Professor permanently deletes a participation grading round and recomputes aggregates."""
    rnd = ParticipationGradeRound.query.get_or_404(round_id)
    class_obj = Class.query.get_or_404(rnd.class_id)
    if class_obj.professor_id != current_user.id:
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403
    class_id = rnd.class_id
    subject_id = rnd.subject_student_id
    grade_date = rnd.date
    PeerParticipationGrade.query.filter_by(round_id=round_id).delete(synchronize_session=False)
    InstructorParticipationGrade.query.filter_by(round_id=round_id).delete(synchronize_session=False)
    db.session.delete(rnd)
    _recompute_subject_participation_grades(class_id, subject_id, grade_date)
    db.session.commit()
    return jsonify({'success': True})


@app.route('/api/update_settings/<int:class_id>', methods=['POST'])
@login_required
def update_settings(class_id):
    class_obj = Class.query.get_or_404(class_id)
    if class_obj.professor_id != current_user.id:
        return jsonify({'success': False, 'error': 'Unauthorized'})
    
    settings = ClassSettings.query.filter_by(class_id=class_id).first()
    if not settings:
        settings = ClassSettings(class_id=class_id)
        db.session.add(settings)
    
    data = request.get_json()
    
    # Update class name and code if provided
    if 'class_name' in data:
        class_obj.name = data.get('class_name')
    if 'class_code' in data:
        new_class_code = data.get('class_code')
        # Check if class code is unique (excluding current class)
        existing_class = Class.query.filter_by(class_code=new_class_code).first()
        if existing_class and existing_class.id != class_id:
            return jsonify({'success': False, 'error': 'Class code already exists'})
        class_obj.class_code = new_class_code
    
    # Update settings
    settings.show_first_name_only = data.get('show_first_name_only', False)
    settings.quiet_mode = data.get('quiet_mode', False)
    
    active_session = ClassSession.query.filter_by(
        class_id=class_id,
        end_time=None
    ).order_by(ClassSession.start_time.desc()).first()
    
    if 'exclude_from_grading' in data and active_session:
        active_session.exclude_from_grading = data.get('exclude_from_grading', False)
    
    db.session.commit()
    
    settings_payload = {
        'class_id': class_id,
        'show_first_name_only': settings.show_first_name_only,
        'quiet_mode': settings.quiet_mode,
        'exclude_from_grading': active_session.exclude_from_grading if active_session else False,
    }
    socketio.emit('settings_updated', settings_payload, room=f'class_{class_id}')
    socketio.emit('settings_updated', settings_payload, room=f'enrolled_{class_id}')
    
    return jsonify({'success': True})

@app.route('/api/grading_weights/<int:class_id>', methods=['GET'])
@login_required
def get_grading_weights(class_id):
    class_obj = Class.query.get_or_404(class_id)
    if class_obj.professor_id != current_user.id:
        return jsonify({'error': 'Unauthorized'}), 403
    
    grading_weights = GradingWeights.query.filter_by(class_id=class_id).first()
    
    if not grading_weights:
        grading_weights = GradingWeights(
            class_id=class_id,
            attendance_weight=25.0,
            participation_weight=50.0,
            participation_instructor_share=50.0,
            poll_weight=25.0,
            quiz_weight=0.0,
            quiz_count_target=0,
        )
        db.session.add(grading_weights)
        db.session.commit()

    peer_share = 100.0 - float(grading_weights.participation_instructor_share)
    return jsonify({
        'attendance_weight': grading_weights.attendance_weight,
        'participation_weight': grading_weights.participation_weight,
        'participation_instructor_share': grading_weights.participation_instructor_share,
        'participation_peer_share': peer_share,
        'poll_weight': grading_weights.poll_weight,
        'quiz_weight': float(getattr(grading_weights, 'quiz_weight', 0.0) or 0.0),
        'quiz_count_target': int(getattr(grading_weights, 'quiz_count_target', 0) or 0),
    })

@app.route('/api/grading_weights/<int:class_id>', methods=['POST'])
@login_required
def update_grading_weights(class_id):
    class_obj = Class.query.get_or_404(class_id)
    if class_obj.professor_id != current_user.id:
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403
    
    data = request.get_json()
    
    attendance_weight = float(data.get('attendance_weight', 25.0))
    participation_weight = float(data.get('participation_weight', 50.0))
    poll_weight = float(data.get('poll_weight', 25.0))
    quiz_weight = float(data.get('quiz_weight', 0.0))
    quiz_count_target = int(data.get('quiz_count_target', 0))
    participation_instructor_share = float(data.get('participation_instructor_share', 50.0))
    
    total_weight = attendance_weight + participation_weight + poll_weight + quiz_weight
    if abs(total_weight - 100.0) > 0.01:
        return jsonify({'success': False, 'error': f'Weights must sum to 100%. Current total: {total_weight}%'}), 400
    
    if any(w < 0 for w in [attendance_weight, participation_weight, poll_weight, quiz_weight]):
        return jsonify({'success': False, 'error': 'All weights must be non-negative'}), 400
    if quiz_count_target < 0:
        return jsonify({'success': False, 'error': 'Number of quizzes must be non-negative'}), 400
    if quiz_weight > 0.0001 and quiz_count_target < 1:
        return jsonify({'success': False, 'error': 'When quiz weight is greater than 0, set number of quizzes to at least 1.'}), 400
    if participation_instructor_share < 0 or participation_instructor_share > 100:
        return jsonify({'success': False, 'error': 'Instructor share of participation must be between 0 and 100%'}), 400
    
    grading_weights = GradingWeights.query.filter_by(class_id=class_id).first()
    
    if not grading_weights:
        grading_weights = GradingWeights(class_id=class_id)
        db.session.add(grading_weights)
    
    grading_weights.attendance_weight = attendance_weight
    grading_weights.participation_weight = participation_weight
    grading_weights.participation_instructor_share = participation_instructor_share
    grading_weights.poll_weight = poll_weight
    grading_weights.quiz_weight = quiz_weight
    grading_weights.quiz_count_target = quiz_count_target
    grading_weights.updated_at = datetime.utcnow()
    
    db.session.commit()
    
    # Recalculate all overall grades for this class with new weights (Task 14)
    # Note: Overall grades are calculated on-the-fly in get_gradebook, so no manual update needed
    # But we can trigger a socket event to notify clients to refresh if needed
    
    return jsonify({'success': True})

@app.route('/api/create_poll/<int:class_id>', methods=['POST'])
@login_required
def create_poll(class_id):
    class_obj = Class.query.get_or_404(class_id)
    if class_obj.professor_id != current_user.id:
        return jsonify({'success': False, 'error': 'Unauthorized'})
    
    data = request.get_json()
    question = data.get('question')
    options = data.get('options', [])
    correct_answer = data.get('correct_answer')
    is_graded = data.get('is_graded', False)
    is_anonymous = data.get('is_anonymous', False)
    show_results_when_stopped = bool(data.get('show_results_when_stopped', True))

    # Validate options
    if not options or len(options) < 2:
        return jsonify({'success': False, 'error': 'At least 2 options are required'}), 400
    
    # Check for duplicate options (case-insensitive, trimmed)
    options_lower = [opt.strip().lower() for opt in options if opt]
    if len(options_lower) != len(set(options_lower)):
        return jsonify({'success': False, 'error': 'Duplicate options are not allowed. Each option must be unique.'}), 400
    
    # Deactivate any existing active polls and quiz runs (single live activity)
    Poll.query.filter_by(class_id=class_id, is_active=True).update({'is_active': False})
    quiz_run_ids = deactivate_active_quiz_runs_for_class(class_id)

    poll = Poll(
        class_id=class_id,
        question=question,
        options=json.dumps(options),
        correct_answer=correct_answer,
        is_graded=is_graded,
        is_anonymous=is_anonymous,
        show_results_when_stopped=show_results_when_stopped,
        is_active=True
    )
    db.session.add(poll)
    db.session.commit()

    for rid in quiz_run_ids:
        socketio.emit(
            'quiz_stopped',
            {'quiz_run_id': rid, 'class_id': class_id},
            room=f'class_{class_id}',
        )
    
    socketio.emit('poll_started', {
        'poll_id': poll.id,
        'question': question,
        'options': options,
        'is_graded': is_graded,
        'is_anonymous': is_anonymous
    }, room=f'class_{class_id}')
    
    return jsonify({'success': True, 'poll_id': poll.id})

@app.route('/api/stop_poll/<int:poll_id>', methods=['POST'])
@login_required
def stop_poll(poll_id):
    poll = Poll.query.get_or_404(poll_id)
    class_obj = Class.query.get_or_404(poll.class_id)
    if class_obj.professor_id != current_user.id:
        return jsonify({'success': False, 'error': 'Unauthorized'})
    
    poll.is_active = False
    db.session.commit()

    socketio.emit(
        'poll_stopped',
        {
            'poll_id': poll_id,
            'class_id': poll.class_id,
            'results': poll_results_payload(poll_id),
        },
        room=f'class_{poll.class_id}',
    )

    return jsonify({'success': True})

@app.route('/api/toggle_poll_graded/<int:poll_id>', methods=['POST'])
@login_required
def toggle_poll_graded(poll_id):
    poll = Poll.query.get_or_404(poll_id)
    class_obj = Class.query.get_or_404(poll.class_id)
    if class_obj.professor_id != current_user.id:
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403
    
    data = request.get_json()
    is_graded = data.get('is_graded', not poll.is_graded)
    
    poll.is_graded = is_graded
    db.session.commit()

    return jsonify({'success': True, 'is_graded': poll.is_graded})

@app.route('/api/clear_poll_responses/<int:poll_id>', methods=['POST'])
@login_required
def clear_poll_responses(poll_id):
    """Delete all student responses for a given poll (professor only)."""
    poll = Poll.query.get_or_404(poll_id)
    class_obj = Class.query.get_or_404(poll.class_id)
    if class_obj.professor_id != current_user.id:
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403
    PollResponse.query.filter_by(poll_id=poll_id).delete()
    db.session.commit()
    socketio.emit('poll_responses_cleared', {'poll_id': poll_id}, room=f'class_{poll.class_id}')
    return jsonify({'success': True})


@app.route('/api/quiz_template.xlsx', methods=['GET'])
@login_required
def download_quiz_template():
    """Excel layout for bulk quiz questions (Question Description, Option A–F, Correct Answer)."""
    wb = Workbook()
    ws = wb.active
    ws.title = 'Questions'
    headers = [
        'Question #',
        'Question Description',
        '# of Options',
        'Option A',
        'Option B',
        'Option C',
        'Option D',
        'Option E',
        'Option F',
        'Correct Answer',
    ]
    ws.append(headers)

    header_fill = PatternFill(start_color="2A1A40", end_color="2A1A40", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    example_row = [1, 'Sample: What is 2+2?', 4, '2', '3', '4', '5', '', '', 'C']
    ws.append(example_row)

    columns_data = [
        ('A', headers[0], example_row[0]),
        ('B', headers[1], example_row[1]),
        ('C', headers[2], example_row[2]),
        ('D', headers[3], example_row[3]),
        ('E', headers[4], example_row[4]),
        ('F', headers[5], example_row[5]),
        ('G', headers[6], example_row[6]),
        ('H', headers[7], example_row[7]),
        ('I', headers[8], example_row[8]),
        ('J', headers[9], example_row[9]),
    ]
    for col_letter, header_text, example_text in columns_data:
        max_content_length = max(len(str(header_text)), len(str(example_text)))
        column_width = max(max_content_length * 1.15 + 2, 12)
        column_width = min(column_width, 50)
        ws.column_dimensions[col_letter].width = column_width

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return send_file(
        bio,
        as_attachment=True,
        download_name='Quiz_Template.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )


@app.route('/api/poll_template.xlsx', methods=['GET'])
@login_required
def download_poll_template():
    """Excel layout for bulk poll-bank questions (quiz-compatible columns)."""
    wb = Workbook()
    ws = wb.active
    ws.title = 'Polls'
    headers = [
        'Question #',
        'Question Description',
        '# of Options',
        'Option A',
        'Option B',
        'Option C',
        'Option D',
        'Option E',
        'Option F',
        'Correct Answer',
    ]
    ws.append(headers)

    header_fill = PatternFill(start_color="2A1A40", end_color="2A1A40", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    example_row = [1, 'Sample: Which topic should we review first?', 4, 'Topic A', 'Topic B', 'Topic C', 'Topic D', '', '', '']
    ws.append(example_row)

    columns_data = [
        ('A', headers[0], example_row[0]),
        ('B', headers[1], example_row[1]),
        ('C', headers[2], example_row[2]),
        ('D', headers[3], example_row[3]),
        ('E', headers[4], example_row[4]),
        ('F', headers[5], example_row[5]),
        ('G', headers[6], example_row[6]),
        ('H', headers[7], example_row[7]),
        ('I', headers[8], example_row[8]),
        ('J', headers[9], example_row[9]),
    ]
    for col_letter, header_text, example_text in columns_data:
        max_content_length = max(len(str(header_text)), len(str(example_text)))
        column_width = max(max_content_length * 1.15 + 2, 12)
        column_width = min(column_width, 50)
        ws.column_dimensions[col_letter].width = column_width

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return send_file(
        bio,
        as_attachment=True,
        download_name='Poll_Template.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )


@app.route('/api/poll_bank/<int:class_id>', methods=['GET'])
@login_required
def list_poll_bank(class_id):
    class_obj = Class.query.get_or_404(class_id)
    if class_obj.professor_id != current_user.id:
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403
    sets = PollBankSet.query.filter_by(class_id=class_id).order_by(PollBankSet.name.asc()).all()
    set_rows = []
    flat_rows = []
    for ps in sets:
        qrows = (
            PollBankQuestion.query
            .filter_by(class_id=class_id, set_id=ps.id)
            .order_by(PollBankQuestion.poll_index.asc())
            .all()
        )
        qout = []
        for row in qrows:
            options = json.loads(row.options or '[]')
            item = {
                'id': row.id,
                'set_id': ps.id,
                'set_name': ps.name,
                'poll_index': row.poll_index,
                'title': row.title,
                'question': row.question,
                'options': options,
                'correct_answer': row.correct_answer,
                'option_count': len(options),
                'source_filename': row.source_filename,
            }
            qout.append(item)
            flat_rows.append(item)
        set_rows.append({'id': ps.id, 'name': ps.name, 'question_count': len(qout), 'questions': qout})
    return jsonify({'success': True, 'poll_sets': set_rows, 'poll_bank': flat_rows})


@app.route('/api/poll_bank_upload/<int:class_id>', methods=['POST'])
@login_required
def poll_bank_upload(class_id):
    class_obj = Class.query.get_or_404(class_id)
    if class_obj.professor_id != current_user.id:
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403
    if 'file' not in request.files:
        return jsonify({'success': False, 'error': 'No file uploaded'}), 400
    file = request.files['file']
    if not file.filename or not file.filename.lower().endswith(('.xlsx', '.xls')):
        return jsonify({'success': False, 'error': 'Please upload an Excel file (.xlsx)'}), 400

    set_name = (request.form.get('set_name') or '').strip()
    if not set_name:
        return jsonify({'success': False, 'error': 'set_name is required'}), 400
    base_title = (request.form.get('title') or 'Poll').strip() or 'Poll'
    results = {'success': True, 'added': 0, 'errors': []}
    try:
        poll_set = PollBankSet.query.filter_by(class_id=class_id, name=set_name).first()
        if not poll_set:
            poll_set = PollBankSet(class_id=class_id, name=set_name)
            db.session.add(poll_set)
            db.session.flush()
        wb = load_workbook(file, read_only=True, data_only=True)
        ws = wb.active
        poll_rows, parse_errors = _parse_poll_bank_workbook(ws)
        results['errors'].extend(parse_errors)
        if results['errors'] and not poll_rows:
            results['success'] = False
            return jsonify(results), 400
        if not poll_rows:
            results['success'] = False
            results['errors'].append('No poll rows found in the spreadsheet.')
            return jsonify(results), 400

        added = 0
        for row in poll_rows:
            if not row['question'] or len(row['options']) < 2:
                continue
            idx = int(row['poll_index'])
            existing = PollBankQuestion.query.filter_by(class_id=class_id, set_id=poll_set.id, poll_index=idx).first()
            title = base_title if len(poll_rows) == 1 else f'{base_title} {idx}'
            if existing:
                existing.title = title
                existing.question = row['question']
                existing.options = json.dumps(row['options'])
                existing.correct_answer = row['correct_answer']
                existing.source_filename = secure_filename(file.filename) if file.filename else None
            else:
                db.session.add(
                    PollBankQuestion(
                        class_id=class_id,
                            set_id=poll_set.id,
                        poll_index=idx,
                        title=title,
                        question=row['question'],
                        options=json.dumps(row['options']),
                        correct_answer=row['correct_answer'],
                        source_filename=secure_filename(file.filename) if file.filename else None,
                    )
                )
            added += 1
        if added == 0:
            results['success'] = False
            results['errors'].append('No valid poll rows found. Each row needs a question and at least 2 options.')
            return jsonify(results), 400
        db.session.commit()
        results['added'] = added
        results['set_id'] = poll_set.id
        results['set_name'] = poll_set.name
        return jsonify(results)
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/quizzes/<int:class_id>', methods=['GET'])
@login_required
def list_quizzes(class_id):
    class_obj = Class.query.get_or_404(class_id)
    if class_obj.professor_id != current_user.id:
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403
    quizzes = Quiz.query.filter_by(class_id=class_id).order_by(Quiz.quiz_index.asc()).all()
    out = []
    for qz in quizzes:
        nq = QuizQuestion.query.filter_by(quiz_id=qz.id).count()
        out.append({
            'id': qz.id,
            'title': qz.title,
            'quiz_index': qz.quiz_index,
            'time_limit_seconds': qz.time_limit_seconds,
            'question_count': nq,
            'source_filename': qz.source_filename,
        })
    return jsonify({'success': True, 'quizzes': out})


@app.route('/api/quiz_upload/<int:class_id>', methods=['POST'])
@login_required
def quiz_upload(class_id):
    class_obj = Class.query.get_or_404(class_id)
    if class_obj.professor_id != current_user.id:
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403

    if 'file' not in request.files:
        return jsonify({'success': False, 'error': 'No file uploaded'}), 400
    file = request.files['file']
    if not file.filename or not file.filename.lower().endswith(('.xlsx', '.xls')):
        return jsonify({'success': False, 'error': 'Please upload an Excel file (.xlsx)'}), 400

    try:
        quiz_index = int(request.form.get('quiz_index') or request.args.get('quiz_index') or '0')
    except (TypeError, ValueError):
        return jsonify({'success': False, 'error': 'Invalid quiz_index'}), 400
    if quiz_index < 1:
        return jsonify({'success': False, 'error': 'quiz_index must be at least 1'}), 400

    gw = GradingWeights.query.filter_by(class_id=class_id).first()
    n_target = int(getattr(gw, 'quiz_count_target', 0) or 0) if gw else 0
    if n_target < 1:
        return jsonify({
            'success': False,
            'error': (
                'Set the number of quizzes in Class Data → Gradebook → Grading (at least 1) '
                'before uploading a quiz.'
            ),
        }), 400
    if n_target > 0 and quiz_index > n_target:
        return jsonify({
            'success': False,
            'error': f'quiz_index must be between 1 and the configured number of quizzes ({n_target}).',
        }), 400

    title = (request.form.get('title') or f'Quiz {quiz_index}').strip() or f'Quiz {quiz_index}'
    raw_min = request.form.get('time_limit_minutes')
    if raw_min is not None and str(raw_min).strip() != '':
        try:
            time_limit_seconds = int(round(float(raw_min) * 60))
        except (TypeError, ValueError):
            time_limit_seconds = 300
        if time_limit_seconds < 60:
            time_limit_seconds = 60
    else:
        try:
            time_limit_seconds = int(request.form.get('time_limit_seconds') or 300)
        except (TypeError, ValueError):
            time_limit_seconds = 300
        if time_limit_seconds < 30:
            time_limit_seconds = 30
    if time_limit_seconds > 24 * 3600:
        time_limit_seconds = 24 * 3600

    results = {'success': True, 'added': 0, 'errors': []}
    try:
        wb = load_workbook(file, read_only=True, data_only=True)
        ws = wb.active
        questions_to_add, parse_errors = _parse_quiz_workbook(ws)
        results['errors'].extend(parse_errors)

        if results['errors'] and not questions_to_add:
            results['success'] = False
            return jsonify(results), 400
        if not questions_to_add:
            results['success'] = False
            results['errors'].append('No question rows found in the spreadsheet.')
            return jsonify(results), 400

        existing = Quiz.query.filter_by(class_id=class_id, quiz_index=quiz_index).first()
        if existing:
            _delete_quiz_and_related(existing)

        quiz = Quiz(
            class_id=class_id,
            title=title,
            time_limit_seconds=time_limit_seconds,
            quiz_index=quiz_index,
            source_filename=secure_filename(file.filename) if file.filename else None,
        )
        db.session.add(quiz)
        db.session.flush()

        for qd in questions_to_add:
            db.session.add(
                QuizQuestion(
                    quiz_id=quiz.id,
                    order=qd['order'],
                    prompt=qd['prompt'],
                    options=json.dumps(qd['options']),
                    correct_index=qd['correct_index'],
                )
            )
        db.session.commit()
        results['quiz_id'] = quiz.id
        results['added'] = len(questions_to_add)
        return jsonify(results)
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/quiz/<int:quiz_id>/start', methods=['POST'])
@login_required
def start_quiz(quiz_id):
    quiz = Quiz.query.options(joinedload(Quiz.questions)).get_or_404(quiz_id)
    class_obj = Class.query.get_or_404(quiz.class_id)
    if class_obj.professor_id != current_user.id:
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403
    if not class_obj.is_active:
        return jsonify({'success': False, 'error': 'Start the class session before launching a quiz.'}), 400
    if not get_active_class_session(quiz.class_id):
        return jsonify({'success': False, 'error': 'No active class session.'}), 400
    if not quiz.questions:
        return jsonify({'success': False, 'error': 'This quiz has no questions. Upload a spreadsheet first.'}), 400

    gw = GradingWeights.query.filter_by(class_id=quiz.class_id).first()
    n_target = int(getattr(gw, 'quiz_count_target', 0) or 0) if gw else 0
    if n_target < 1:
        return jsonify({
            'success': False,
            'error': (
                'Set the number of quizzes in Class Data → Gradebook → Grading (at least 1) '
                'before starting a quiz.'
            ),
        }), 400

    class_id = quiz.class_id
    poll_ids = deactivate_active_polls_for_class(class_id)
    quiz_run_ids = deactivate_active_quiz_runs_for_class(class_id)

    started = datetime.utcnow()
    deadline = started + timedelta(seconds=int(quiz.time_limit_seconds or 300))
    run = QuizRun(
        quiz_id=quiz.id,
        class_id=class_id,
        started_at=started,
        deadline_at=deadline,
        ended_at=None,
        is_active=True,
    )
    db.session.add(run)
    db.session.commit()

    emit_poll_stopped_events(class_id, poll_ids)
    for qrid in quiz_run_ids:
        socketio.emit(
            'quiz_stopped',
            {'quiz_run_id': qrid, 'class_id': class_id},
            room=f'class_{class_id}',
        )
    socketio.emit('quiz_started', quiz_run_public_payload(run), room=f'class_{class_id}')
    return jsonify({'success': True, 'quiz_run_id': run.id})


@app.route('/api/quiz_run/<int:run_id>/stop', methods=['POST'])
@login_required
def stop_quiz_run(run_id):
    run = QuizRun.query.get_or_404(run_id)
    class_obj = Class.query.get_or_404(run.class_id)
    if class_obj.professor_id != current_user.id:
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403
    now = datetime.utcnow()
    run.is_active = False
    if run.ended_at is None:
        run.ended_at = now
    db.session.commit()
    socketio.emit(
        'quiz_stopped',
        {'quiz_run_id': run.id, 'class_id': run.class_id},
        room=f'class_{run.class_id}',
    )
    return jsonify({'success': True})


@app.route('/api/student/quiz_submit', methods=['POST'])
def student_quiz_submit():
    student_id = _authenticated_student_id()
    if not student_id:
        return jsonify({'success': False, 'error': 'Not logged in'}), 401
    data = request.get_json() or {}
    run_id = data.get('quiz_run_id')
    answers = data.get('answers') or {}
    if not run_id:
        return jsonify({'success': False, 'error': 'quiz_run_id required'}), 400
    try:
        run_id = int(run_id)
    except (TypeError, ValueError):
        return jsonify({'success': False, 'error': 'Invalid quiz_run_id'}), 400

    run = QuizRun.query.options(joinedload(QuizRun.quiz).joinedload(Quiz.questions)).get_or_404(run_id)
    class_id = run.class_id
    if Enrollment.query.filter_by(
        class_id=class_id,
        student_id=student_id,
        is_active=True,
    ).first() is None:
        return jsonify({'success': False, 'error': 'Not enrolled in this class'}), 403

    now = datetime.utcnow()
    if not run.is_active:
        return jsonify({'success': False, 'error': 'This quiz is no longer active.'}), 400
    if now > run.deadline_at:
        return jsonify({'success': False, 'error': 'The time limit for this quiz has passed.'}), 400

    if QuizAnswer.query.filter_by(quiz_run_id=run.id, student_id=student_id).first():
        return jsonify({'success': False, 'error': 'You have already submitted this quiz.'}), 400

    quiz = run.quiz
    q_rows = sorted(quiz.questions, key=lambda q: q.order)
    if len(answers) != len(q_rows):
        return jsonify({
            'success': False,
            'error': f'Expected an answer for each question ({len(q_rows)} total).',
        }), 400

    correct_n = 0
    for q in q_rows:
        key = str(q.id)
        if key not in answers:
            return jsonify({'success': False, 'error': f'Missing answer for question id {q.id}'}), 400
        try:
            sel = int(answers[key])
        except (TypeError, ValueError):
            return jsonify({'success': False, 'error': f'Invalid answer for question {q.id}'}), 400
        try:
            opts = json.loads(q.options or '[]')
        except (json.JSONDecodeError, TypeError):
            opts = []
        if sel < 0 or sel >= len(opts):
            return jsonify({'success': False, 'error': f'Answer out of range for question {q.id}'}), 400
        is_ok = sel == q.correct_index
        if is_ok:
            correct_n += 1
        db.session.add(
            QuizAnswer(
                quiz_run_id=run.id,
                student_id=student_id,
                question_id=q.id,
                selected_index=sel,
                is_correct=is_ok,
                submitted_at=now,
            )
        )
    db.session.commit()
    return jsonify({
        'success': True,
        'correct_count': correct_n,
        'total_questions': len(q_rows),
    })


@app.route('/api/create_class', methods=['POST'])
@login_required
def create_class():
    data = request.get_json()
    name = data.get('name')
    class_code = data.get('class_code')
    
    if not name or not class_code:
        return jsonify({'success': False, 'error': 'Name and class code required'})
    
    if Class.query.filter_by(class_code=class_code).first():
        return jsonify({'success': False, 'error': 'Class code already exists'})
    
    class_obj = Class(
        professor_id=current_user.id,
        name=name,
        class_code=class_code
    )
    db.session.add(class_obj)
    db.session.commit()
    
    return jsonify({'success': True, 'class_id': class_obj.id})

@app.route('/api/delete_class/<int:class_id>', methods=['DELETE'])
@login_required
def delete_class(class_id):
    class_obj = Class.query.get_or_404(class_id)
    if class_obj.professor_id != current_user.id:
        return jsonify({'success': False, 'error': 'Unauthorized'})
    
    try:
        # Delete all related records
        # First, delete poll responses for polls in this class
        polls = Poll.query.filter_by(class_id=class_id).all()
        for poll in polls:
            PollResponse.query.filter_by(poll_id=poll.id).delete()
        
        # Delete polls
        Poll.query.filter_by(class_id=class_id).delete()

        for qz in Quiz.query.filter_by(class_id=class_id).all():
            _delete_quiz_and_related(qz)
        
        # Delete participations
        Participation.query.filter_by(class_id=class_id).delete()
        
        # Delete attendances
        Attendance.query.filter_by(class_id=class_id).delete()
        
        # Delete enrollments
        Enrollment.query.filter_by(class_id=class_id).delete()
        
        # Delete class settings
        ClassSettings.query.filter_by(class_id=class_id).delete()

        GradingWeights.query.filter_by(class_id=class_id).delete()
        
        # Finally, delete the class
        db.session.delete(class_obj)
        db.session.commit()
        
        return jsonify({'success': True})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/add_student_to_class', methods=['POST'])
@login_required
def add_student_to_class():
    data = request.get_json()
    class_id = data.get('class_id')
    student_id = data.get('student_id')
    
    class_obj = Class.query.get_or_404(class_id)
    if class_obj.professor_id != current_user.id:
        return jsonify({'success': False, 'error': 'Unauthorized'})
    
    enrollment = Enrollment.query.filter_by(
        class_id=class_id,
        student_id=student_id
    ).first()
    
    if enrollment:
        return jsonify({'success': False, 'error': 'Student already enrolled'})
    
    enrollment = Enrollment(class_id=class_id, student_id=student_id, is_active=True)
    db.session.add(enrollment)
    db.session.commit()
    
    return jsonify({'success': True})

@app.route('/api/create_and_add_student/<int:class_id>', methods=['POST'])
@login_required
def create_and_add_student(class_id):
    """Create a new student and add them to the class"""
    class_obj = Class.query.get_or_404(class_id)
    if class_obj.professor_id != current_user.id:
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403
    
    data = request.get_json()
    
    # Safely extract and strip fields, handling None values
    first_name = (data.get('first_name') or '').strip() if data.get('first_name') is not None else ''
    last_name = (data.get('last_name') or '').strip() if data.get('last_name') is not None else ''
    preferred_name_raw = data.get('preferred_name')
    preferred_name = preferred_name_raw.strip() if preferred_name_raw else None
    student_number = (data.get('student_number') or '').strip() if data.get('student_number') is not None else ''
    email = (data.get('email') or '').strip() if data.get('email') is not None else ''
    rfid_card_id_raw = data.get('rfid_card_id')
    rfid_card_id = rfid_card_id_raw.strip() if rfid_card_id_raw else None
    
    # Validate required fields
    if not first_name or not last_name or not student_number or not email:
        return jsonify({'success': False, 'error': 'First name, last name, student number, and email are required'}), 400
    
    # Validate student number format (exactly 9 digits, no more, no less)
    if not student_number.isdigit():
        return jsonify({'success': False, 'error': 'Student number must contain only digits'}), 400
    
    if len(student_number) != 9:
        return jsonify({'success': False, 'error': f'Student number must be exactly 9 digits. You entered {len(student_number)} digit(s).'}), 400
    
    # Validate email format
    if '@' not in email:
        return jsonify({'success': False, 'error': 'Invalid email format'}), 400
    
    # Check if student already exists by student_number OR email (to ensure one account per student)
    existing_student = Student.query.filter_by(student_number=student_number).first()
    if not existing_student:
        # Also check by email to ensure uniqueness across classes
        existing_student = Student.query.filter_by(email=email).first()
    
    if existing_student:
        # Student exists, just enroll them if not already enrolled
        enrollment = Enrollment.query.filter_by(
            class_id=class_id,
            student_id=existing_student.id
        ).first()
        
        if enrollment:
            return jsonify({'success': False, 'error': 'Student is already enrolled in this class'}), 400
        
        # Update student info if provided (but don't change student_number or email if they differ)
        if existing_student.first_name != first_name:
            existing_student.first_name = first_name
        if existing_student.last_name != last_name:
            existing_student.last_name = last_name
        if existing_student.preferred_name != preferred_name:
            existing_student.preferred_name = preferred_name
        # Only update email if it matches (to maintain account uniqueness)
        if existing_student.email != email and existing_student.student_number == student_number:
            existing_student.email = email
        if rfid_card_id and existing_student.rfid_card_id != rfid_card_id:
            # Check if RFID is already taken by another student
            rfid_student = Student.query.filter_by(rfid_card_id=rfid_card_id).first()
            if rfid_student and rfid_student.id != existing_student.id:
                return jsonify({'success': False, 'error': 'RFID card ID already in use by another student'}), 400
            existing_student.rfid_card_id = rfid_card_id
        
        student = existing_student
    else:
        # Create new student - check if student_number or email conflicts
        # Check if student number is already taken
        if Student.query.filter_by(student_number=student_number).first():
            return jsonify({'success': False, 'error': 'Student number already exists'}), 400
        # Check if email is already taken
        if Student.query.filter_by(email=email).first():
            return jsonify({'success': False, 'error': 'Email already exists'}), 400
        # Check if RFID is already taken
        if rfid_card_id and Student.query.filter_by(rfid_card_id=rfid_card_id).first():
            return jsonify({'success': False, 'error': 'RFID card ID already in use'}), 400
        
        student = Student(
            student_number=student_number,
            first_name=first_name,
            last_name=last_name,
            preferred_name=preferred_name,
            email=email,
            rfid_card_id=rfid_card_id
        )
        db.session.add(student)
        db.session.flush()
    
    # Enroll student in class
    enrollment = Enrollment(class_id=class_id, student_id=student.id, is_active=True)
    db.session.add(enrollment)
    db.session.commit()
    
    return jsonify({
        'success': True,
        'student': {
            'id': student.id,
            'student_number': student.student_number,
            'first_name': student.first_name,
            'last_name': student.last_name,
            'preferred_name': student.preferred_name,
            'email': student.email
        }
    })

@app.route('/api/toggle_student_status/<int:class_id>/<int:student_id>', methods=['POST'])
@login_required
def toggle_student_status(class_id, student_id):
    class_obj = Class.query.get_or_404(class_id)
    if class_obj.professor_id != current_user.id:
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403
    
    enrollment = Enrollment.query.filter_by(
        class_id=class_id,
        student_id=student_id
    ).first()
    
    if not enrollment:
        return jsonify({'success': False, 'error': 'Student not enrolled in this class'}), 404
    
    # Toggle the status
    enrollment.is_active = not enrollment.is_active
    db.session.commit()
    
    return jsonify({'success': True, 'is_active': enrollment.is_active})

@app.route('/api/update_student/<int:student_id>', methods=['POST'])
@login_required
def update_student(student_id):
    student = Student.query.get_or_404(student_id)
    
    # Check if any class the student is enrolled in belongs to current user
    enrollments = Enrollment.query.filter_by(student_id=student_id).all()
    if not enrollments:
        return jsonify({'success': False, 'error': 'Student not found in any of your classes'}), 404
    
    # Check if student is enrolled in at least one class owned by current user
    has_access = any(e.class_obj.professor_id == current_user.id for e in enrollments)
    if not has_access:
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403
    
    data = request.get_json()
    
    first_name = data.get('first_name', '').strip()
    last_name = data.get('last_name', '').strip()
    preferred_name = data.get('preferred_name', '').strip() or None
    student_number = data.get('student_number', '').strip()
    email = data.get('email', '').strip()
    
    # Validate required fields
    if not first_name or not last_name or not student_number or not email:
        return jsonify({'success': False, 'error': 'First name, last name, student number, and email are required'}), 400
    
    # Validate student number format (9 digits)
    if len(student_number) != 9 or not student_number.isdigit():
        return jsonify({'success': False, 'error': 'Student number must be exactly 9 digits'}), 400
    
    # Validate email format
    if '@' not in email:
        return jsonify({'success': False, 'error': 'Invalid email format'}), 400
    
    # Check if student number is already taken by another student
    existing_student = Student.query.filter_by(student_number=student_number).first()
    if existing_student and existing_student.id != student_id:
        return jsonify({'success': False, 'error': 'Student number already exists'}), 400
    
    # Update student information
    student.first_name = first_name
    student.last_name = last_name
    student.preferred_name = preferred_name
    student.student_number = student_number
    student.email = email
    
    db.session.commit()
    
    return jsonify({'success': True})

@app.route('/api/remove_student_from_class/<int:class_id>/<int:student_id>', methods=['DELETE'])
@login_required
def remove_student_from_class(class_id, student_id):
    class_obj = Class.query.get_or_404(class_id)
    if class_obj.professor_id != current_user.id:
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403
    
    student = Student.query.get_or_404(student_id)
    
    # Find enrollment
    enrollment = Enrollment.query.filter_by(
        class_id=class_id,
        student_id=student_id
    ).first()
    
    if not enrollment:
        return jsonify({'success': False, 'error': 'Student not enrolled in this class'}), 404
    
    # Delete all class-related data for this student in this class
    # Delete attendance records
    Attendance.query.filter_by(
        class_id=class_id,
        student_id=student_id
    ).delete()
    
    # Delete participation records
    Participation.query.filter_by(
        class_id=class_id,
        student_id=student_id
    ).delete()
    
    # Delete poll responses for polls in this class
    polls = Poll.query.filter_by(class_id=class_id).all()
    poll_ids = [poll.id for poll in polls]
    if poll_ids:
        PollResponse.query.filter(
            PollResponse.student_id == student_id,
            PollResponse.poll_id.in_(poll_ids)
        ).delete()
    
    # Delete enrollment
    db.session.delete(enrollment)
    
    db.session.commit()
    
    return jsonify({'success': True})

@app.route('/api/download_student_template/<int:class_id>')
@login_required
def download_student_template(class_id):
    class_obj = Class.query.get_or_404(class_id)
    if class_obj.professor_id != current_user.id:
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403
    
    # Create a new workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Students"
    
    # Define the header row with exact format (5 columns)
    # Column order: A=First Name, B=Last Name, C=Preferred Name, D=Student Number, E=Email
    headers = ['Student First Name', 'Last Name', 'Student Preferred Name', 'Student Number', 'Email']
    ws.append(headers)
    
    # Style the header row
    header_fill = PatternFill(start_color="2A1A40", end_color="2A1A40", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    
    # Add one example row with 9-digit student number (Student Number in column D, Email in column E)
    # Column order: First Name, Last Name, Preferred Name, Student Number, Email
    example_row = ['John', 'Doe', 'Johnny', '123456789', 'john.doe@example.com']
    ws.append(example_row)
    
    # Auto-fit column widths based on content
    from openpyxl.utils import get_column_letter
    
    # Calculate column widths based on header and example content
    # Excel column width units are approximately equal to the width of one character
    # We add extra padding for better readability
    # Columns: A=First Name, B=Last Name, C=Preferred Name, D=Student Number, E=Email
    columns_data = [
        ('A', 'Student First Name', example_row[0]),
        ('B', 'Last Name', example_row[1]),
        ('C', 'Student Preferred Name', example_row[2]),
        ('D', 'Student Number', example_row[3]),
        ('E', 'Email', example_row[4])
    ]
    
    for col_letter, header_text, example_text in columns_data:
        # Calculate width based on the longest content (header or example)
        # Multiply by 1.2 for padding and convert to Excel width units
        max_content_length = max(len(header_text), len(str(example_text)))
        # Excel width calculation: add padding and ensure minimum readable width
        column_width = max(max_content_length * 1.15 + 2, 12)
        # Cap maximum width at 50 to prevent extremely wide columns
        column_width = min(column_width, 50)
        ws.column_dimensions[col_letter].width = column_width
    
    # Create BytesIO object to store the Excel file in memory
    excel_buffer = BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    return send_file(
        excel_buffer,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'student_template_{class_obj.name.replace(" ", "_")}.xlsx'
    )

@app.route('/api/export_students/<int:class_id>')
@login_required
def export_students(class_id):
    class_obj = Class.query.get_or_404(class_id)
    if class_obj.professor_id != current_user.id:
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403
    
    # Get active students only (matching template format)
    active_students = db.session.query(Student).join(Enrollment).filter(
        Enrollment.class_id == class_id,
        Enrollment.is_active == True
    ).order_by(Student.last_name, Student.first_name).all()
    
    # Create a new workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Students"  # Match template sheet name
    
    # Define the header row with exact format (5 columns) - matching template exactly
    # Column order: A=First Name, B=Last Name, C=Preferred Name, D=Student Number, E=Email
    headers = ['Student First Name', 'Last Name', 'Student Preferred Name', 'Student Number', 'Email']
    ws.append(headers)
    
    # Style the header row - matching template styling
    header_fill = PatternFill(start_color="2A1A40", end_color="2A1A40", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    
    # Add active student rows with all 5 columns
    for student in active_students:
        row = [
            student.first_name,
            student.last_name,
            student.preferred_name if student.preferred_name else '',
            student.student_number,
            student.email if hasattr(student, 'email') and student.email else ''
        ]
        ws.append(row)
    
    # Auto-fit column widths - matching template calculation method
    from openpyxl.utils import get_column_letter
    
    if active_students:
        # Calculate max lengths including headers
        max_first_name = max([len(student.first_name) for student in active_students] + [len('Student First Name')])
        max_last_name = max([len(student.last_name) for student in active_students] + [len('Last Name')])
        max_preferred_name = max([len(student.preferred_name or '') for student in active_students] + [len('Student Preferred Name')])
        max_student_number = max([len(str(student.student_number)) for student in active_students] + [len('Student Number')])
        max_email = max([len(getattr(student, 'email', '') or '') for student in active_students] + [len('Email')])
    else:
        # If no students, use header lengths
        max_first_name = len('Student First Name')
        max_last_name = len('Last Name')
        max_preferred_name = len('Student Preferred Name')
        max_student_number = len('Student Number')
        max_email = len('Email')
    
    # Columns: A=First Name, B=Last Name, C=Preferred Name, D=Student Number, E=Email
    columns_data = [
        ('A', 'Student First Name', max_first_name),
        ('B', 'Last Name', max_last_name),
        ('C', 'Student Preferred Name', max_preferred_name),
        ('D', 'Student Number', max_student_number),
        ('E', 'Email', max_email)
    ]
    
    for col_letter, header_text, max_content_length in columns_data:
        # Calculate width based on the longest content (header or data)
        # Multiply by 1.15 for padding and add 2 for extra spacing
        column_width = max(max_content_length * 1.15 + 2, 12)
        # Cap maximum width at 50 to prevent extremely wide columns
        column_width = min(column_width, 50)
        ws.column_dimensions[col_letter].width = column_width
    
    # Create BytesIO object to store the Excel file in memory
    excel_buffer = BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    return send_file(
        excel_buffer,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'class_list_{class_obj.name.replace(" ", "_")}_{datetime.now().strftime("%Y%m%d")}.xlsx'
    )

@app.route('/api/upload_students/<int:class_id>', methods=['POST'])
@login_required
def upload_students(class_id):
    class_obj = Class.query.get_or_404(class_id)
    if class_obj.professor_id != current_user.id:
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403
    
    if 'file' not in request.files:
        return jsonify({'success': False, 'error': 'No file uploaded'}), 400
    
    file = request.files['file']
    
    if file.filename == '':
        return jsonify({'success': False, 'error': 'No file selected'}), 400
    
    # Check if file is Excel
    if not file.filename.endswith(('.xlsx', '.xls')):
        return jsonify({'success': False, 'error': 'Invalid file type. Please upload an Excel file (.xlsx or .xls)'}), 400
    
    try:
        # Load the workbook
        wb = load_workbook(file, read_only=True, data_only=True)
        
        # Expected headers (5 columns) - Column order: First Name, Last Name, Preferred Name, Student Number, Email
        expected_headers = ['Student First Name', 'Last Name', 'Student Preferred Name', 'Student Number', 'Email']
        
        # Try to find "Active Students" sheet first, then fall back to active sheet
        ws_active = None
        ws_inactive = None
        
        if 'Active Students' in wb.sheetnames:
            ws_active = wb['Active Students']
        else:
            # Fall back to active sheet (for backward compatibility with single-sheet uploads)
            ws_active = wb.active
        
        # If "Inactive Students" sheet exists, use it (optional)
        if 'Inactive Students' in wb.sheetnames:
            ws_inactive = wb['Inactive Students']
        
        # Validate active sheet (required)
        if ws_active:
            header_row = []
            for cell in ws_active[1]:
                if cell.value is None:
                    header_row.append('')
                else:
                    header_row.append(str(cell.value).strip())
            
            header_row = header_row[:5]
            expected_normalized = [h.strip() for h in expected_headers]
            
            if header_row != expected_normalized:
                return jsonify({
                    'success': False, 
                    'error': f'Invalid file format. The "Active Students" sheet first row must match exactly: {expected_headers}. Your file has: {header_row}. Please download the template and use it exactly as provided (including exact column names and order).'
                }), 400
        
        # Validate inactive sheet if it exists (optional)
        if ws_inactive:
            header_row = []
            for cell in ws_inactive[1]:
                if cell.value is None:
                    header_row.append('')
                else:
                    header_row.append(str(cell.value).strip())
            
            header_row = header_row[:5]
            expected_normalized = [h.strip() for h in expected_headers]
            
            if header_row != expected_normalized:
                return jsonify({
                    'success': False, 
                    'error': f'Invalid file format. The "Inactive Students" sheet first row must match exactly: {expected_headers}. Your file has: {header_row}. Please download the template and use it exactly as provided (including exact column names and order).'
                }), 400
        
        # Process data rows (skip header row)
        results = {
            'success': True,
            'added': 0,
            'updated': 0,
            'activated': 0,
            'deactivated': 0,
            'skipped': 0,
            'errors': []
        }
        
        # Track all student numbers processed in this upload to handle duplicates within the upload
        processed_student_numbers_in_upload = set()
        
        # Get all existing student numbers in the database to avoid duplicate checks
        existing_student_numbers_db = {s.student_number for s in Student.query.all()}
        
        # Process Active Students sheet
        if ws_active:
            for row_idx, row in enumerate(ws_active.iter_rows(min_row=2, values_only=False), start=2):
                # Rule 1: Check that data only appears in columns A, B, C, D, E
                has_extra_data = False
                if len(row) > 5:
                    for cell in row[5:]:
                        if cell.value is not None and str(cell.value).strip():
                            has_extra_data = True
                            break
                
                if has_extra_data:
                    results['skipped'] += 1
                    results['errors'].append(f'Active Sheet Row {row_idx}: Data found beyond columns A-E. All data must be in columns A, B, C, D, and E only.')
                    continue
                
                # Extract values from first 5 columns (A, B, C, D, E)
                cell_values = []
                for cell in row[:5]:
                    if cell.value is None:
                        cell_values.append(None)
                    else:
                        raw_value = cell.value
                        cell_values.append(str(raw_value).strip() if raw_value else None)
                
                # Skip if all cells in first 5 columns are empty
                if not any(cell_values):
                    continue
                
                # Extract values based on column order:
                # A=First Name, B=Last Name, C=Preferred Name, D=Student Number, E=Email
                first_name = cell_values[0] if cell_values[0] else None
                last_name = cell_values[1] if cell_values[1] else None
                preferred_name_raw = cell_values[2] if cell_values[2] else None
                student_number_raw = cell_values[3] if cell_values[3] else None
                email_raw = cell_values[4] if cell_values[4] else None
                
                # Rule 5: Extract only digits from student number (ignore formatting)
                if student_number_raw:
                    student_number = ''.join(re.findall(r'\d', student_number_raw))
                else:
                    student_number = None
                
                # Preferred name: extract text
                if preferred_name_raw and preferred_name_raw.strip():
                    preferred_name = preferred_name_raw.strip()
                else:
                    preferred_name = None
                
                # Validate required fields
                if not student_number or not first_name or not last_name or not email_raw:
                    results['skipped'] += 1
                    results['errors'].append(f'Active Sheet Row {row_idx}: Missing required fields (Student Number, Student First Name, Last Name, or Email)')
                    continue
                
                # Validate email format
                email = email_raw.strip() if email_raw else None
                if not email or '@' not in email:
                    results['skipped'] += 1
                    results['errors'].append(f'Active Sheet Row {row_idx}: Invalid email format: "{email_raw}"')
                    continue
                
                # Rule 2: Validate Student Number is exactly 9 digits
                if len(student_number) != 9:
                    results['skipped'] += 1
                    results['errors'].append(f'Active Sheet Row {row_idx}: Student Number must be exactly 9 digits. Found {len(student_number)} digit(s) in: "{student_number_raw}"')
                    continue
                
                # Check for duplicates within this upload (skip if already processed in this upload)
                if student_number in processed_student_numbers_in_upload:
                    results['skipped'] += 1
                    results['errors'].append(f'Active Sheet Row {row_idx}: Duplicate student number "{student_number}" found in this upload. Only the first occurrence will be processed.')
                    continue
                
                # Mark this student number as processed in this upload
                processed_student_numbers_in_upload.add(student_number)
                
                # Normalize names
                first_name = first_name.strip() if first_name else None
                last_name = last_name.strip() if last_name else None
                
                # Check if student exists by student_number OR email (to ensure one account per student)
                existing_student = Student.query.filter_by(student_number=student_number).first()
                if not existing_student:
                    # Also check by email to ensure uniqueness across classes
                    existing_student = Student.query.filter_by(email=email).first()
                
                if existing_student:
                    # Update existing student if needed (but maintain account uniqueness)
                    updated = False
                    if existing_student.first_name != first_name:
                        existing_student.first_name = first_name
                        updated = True
                    if existing_student.last_name != last_name:
                        existing_student.last_name = last_name
                        updated = True
                    if existing_student.preferred_name != preferred_name:
                        existing_student.preferred_name = preferred_name
                        updated = True
                    # Only update email if student_number matches (to maintain account uniqueness)
                    if existing_student.email != email and existing_student.student_number == student_number:
                        existing_student.email = email
                        updated = True
                    
                    if updated:
                        results['updated'] += 1
                else:
                    # Create new student
                    new_student = Student(
                        student_number=student_number,
                        first_name=first_name,
                        preferred_name=preferred_name,
                        last_name=last_name,
                        email=email,
                        rfid_card_id=None
                    )
                    db.session.add(new_student)
                    db.session.flush()
                    existing_student = new_student
                    results['added'] += 1
                
                # Check if student is already enrolled in this class
                enrollment = Enrollment.query.filter_by(
                    class_id=class_id,
                    student_id=existing_student.id
                ).first()
                
                if not enrollment:
                    # Only create enrollment if student is not already enrolled
                    enrollment = Enrollment(class_id=class_id, student_id=existing_student.id, is_active=True)
                    db.session.add(enrollment)
                else:
                    # If re-enrolling an inactive student, activate them
                    if not enrollment.is_active:
                        enrollment.is_active = True
                        results['activated'] += 1
                    # If already enrolled and active, no action needed (not a duplicate error, just skip)
        
        # Process Inactive Students sheet (optional)
        if ws_inactive:
            for row_idx, row in enumerate(ws_inactive.iter_rows(min_row=2, values_only=False), start=2):
                # Rule 1: Check that data only appears in columns A, B, C, D, E
                has_extra_data = False
                if len(row) > 5:
                    for cell in row[5:]:
                        if cell.value is not None and str(cell.value).strip():
                            has_extra_data = True
                            break
                
                if has_extra_data:
                    results['skipped'] += 1
                    results['errors'].append(f'Inactive Sheet Row {row_idx}: Data found beyond columns A-E. All data must be in columns A, B, C, D, and E only.')
                    continue
                
                # Extract values from first 5 columns (A, B, C, D, E)
                cell_values = []
                for cell in row[:5]:
                    if cell.value is None:
                        cell_values.append(None)
                    else:
                        raw_value = cell.value
                        cell_values.append(str(raw_value).strip() if raw_value else None)
                
                # Skip if all cells in first 5 columns are empty
                if not any(cell_values):
                    continue
                
                # Extract values based on column order
                # A=First Name, B=Last Name, C=Preferred Name, D=Student Number, E=Email
                first_name = cell_values[0] if cell_values[0] else None
                last_name = cell_values[1] if cell_values[1] else None
                preferred_name_raw = cell_values[2] if cell_values[2] else None
                student_number_raw = cell_values[3] if cell_values[3] else None
                email_raw = cell_values[4] if cell_values[4] else None
                
                # Rule 5: Extract only digits from student number
                if student_number_raw:
                    student_number = ''.join(re.findall(r'\d', student_number_raw))
                else:
                    student_number = None
                
                # Skip if already processed in active sheet (active takes precedence)
                if student_number and student_number in processed_student_numbers_in_upload:
                    continue  # Skip - already processed in active sheet
                
                # Check for duplicates within inactive sheet (skip if already processed in this upload's inactive section)
                if student_number and student_number in processed_student_numbers_in_upload:
                    results['skipped'] += 1
                    results['errors'].append(f'Inactive Sheet Row {row_idx}: Duplicate student number "{student_number}" found in this upload. Only the first occurrence will be processed.')
                    continue
                
                # Preferred name: extract text
                if preferred_name_raw and preferred_name_raw.strip():
                    preferred_name = preferred_name_raw.strip()
                else:
                    preferred_name = None
                
                # Validate required fields
                if not student_number or not first_name or not last_name or not email_raw:
                    results['skipped'] += 1
                    results['errors'].append(f'Inactive Sheet Row {row_idx}: Missing required fields (Student Number, Student First Name, Last Name, or Email)')
                    continue
                
                # Validate email format
                email = email_raw.strip() if email_raw else None
                if not email or '@' not in email:
                    results['skipped'] += 1
                    results['errors'].append(f'Inactive Sheet Row {row_idx}: Invalid email format: "{email_raw}"')
                    continue
                
                # Rule 2: Validate Student Number is exactly 9 digits
                if len(student_number) != 9:
                    results['skipped'] += 1
                    results['errors'].append(f'Inactive Sheet Row {row_idx}: Student Number must be exactly 9 digits. Found {len(student_number)} digit(s) in: "{student_number_raw}"')
                    continue
                
                # Mark this student number as processed in this upload (for inactive sheet)
                processed_student_numbers_in_upload.add(student_number)
                
                # Normalize names
                first_name = first_name.strip() if first_name else None
                last_name = last_name.strip() if last_name else None
                
                # Check if student exists by student_number OR email (to ensure one account per student)
                existing_student = Student.query.filter_by(student_number=student_number).first()
                if not existing_student:
                    # Also check by email to ensure uniqueness across classes
                    existing_student = Student.query.filter_by(email=email).first()
                
                if existing_student:
                    # Update existing student if needed (but maintain account uniqueness)
                    updated = False
                    if existing_student.first_name != first_name:
                        existing_student.first_name = first_name
                        updated = True
                    if existing_student.last_name != last_name:
                        existing_student.last_name = last_name
                        updated = True
                    if existing_student.preferred_name != preferred_name:
                        existing_student.preferred_name = preferred_name
                        updated = True
                    # Only update email if student_number matches (to maintain account uniqueness)
                    if existing_student.email != email and existing_student.student_number == student_number:
                        existing_student.email = email
                        updated = True
                    
                    if updated:
                        results['updated'] += 1
                else:
                    # Create new student
                    new_student = Student(
                        student_number=student_number,
                        first_name=first_name,
                        preferred_name=preferred_name,
                        last_name=last_name,
                        email=email,
                        rfid_card_id=None
                    )
                    db.session.add(new_student)
                    db.session.flush()
                    existing_student = new_student
                    results['added'] += 1
                
                # Check if student is already enrolled in this class
                enrollment = Enrollment.query.filter_by(
                    class_id=class_id,
                    student_id=existing_student.id
                ).first()
                
                if not enrollment:
                    # Only create enrollment if student is not already enrolled
                    enrollment = Enrollment(class_id=class_id, student_id=existing_student.id, is_active=False)
                    db.session.add(enrollment)
                else:
                    # Mark as inactive if currently active
                    if enrollment.is_active:
                        enrollment.is_active = False
                        results['deactivated'] += 1
                    # If already enrolled and inactive, no action needed (not a duplicate error, just skip)
        
        # Commit all changes
        db.session.commit()
        
        return jsonify(results)
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': f'Error processing file: {str(e)}'}), 500

# Student routes
@app.route('/student')
def student_interface():
    html = render_template('student_interface.html')
    resp = make_response(html)
    # Prevent proxies/browsers from serving an old student UI after deploy
    resp.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
    resp.headers['Pragma'] = 'no-cache'
    resp.headers['Expires'] = '0'
    return resp

def student_on_active_roster(student_id):
    """Student must appear on at least one class roster with active enrollment."""
    return Enrollment.query.filter_by(student_id=student_id, is_active=True).first() is not None


def _student_login_payload(stu):
    return {
        'id': stu.id,
        'student_number': stu.student_number,
        'first_name': stu.first_name,
        'last_name': stu.last_name,
        'preferred_name': stu.preferred_name,
        'email': stu.email,
    }


def _find_student_by_identifier(identifier):
    """Resolve student by 9-digit student number or email."""
    if not identifier or not str(identifier).strip():
        return None
    s = str(identifier).strip()
    if s.isdigit() and len(s) == 9:
        return Student.query.filter_by(student_number=s).first()
    return Student.query.filter_by(email=s).first()


@app.route('/api/student/login', methods=['POST'])
def student_login():
    data = request.get_json() or {}
    rfid_card_id = data.get('rfid_card_id')
    identifier = (data.get('identifier') or data.get('email') or '').strip()
    password = data.get('password') or ''

    student = None

    # RFID (hardware nameplates)
    if rfid_card_id:
        student = Student.query.filter_by(rfid_card_id=rfid_card_id.strip()).first()
        if student:
            if not student_on_active_roster(student.id):
                return jsonify({'success': False, 'error': 'You are not registered in any class. Please contact your professor.'})
            if not student.password_hash:
                return jsonify({
                    'success': True,
                    'needs_password': True,
                    'student': _student_login_payload(student),
                    'token': issue_student_token(student.id, needs_password=True),
                })
            return jsonify({
                'success': True,
                'needs_password': False,
                'student': {k: v for k, v in _student_login_payload(student).items() if k != 'email'},
                'token': issue_student_token(student.id, needs_password=False),
            })

    # Student number or email + password
    if identifier:
        student = _find_student_by_identifier(identifier)
        if not student:
            return jsonify({'success': False, 'error': 'Invalid student number, email, or password.'})
        if not student_on_active_roster(student.id):
            return jsonify({'success': False, 'error': 'You are not registered in any class. Please contact your professor.'})
        if not student.password_hash:
            return jsonify({
                'success': True,
                'needs_password': True,
                'student': _student_login_payload(student),
                'token': issue_student_token(student.id, needs_password=True),
            })
        if not password:
            return jsonify({'success': False, 'error': 'Please enter your password.'})
        if check_password_hash(student.password_hash, password):
            return jsonify({
                'success': True,
                'needs_password': False,
                'student': {k: v for k, v in _student_login_payload(student).items() if k != 'email'},
                'token': issue_student_token(student.id, needs_password=False),
            })
        return jsonify({'success': False, 'error': 'Invalid student number, email, or password.'})

    return jsonify({'success': False, 'error': 'Please enter your student number or email and password.'})

@app.route('/api/student/find_for_password', methods=['POST'])
def find_student_for_password():
    """Find student by student number or email for password setup"""
    data = request.get_json()
    identifier = data.get('identifier', '').strip()
    
    if not identifier:
        return jsonify({'success': False, 'error': 'Please enter your student number or email'})
    
    # Try to find by student number (9 digits) or email
    student = None
    if identifier.isdigit() and len(identifier) == 9:
        student = Student.query.filter_by(student_number=identifier).first()
    else:
        student = Student.query.filter_by(email=identifier).first()
    
    if not student:
        return jsonify({'success': False, 'error': 'Student not found. Please make sure you are registered in a class by your professor.'})
    
    if not student_on_active_roster(student.id):
        return jsonify({'success': False, 'error': 'You are not registered in any class. Please contact your professor.'})
    
    return jsonify({
        'success': True,
        'student': {
            'id': student.id,
            'student_number': student.student_number,
            'first_name': student.first_name,
            'last_name': student.last_name,
            'email': student.email,
            'has_password': student.password_hash is not None
        },
        'token': issue_student_token(student.id, needs_password=student.password_hash is None),
    })

@app.route('/api/student/set_password', methods=['POST'])
def student_set_password():
    """Set password for student on first login"""
    student_id = _authenticated_student_id()
    if not student_id:
        return jsonify({'success': False, 'error': 'Not authenticated. Please find your account first using the "Set Password / Register" option.'})
    
    data = request.get_json()
    password = data.get('password')
    confirm_password = data.get('confirm_password')
    
    if not password or not confirm_password:
        return jsonify({'success': False, 'error': 'Password and confirmation are required'})
    
    if password != confirm_password:
        return jsonify({'success': False, 'error': 'Passwords do not match'})
    
    if len(password) < 6:
        return jsonify({'success': False, 'error': 'Password must be at least 6 characters long'})
    
    student = Student.query.get(student_id)
    if not student:
        return jsonify({'success': False, 'error': 'Student not found'})
    
    if not student_on_active_roster(student.id):
        return jsonify({'success': False, 'error': 'You are not registered in any class. Please contact your professor.'})
    
    # Set the password
    student.password_hash = generate_password_hash(password)
    db.session.commit()
    
    return jsonify({
        'success': True,
        'message': 'Password set successfully',
        'student': {
            'id': student.id,
            'student_number': student.student_number,
            'first_name': student.first_name,
            'last_name': student.last_name,
            'preferred_name': student.preferred_name,
        },
        'token': issue_student_token(student.id, needs_password=False),
    })

@app.route('/api/student/current', methods=['GET'])
def get_current_student():
    """Get current logged-in student info"""
    student_id = _authenticated_student_id()
    if not student_id:
        return jsonify({'success': False, 'error': 'Not logged in'}), 401
    
    student = Student.query.get(student_id)
    if not student:
        return jsonify({'success': False, 'error': 'Student not found'}), 404
    
    return jsonify({
        'success': True,
        'student': {
            'id': student.id,
            'student_number': student.student_number,
            'first_name': student.first_name,
            'last_name': student.last_name,
            'preferred_name': student.preferred_name,
            'email': student.email
        }
    })

@app.route('/api/student/classes')
def get_active_classes():
    """All classes where student is actively enrolled. Includes live and not-live sessions."""
    student_id = _authenticated_student_id()
    if not student_id:
        return jsonify({'success': False, 'error': 'Not logged in'}), 401

    rows = db.session.query(Class).join(Enrollment).filter(
        Enrollment.student_id == student_id,
        Enrollment.is_active == True,
    ).order_by(Class.name).all()

    out = []
    for c in rows:
        settings = ClassSettings.query.filter_by(class_id=c.id).first()
        show_fn = bool(settings and settings.show_first_name_only)
        fn_labels = _first_name_only_labels_for_class(c.id) if show_fn else {}
        row = {
            'id': c.id,
            'name': c.name,
            'class_code': c.class_code,
            'is_active': c.is_active,
            'show_first_name_only': show_fn,
            'quiet_mode': bool(settings and settings.quiet_mode),
        }
        if show_fn:
            row['first_name_only_display'] = fn_labels.get(student_id, '')
        out.append(row)
    return jsonify(out)

@app.route('/api/student/join_class', methods=['POST'])
def student_join_class():
    student_id = _authenticated_student_id()
    if not student_id:
        return jsonify({'success': False, 'error': 'Not logged in'})
    
    data = request.get_json()
    class_id = data.get('class_id')
    
    class_obj = Class.query.get_or_404(class_id)
    if not class_obj.is_active:
        return jsonify({'success': False, 'error': 'Class is not active'})

    enrollment = Enrollment.query.filter_by(
        class_id=class_id,
        student_id=student_id,
        is_active=True,
    ).first()
    if not enrollment:
        return jsonify({'success': False, 'error': 'You are not enrolled in this class.'})
    
    active_session = get_active_class_session(class_id)
    if not active_session:
        return jsonify({'success': False, 'error': 'No active class session. Ask your instructor to start the class.'})

    join_time = datetime.utcnow()
    session_date = active_session.start_time.date()
    attendance = Attendance.query.filter_by(
        class_id=class_id,
        student_id=student_id,
        class_session_id=active_session.id,
    ).first()

    if not attendance:
        attendance = Attendance(
            class_id=class_id,
            student_id=student_id,
            class_session_id=active_session.id,
            date=session_date,
            present=True,
            join_time=join_time,
            leave_time=None,
        )
        db.session.add(attendance)
    else:
        attendance.present = True
        attendance.leave_time = None
        if not attendance.join_time:
            attendance.join_time = join_time

    db.session.commit()
    
    socketio.emit('student_joined', {
        'student_id': student_id,
        'class_id': class_id
    }, room=f'class_{class_id}')
    
    class_settings = ClassSettings.query.filter_by(class_id=class_id).first()
    show_fn = bool(class_settings and class_settings.show_first_name_only)
    fn_labels = _first_name_only_labels_for_class(class_id) if show_fn else {}
    active_poll = Poll.query.filter_by(class_id=class_id, is_active=True).first()
    payload = {
        'success': True,
        'class_id': class_id,
        'show_first_name_only': show_fn,
    }
    if show_fn:
        payload['first_name_only_display'] = fn_labels.get(student_id, '')
    if active_poll:
        payload['active_poll'] = {
            'poll_id': active_poll.id,
            'question': active_poll.question,
            'options': json.loads(active_poll.options or '[]'),
            'is_anonymous': bool(active_poll.is_anonymous),
            'is_graded': bool(active_poll.is_graded),
        }
    active_quiz_run = (
        QuizRun.query.options(joinedload(QuizRun.quiz).joinedload(Quiz.questions))
        .filter_by(class_id=class_id, is_active=True)
        .first()
    )
    if active_quiz_run:
        payload['active_quiz'] = quiz_run_public_payload(active_quiz_run)
    return jsonify(payload)


@app.route('/api/student/leave_class', methods=['POST'])
def student_leave_class():
    """Record leave time for today without ending student session (return to class list)."""
    student_id = _authenticated_student_id()
    if not student_id:
        return jsonify({'success': False, 'error': 'Not logged in'})
    data = request.get_json() or {}
    class_id = data.get('class_id')
    if not class_id:
        return jsonify({'success': False, 'error': 'Class ID required'})
    leave_time = datetime.utcnow()
    active_session = get_active_class_session(class_id)
    if active_session:
        attendance = Attendance.query.filter_by(
            class_id=class_id,
            student_id=student_id,
            class_session_id=active_session.id,
        ).first()
        if attendance:
            attendance.leave_time = leave_time
            db.session.commit()
    return jsonify({'success': True})


@app.route('/api/student/logout', methods=['POST'])
def student_logout():
    student_id = _authenticated_student_id()
    if not student_id:
        return jsonify({'success': False, 'error': 'Not logged in'})
    
    data = request.get_json()
    class_id = data.get('class_id')
    
    if class_id:
        leave_time = datetime.utcnow()
        active_session = get_active_class_session(class_id)
        if active_session:
            attendance = Attendance.query.filter_by(
                class_id=class_id,
                student_id=student_id,
                class_session_id=active_session.id,
            ).first()
            if attendance:
                attendance.leave_time = leave_time
                db.session.commit()
    
    session.pop('student_id', None)
    session.pop('needs_password', None)
    
    return jsonify({'success': True})

@app.route('/api/student/interaction', methods=['POST'])
def student_interaction():
    try:
        student_id = _authenticated_student_id()
        if not student_id:
            return jsonify({'success': False, 'error': 'Not logged in'})
        
        data = request.get_json()
        if not data:
            return jsonify({'success': False, 'error': 'No data provided'})
        
        class_id = data.get('class_id')
        interaction_type = data.get('type')  # 'hand_raise', 'thumbs_up', 'thumbs_down'
        
        if not class_id:
            return jsonify({'success': False, 'error': 'Class ID required'})
        try:
            class_id = int(class_id)
        except (TypeError, ValueError):
            return jsonify({'success': False, 'error': 'Invalid class id'}), 400

        if not interaction_type:
            return jsonify({'success': False, 'error': 'Interaction type required'})

        if Enrollment.query.filter_by(
            class_id=class_id,
            student_id=student_id,
            is_active=True,
        ).first() is None:
            return jsonify({'success': False, 'error': 'Not enrolled in this class'}), 403

        # Check if quiet mode is enabled
        settings = ClassSettings.query.filter_by(class_id=class_id).first()
        allow_quiet = (
            data.get('action') == 'auto_off'
            and interaction_type in ('thumbs_up', 'thumbs_down')
        )
        if settings and settings.quiet_mode and interaction_type in ['hand_raise', 'thumbs_up', 'thumbs_down'] and not allow_quiet:
            return jsonify({'success': False, 'error': 'Quiet mode is enabled. Participation is disabled.'})
        
        today = datetime.utcnow().date()
        participation = Participation.query.filter_by(
            class_id=class_id,
            student_id=student_id,
            date=today
        ).first()
        
        if not participation:
            participation = Participation(
                class_id=class_id,
                student_id=student_id,
                date=today
            )
            db.session.add(participation)
        
        # Ensure fields are initialized to 0 if None
        if participation.hand_raises is None:
            participation.hand_raises = 0
        if participation.thumbs_up is None:
            participation.thumbs_up = 0
        if participation.thumbs_down is None:
            participation.thumbs_down = 0
        
        # Check current state for toggle behavior
        action = 'toggle'  # Default to toggle
        if 'action' in data:
            action = data.get('action')  # 'raise' or 'lower'
        
        if interaction_type == 'hand_raise':
            # Check if student already has an active (not cleared) hand raise
            active_hand_raise = HandRaise.query.filter_by(
                class_id=class_id,
                student_id=student_id,
                cleared=False
            ).first()
            
            if active_hand_raise:
                # Lower hand (toggle off)
                active_hand_raise.cleared = True
                # Don't increment participation count when lowering
            else:
                # Raise hand (toggle on)
                participation.hand_raises += 1
                hand_raise = HandRaise(
                    class_id=class_id,
                    student_id=student_id,
                    timestamp=datetime.utcnow()
                )
                db.session.add(hand_raise)
                
        elif interaction_type == 'thumbs_up':
            cu = participation.thumbs_up or 0
            cd = participation.thumbs_down or 0
            if data.get('action') == 'auto_off':
                # Client 10s timer: turn off only if still active (no toggle — avoids turning on by mistake)
                if cu % 2 == 1:
                    participation.thumbs_up = max(0, cu - 1)
            else:
                # Mutually exclusive with thumbs_down: only one can be "on" per student
                if cu % 2 == 1 and cd % 2 == 1:
                    participation.thumbs_down = max(0, cd - 1)
                    cd = participation.thumbs_down
                if cu % 2 == 1:
                    participation.thumbs_up = max(0, cu - 1)
                else:
                    if cd % 2 == 1:
                        participation.thumbs_down = max(0, cd - 1)
                    participation.thumbs_up = cu + 1
        elif interaction_type == 'thumbs_down':
            cu = participation.thumbs_up or 0
            cd = participation.thumbs_down or 0
            if data.get('action') == 'auto_off':
                if cd % 2 == 1:
                    participation.thumbs_down = max(0, cd - 1)
            else:
                if cu % 2 == 1 and cd % 2 == 1:
                    participation.thumbs_up = max(0, cu - 1)
                    cu = participation.thumbs_up
                if cd % 2 == 1:
                    participation.thumbs_down = max(0, cd - 1)
                else:
                    if cu % 2 == 1:
                        participation.thumbs_up = max(0, cu - 1)
                    participation.thumbs_down = cd + 1
        else:
            return jsonify({'success': False, 'error': 'Invalid interaction type'})
        
        db.session.commit()
        
        # Determine if interaction is now active
        is_active = False
        if interaction_type == 'hand_raise':
            is_active = HandRaise.query.filter_by(
                class_id=class_id,
                student_id=student_id,
                cleared=False
            ).first() is not None
        elif interaction_type == 'thumbs_up':
            is_active = (participation.thumbs_up or 0) % 2 == 1
        elif interaction_type == 'thumbs_down':
            is_active = (participation.thumbs_down or 0) % 2 == 1
        
        # Refresh participation to get updated values
        db.session.refresh(participation)
        
        socketio.emit('student_interaction', {
            'student_id': student_id,
            'class_id': class_id,
            'type': interaction_type,
            'is_active': is_active
        }, room=f'class_{class_id}')
        
        out = {'success': True, 'is_active': is_active}
        if interaction_type in ('thumbs_up', 'thumbs_down'):
            out['thumbs_up_active'] = (participation.thumbs_up or 0) % 2 == 1
            out['thumbs_down_active'] = (participation.thumbs_down or 0) % 2 == 1
        return jsonify(out)
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/student/poll_response', methods=['POST'])
def student_poll_response():
    student_id = _authenticated_student_id()
    if not student_id:
        return jsonify({'success': False, 'error': 'Not logged in'})
    
    data = request.get_json()
    poll_id = data.get('poll_id')
    answer = data.get('answer')
    
    poll = Poll.query.get_or_404(poll_id)
    if not poll.is_active:
        return jsonify({'success': False, 'error': 'Poll is not active'})

    if Enrollment.query.filter_by(
        class_id=poll.class_id,
        student_id=student_id,
        is_active=True,
    ).first() is None:
        return jsonify({'success': False, 'error': 'Not enrolled in this class'}), 403

    # Check if already responded
    existing = PollResponse.query.filter_by(
        poll_id=poll_id,
        student_id=student_id
    ).first()
    
    if existing:
        return jsonify({'success': False, 'error': 'Already responded'})
    
    is_correct = (poll.correct_answer is not None and answer == poll.correct_answer)
    
    response = PollResponse(
        poll_id=poll_id,
        student_id=student_id,
        answer=answer,
        is_correct=is_correct
    )
    db.session.add(response)
    db.session.commit()
    
    has_key = poll.correct_answer is not None
    socketio.emit('poll_response', {
        'poll_id': poll_id,
        'student_id': student_id,
        'answer': answer,
        'is_correct': is_correct,
        'is_anonymous': poll.is_anonymous
    }, room=f'class_{poll.class_id}')
    
    return jsonify({
        'success': True,
        'is_correct': is_correct,
        'has_correct_answer': has_key,
        'correct_answer_index': poll.correct_answer,
        'selected_index': answer,
    })

# Live Dashboard APIs
@app.route('/api/live_dashboard/<int:class_id>')
@login_required
def live_dashboard(class_id):
    class_obj = Class.query.get_or_404(class_id)
    if class_obj.professor_id != current_user.id:
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403
    
    today = datetime.utcnow().date()
    
    # Get settings
    settings = ClassSettings.query.filter_by(class_id=class_id).first()
    if not settings:
        settings = ClassSettings(class_id=class_id)
        db.session.add(settings)
        db.session.commit()
    
    # Get total enrolled students
    total_students = db.session.query(Student).join(Enrollment).filter(
        Enrollment.class_id == class_id,
        Enrollment.is_active == True
    ).count()
    
    active_session = get_active_class_session(class_id)
    if active_session:
        present_students = db.session.query(Student).join(Attendance).filter(
            Attendance.class_id == class_id,
            Attendance.class_session_id == active_session.id,
            Attendance.leave_time == None,
        ).all()
    else:
        present_students = []

    # Get hands raised (not cleared, ordered by timestamp)
    hands_raised = db.session.query(HandRaise, Student).join(Student).filter(
        HandRaise.class_id == class_id,
        HandRaise.cleared == False
    ).order_by(HandRaise.timestamp.asc()).all()

    show_first_only = bool(settings and settings.show_first_name_only)
    fn_labels = _first_name_only_labels_for_class(class_id) if show_first_only else {}
    
    hands_raised_list = []
    for hand_raise, student in hands_raised:
        if show_first_only:
            display_name = fn_labels.get(
                student.id,
                (student.preferred_name or student.first_name or '').strip(),
            )
        else:
            display_name = f'{student.preferred_name or student.first_name} {student.last_name}'.strip()
        hands_raised_list.append({
            'student_id': student.id,
            'student_number': student.student_number,
            'first_name': student.first_name,
            'last_name': student.last_name,
            'preferred_name': student.preferred_name,
            'display_name': display_name,
            'timestamp': hand_raise.timestamp.isoformat()
        })
    
    # Count of currently active (not cleared) hand raises
    active_hand_raises_count = len(hands_raised_list)
    
    # Get participation counts
    participations = Participation.query.filter_by(
        class_id=class_id,
        date=today
    ).all()
    
    # Hand raises only — thumbs up/down must not affect this metric (faculty live "Unique Participation")
    unique_participants = len({p.student_id for p in participations if (p.hand_raises or 0) > 0})
    
    # Get thumbs up/down counts (only from current session)
    thumbs_up_count = 0
    thumbs_down_count = 0
    if active_session:
        # Get thumbs from current session only - count only active (odd numbers)
        session_participations = Participation.query.filter_by(
            class_id=class_id,
            date=today
        ).all()
        # Count only students who have thumbs up/down active (odd count)
        thumbs_up_count = sum(1 for p in session_participations if (p.thumbs_up or 0) % 2 == 1)
        thumbs_down_count = sum(1 for p in session_participations if (p.thumbs_down or 0) % 2 == 1)
    
    # Get active poll
    active_poll = Poll.query.filter_by(class_id=class_id, is_active=True).first()
    poll_data = None
    poll_results_data = None
    if active_poll:
        options, option_counts, total = poll_option_counts(active_poll)
        poll_data = {
            'poll_id': active_poll.id,
            'question': active_poll.question,
            'options': options,
            'is_anonymous': active_poll.is_anonymous,
            'is_graded': active_poll.is_graded
        }
        poll_results_data = {
            'success': True,
            'question': active_poll.question,
            'options': options,
            'option_counts': option_counts,
            'total_responses': total
        }

    active_quiz_run = (
        QuizRun.query.options(joinedload(QuizRun.quiz).joinedload(Quiz.questions))
        .filter_by(class_id=class_id, is_active=True)
        .first()
    )
    quiz_data = None
    if active_quiz_run:
        quiz_data = quiz_run_public_payload(active_quiz_run)
    
    return jsonify({
        'success': True,
        'active_hand_raises_count': active_hand_raises_count,
        'hands_raised': hands_raised_list,
        'unique_participants': unique_participants,
        'thumbs_up_count': thumbs_up_count,
        'thumbs_down_count': thumbs_down_count,
        'present_students': len(present_students),
        'total_students': total_students,
        'active_poll': poll_data,
        'poll_results': poll_results_data,
        'active_quiz': quiz_data,
        'show_first_name_only': settings.show_first_name_only
    })

@app.route('/api/live_attendance/<int:class_id>')
@login_required
def live_attendance(class_id):
    class_obj = Class.query.get_or_404(class_id)
    if class_obj.professor_id != current_user.id:
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403
    
    today = datetime.utcnow().date()

    # Get all enrolled students
    enrolled_students = db.session.query(Student).join(Enrollment).filter(
        Enrollment.class_id == class_id,
        Enrollment.is_active == True
    ).all()

    active_session = get_active_class_session(class_id)
    if active_session:
        present_student_ids = set(
            id[0] for id in db.session.query(Attendance.student_id).filter(
                Attendance.class_id == class_id,
                Attendance.class_session_id == active_session.id,
                Attendance.leave_time == None,
            ).all()
        )
    else:
        present_student_ids = set()

    # Get participation data for all students
    participations = Participation.query.filter_by(
        class_id=class_id,
        date=today
    ).all()
    participation_map = {p.student_id: p for p in participations}

    poll_map = gradebook_poll_responses_by_student(class_id)
    poll_grades = {}
    for student_id, responses in poll_map.items():
        poll_grades[student_id] = {
            'correct': sum(1 for pr in responses if pr.is_correct),
            'total': len(responses),
        }
    
    present_students = []
    absent_students = []
    
    for student in enrolled_students:
        participation = participation_map.get(student.id)
        poll_grade_data = poll_grades.get(student.id, {'correct': 0, 'total': 0})
        poll_grade = (poll_grade_data['correct'] / poll_grade_data['total'] * 100) if poll_grade_data['total'] > 0 else 0
        
        student_data = {
            'id': student.id,
            'student_number': student.student_number,
            'first_name': student.first_name,
            'last_name': student.last_name,
            'preferred_name': student.preferred_name,
            'hand_raises': participation.hand_raises if participation else 0,
            'thumbs_up': participation.thumbs_up if participation else 0,
            'thumbs_down': participation.thumbs_down if participation else 0,
            'poll_grade': round(poll_grade, 1),
            'participation_freq': (participation.hand_raises or 0) + (participation.thumbs_up or 0) + (participation.thumbs_down or 0) if participation else 0
        }
        
        if student.id in present_student_ids:
            present_students.append(student_data)
        else:
            absent_students.append(student_data)
    
    return jsonify({
        'success': True,
        'present_students': present_students,
        'absent_students': absent_students
    })

@app.route('/api/live_preferences/<int:class_id>')
@login_required
def live_preferences(class_id):
    class_obj = Class.query.get_or_404(class_id)
    if class_obj.professor_id != current_user.id:
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403
    
    settings = ClassSettings.query.filter_by(class_id=class_id).first()
    if not settings:
        settings = ClassSettings(class_id=class_id)
        db.session.add(settings)
        db.session.commit()
    
    # Get current session's exclude_from_grading status
    active_session = ClassSession.query.filter_by(
        class_id=class_id,
        end_time=None
    ).order_by(ClassSession.start_time.desc()).first()
    
    exclude_from_grading = active_session.exclude_from_grading if active_session else False
    
    return jsonify({
        'success': True,
        'quiet_mode': settings.quiet_mode,
        'show_first_name_only': settings.show_first_name_only,
        'exclude_from_grading': exclude_from_grading
    })

@app.route('/api/clear_hands_raised/<int:class_id>', methods=['POST'])
@login_required
def clear_hands_raised(class_id):
    class_obj = Class.query.get_or_404(class_id)
    if class_obj.professor_id != current_user.id:
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403
    
    HandRaise.query.filter_by(
        class_id=class_id,
        cleared=False
    ).update({'cleared': True})
    
    db.session.commit()
    
    socketio.emit('all_hands_cleared', {'class_id': class_id}, room=f'class_{class_id}')
    
    return jsonify({'success': True})


@app.route('/api/dismiss_hand_raise/<int:class_id>', methods=['POST'])
@login_required
def dismiss_hand_raise(class_id):
    """Mark one student's active hand raise as cleared (e.g. after they participated)."""
    class_obj = Class.query.get_or_404(class_id)
    if class_obj.professor_id != current_user.id:
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403

    data = request.get_json() or {}
    student_id = data.get('student_id')
    if student_id is None:
        return jsonify({'success': False, 'error': 'student_id required'}), 400
    try:
        student_id = int(student_id)
    except (TypeError, ValueError):
        return jsonify({'success': False, 'error': 'Invalid student_id'}), 400

    hr = HandRaise.query.filter_by(
        class_id=class_id,
        student_id=student_id,
        cleared=False,
    ).first()
    if not hr:
        return jsonify({'success': False, 'error': 'No active hand raise for this student'}), 404

    hr.cleared = True
    db.session.commit()

    socketio.emit(
        'hand_raise_dismissed',
        {'class_id': class_id, 'student_id': student_id},
        room=f'class_{class_id}',
    )

    return jsonify({'success': True})


@app.route('/api/participation_grade/start', methods=['POST'])
@login_required
def participation_grade_start():
    """Begin a participation grading round for a subject student (raised hand)."""
    data = request.get_json() or {}
    class_id = data.get('class_id')
    subject_student_id = data.get('subject_student_id')
    if class_id is None or subject_student_id is None:
        return jsonify({'success': False, 'error': 'class_id and subject_student_id required'}), 400
    try:
        class_id = int(class_id)
        subject_student_id = int(subject_student_id)
    except (TypeError, ValueError):
        return jsonify({'success': False, 'error': 'Invalid ids'}), 400

    class_obj = Class.query.get_or_404(class_id)
    if class_obj.professor_id != current_user.id:
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403

    settings = ClassSettings.query.filter_by(class_id=class_id).first()
    if settings and settings.quiet_mode:
        return jsonify({'success': False, 'error': 'Participation grading is unavailable while quiet mode is on.'}), 400

    active_session = get_active_class_session(class_id)
    if not class_obj.is_active or not active_session:
        return jsonify({'success': False, 'error': 'Class session is not active.'}), 400

    enroll = Enrollment.query.filter_by(
        class_id=class_id,
        student_id=subject_student_id,
        is_active=True,
    ).first()
    if not enroll:
        return jsonify({'success': False, 'error': 'Student is not enrolled in this class.'}), 400

    hr = HandRaise.query.filter_by(
        class_id=class_id,
        student_id=subject_student_id,
        cleared=False,
    ).first()
    hand_raise_id = hr.id if hr else None

    today = datetime.utcnow().date()
    rnd = ParticipationGradeRound(
        class_id=class_id,
        subject_student_id=subject_student_id,
        date=today,
        hand_raise_id=hand_raise_id,
        class_session_id=active_session.id,
    )
    db.session.add(rnd)
    db.session.commit()

    return jsonify({'success': True, 'round_id': rnd.id})


@app.route('/api/participation_grade/instructor', methods=['POST'])
@login_required
def participation_grade_instructor():
    """Professor submits 1–100; notifies class room to collect peer ratings."""
    data = request.get_json() or {}
    class_id = data.get('class_id')
    round_id = data.get('round_id')
    score = data.get('score')
    if class_id is None or round_id is None or score is None:
        return jsonify({'success': False, 'error': 'class_id, round_id, and score required'}), 400
    try:
        class_id = int(class_id)
        round_id = int(round_id)
        score = int(score)
    except (TypeError, ValueError):
        return jsonify({'success': False, 'error': 'Invalid values'}), 400

    if score < 1 or score > 100:
        return jsonify({'success': False, 'error': 'Score must be between 1 and 100.'}), 400

    class_obj = Class.query.get_or_404(class_id)
    if class_obj.professor_id != current_user.id:
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403

    settings = ClassSettings.query.filter_by(class_id=class_id).first()
    if settings and settings.quiet_mode:
        return jsonify({'success': False, 'error': 'Participation grading is unavailable while quiet mode is on.'}), 400

    rnd = ParticipationGradeRound.query.filter_by(id=round_id, class_id=class_id).first()
    if not rnd:
        return jsonify({'success': False, 'error': 'Round not found.'}), 404

    if InstructorParticipationGrade.query.filter_by(round_id=round_id).first():
        return jsonify({'success': False, 'error': 'Instructor grade already submitted for this round.'}), 409

    db.session.add(InstructorParticipationGrade(round_id=round_id, score=score))
    _recompute_subject_participation_grades(class_id, rnd.subject_student_id, rnd.date)
    cleared_hand = _mark_active_hand_raise_cleared(class_id, rnd.subject_student_id)
    db.session.commit()

    subject = Student.query.get(rnd.subject_student_id)
    display_name = _subject_display_name_for_participation_grade(class_id, subject) if subject else ''

    socketio.emit(
        'peer_participation_grade_request',
        {
            'class_id': class_id,
            'round_id': round_id,
            'subject_student_id': rnd.subject_student_id,
            'subject_display_name': display_name,
        },
        room=f'class_{class_id}',
    )
    if cleared_hand:
        socketio.emit(
            'hand_raise_dismissed',
            {'class_id': class_id, 'student_id': rnd.subject_student_id},
            room=f'class_{class_id}',
        )

    return jsonify({'success': True})


@app.route('/api/student/peer_participation_grade', methods=['POST'])
def student_peer_participation_grade():
    """Peer submits rating 0–4 (mapped to %); updates aggregated peer_grade."""
    grader_id = _authenticated_student_id()
    if not grader_id:
        return jsonify({'success': False, 'error': 'Not logged in'}), 401

    data = request.get_json() or {}
    class_id = data.get('class_id')
    round_id = data.get('round_id')
    rating = data.get('rating')
    if class_id is None or round_id is None or rating is None:
        return jsonify({'success': False, 'error': 'class_id, round_id, and rating required'}), 400
    try:
        class_id = int(class_id)
        round_id = int(round_id)
        rating = int(rating)
    except (TypeError, ValueError):
        return jsonify({'success': False, 'error': 'Invalid values'}), 400

    try:
        pct = _peer_rating_to_percent(rating)
    except ValueError:
        return jsonify({'success': False, 'error': 'rating must be 0, 1, 2, 3, or 4'}), 400

    rnd = ParticipationGradeRound.query.filter_by(id=round_id, class_id=class_id).first()
    if not rnd:
        return jsonify({'success': False, 'error': 'Round not found.'}), 404

    if grader_id == rnd.subject_student_id:
        return jsonify({'success': False, 'error': 'You cannot grade yourself.'}), 400

    if not InstructorParticipationGrade.query.filter_by(round_id=round_id).first():
        return jsonify({'success': False, 'error': 'This grading round is not open for peer scores yet.'}), 400

    enroll = Enrollment.query.filter_by(
        class_id=class_id,
        student_id=grader_id,
        is_active=True,
    ).first()
    if not enroll:
        return jsonify({'success': False, 'error': 'Not enrolled in this class.'}), 400

    existing = PeerParticipationGrade.query.filter_by(
        round_id=round_id,
        grader_student_id=grader_id,
    ).first()
    if existing:
        return jsonify({'success': False, 'error': 'You already submitted a rating for this round.'}), 409

    db.session.add(
        PeerParticipationGrade(
            round_id=round_id,
            grader_student_id=grader_id,
            rating=rating,
            score_percent=pct,
        )
    )
    _recompute_subject_participation_grades(class_id, rnd.subject_student_id, rnd.date)
    db.session.commit()

    return jsonify({'success': True})


def _mark_active_hand_raise_cleared(class_id, student_id):
    """Set cleared=True on active HandRaise if present. Returns True if a row was updated."""
    hr = HandRaise.query.filter_by(
        class_id=class_id,
        student_id=student_id,
        cleared=False,
    ).first()
    if hr:
        hr.cleared = True
        return True
    return False


@app.route('/api/participation_grade/cancel', methods=['POST'])
@login_required
def participation_grade_cancel():
    """Abandon a grading round before instructor submits; removes student from queue and lowers hand."""
    data = request.get_json() or {}
    class_id = data.get('class_id')
    round_id = data.get('round_id')
    if class_id is None or round_id is None:
        return jsonify({'success': False, 'error': 'class_id and round_id required'}), 400
    try:
        class_id = int(class_id)
        round_id = int(round_id)
    except (TypeError, ValueError):
        return jsonify({'success': False, 'error': 'Invalid values'}), 400

    class_obj = Class.query.get_or_404(class_id)
    if class_obj.professor_id != current_user.id:
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403

    rnd = ParticipationGradeRound.query.filter_by(id=round_id, class_id=class_id).first()
    if not rnd:
        return jsonify({'success': False, 'error': 'Round not found.'}), 404

    if InstructorParticipationGrade.query.filter_by(round_id=round_id).first():
        return jsonify({'success': False, 'error': 'Cannot cancel after instructor grade is submitted.'}), 400

    subject_id = rnd.subject_student_id
    db.session.delete(rnd)
    cleared_hand = _mark_active_hand_raise_cleared(class_id, subject_id)
    db.session.commit()

    if cleared_hand:
        socketio.emit(
            'hand_raise_dismissed',
            {'class_id': class_id, 'student_id': subject_id},
            room=f'class_{class_id}',
        )

    return jsonify({'success': True})


@app.route('/api/clear_participation_count/<int:class_id>', methods=['POST'])
@login_required
def clear_participation_count(class_id):
    class_obj = Class.query.get_or_404(class_id)
    if class_obj.professor_id != current_user.id:
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403
    
    # This is a conceptual reset - in practice, we might want to track this differently
    # For now, we'll just return success as the count is calculated from participations
    return jsonify({'success': True, 'message': 'Participation count is calculated from current session data'})

@app.route('/api/reset_thumbs_up/<int:class_id>', methods=['POST'])
@login_required
def reset_thumbs_up(class_id):
    class_obj = Class.query.get_or_404(class_id)
    if class_obj.professor_id != current_user.id:
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403
    
    # Reset thumbs up for current session
    today = datetime.utcnow().date()
    participations = Participation.query.filter_by(
        class_id=class_id,
        date=today
    ).all()
    
    for p in participations:
        if p.thumbs_up and p.thumbs_up > 0:
            p.thumbs_up = 0
    
    db.session.commit()
    socketio.emit('thumbs_reactions_cleared', {'class_id': class_id}, room=f'class_{class_id}')
    
    return jsonify({'success': True})

@app.route('/api/reset_thumbs_down/<int:class_id>', methods=['POST'])
@login_required
def reset_thumbs_down(class_id):
    class_obj = Class.query.get_or_404(class_id)
    if class_obj.professor_id != current_user.id:
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403
    
    # Reset thumbs down for current session
    today = datetime.utcnow().date()
    participations = Participation.query.filter_by(
        class_id=class_id,
        date=today
    ).all()
    
    for p in participations:
        if p.thumbs_down and p.thumbs_down > 0:
            p.thumbs_down = 0
    
    db.session.commit()
    socketio.emit('thumbs_reactions_cleared', {'class_id': class_id}, room=f'class_{class_id}')
    
    return jsonify({'success': True})

@app.route('/api/student/interaction_state/<int:class_id>')
def get_student_interaction_state(class_id):
    """Get current interaction states for logged-in student"""
    student_id = _authenticated_student_id()
    if not student_id:
        return jsonify({'success': False, 'error': 'Not logged in'}), 401
    
    # Check hand raise
    hand_raise_active = HandRaise.query.filter_by(
        class_id=class_id,
        student_id=student_id,
        cleared=False
    ).first() is not None
    
    # Check thumbs up/down
    today = datetime.utcnow().date()
    participation = Participation.query.filter_by(
        class_id=class_id,
        student_id=student_id,
        date=today
    ).first()
    
    thumbs_up_active = False
    thumbs_down_active = False
    if participation:
        thumbs_up_active = (participation.thumbs_up or 0) % 2 == 1
        thumbs_down_active = (participation.thumbs_down or 0) % 2 == 1
    
    return jsonify({
        'success': True,
        'hand_raise_active': hand_raise_active,
        'thumbs_up_active': thumbs_up_active,
        'thumbs_down_active': thumbs_down_active
    })

@app.route('/api/poll_results/<int:poll_id>')
@login_required
def poll_results(poll_id):
    poll = Poll.query.get_or_404(poll_id)
    class_obj = Class.query.get_or_404(poll.class_id)
    if class_obj.professor_id != current_user.id:
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403

    options, option_counts, total = poll_option_counts(poll)
    return jsonify({
        'success': True,
        'question': poll.question,
        'options': options,
        'option_counts': option_counts,
        'total_responses': total
    })

# SocketIO Events
@socketio.on('connect')
def on_connect():
    emit('connected', {'data': 'Connected'})

@socketio.on('join_student_enrollments')
def on_join_student_enrollments(data=None):
    """Join Socket.IO rooms so this client receives class_started / class_stopped while on My Classes."""
    student_id = _student_id_from_socket_token(data)
    if not student_id:
        return
    enrollments = Enrollment.query.filter_by(
        student_id=student_id,
        is_active=True,
    ).all()
    for e in enrollments:
        join_room(f'enrolled_{e.class_id}')
    emit('enrolled_feed_ready', {'class_ids': [e.class_id for e in enrollments]})


def _socket_authorize_class_access(data):
    """Return (class_id, kind) if caller may receive events for this class, else (None, None).

    kind is 'prof' when the caller owns the class (Flask-Login session), 'student' when the
    caller is actively enrolled via a valid bearer token in the socket payload.
    """
    if not isinstance(data, dict):
        return None, None
    raw_cid = data.get('class_id')
    try:
        class_id = int(raw_cid)
    except (TypeError, ValueError):
        return None, None

    class_obj = Class.query.get(class_id)
    if class_obj is None:
        return None, None

    try:
        if current_user.is_authenticated and getattr(current_user, 'id', None) == class_obj.professor_id:
            return class_id, 'prof'
    except Exception:
        pass

    sid = _student_id_from_socket_token(data)
    if sid is not None:
        enrolled = Enrollment.query.filter_by(
            class_id=class_id,
            student_id=sid,
            is_active=True,
        ).first()
        if enrolled is not None:
            return class_id, 'student'

    return None, None


@socketio.on('join_class')
def on_join_class(data):
    class_id, _kind = _socket_authorize_class_access(data)
    if class_id is None:
        emit('join_class_denied', {'class_id': (data or {}).get('class_id')})
        return
    join_room(f'class_{class_id}')
    emit('joined_class', {'class_id': class_id})


@socketio.on('leave_class')
def on_leave_class(data):
    if data and data.get('class_id') is not None:
        leave_room(f'class_{data["class_id"]}')

@socketio.on('get_live_stats')
def on_get_live_stats(data):
    class_id, kind = _socket_authorize_class_access(data)
    if class_id is None or kind != 'prof':
        return

    total_students = db.session.query(Enrollment).filter(
        Enrollment.class_id == class_id,
        Enrollment.is_active == True
    ).count()

    active_session = get_active_class_session(class_id)
    if active_session:
        present_students = db.session.query(Attendance).filter(
            Attendance.class_id == class_id,
            Attendance.class_session_id == active_session.id,
            Attendance.leave_time == None,
        ).count()
    else:
        present_students = 0

    today = datetime.utcnow().date()
    participations = Participation.query.filter_by(
        class_id=class_id,
        date=today
    ).all()
    total_hand_raises = sum(p.hand_raises for p in participations)
    total_thumbs_up = sum(p.thumbs_up for p in participations)
    total_thumbs_down = sum(p.thumbs_down for p in participations)

    active_poll = Poll.query.filter_by(class_id=class_id, is_active=True).first()
    poll_stats = None
    if active_poll:
        options, option_counts, total = poll_option_counts(active_poll)
        poll_stats = {
            'poll_id': active_poll.id,
            'question': active_poll.question,
            'options': options,
            'option_counts': option_counts,
            'total_responses': total,
            'is_anonymous': active_poll.is_anonymous
        }

    emit('live_stats', {
        'total_students': total_students,
        'present_students': present_students,
        'total_hand_raises': total_hand_raises,
        'total_thumbs_up': total_thumbs_up,
        'total_thumbs_down': total_thumbs_down,
        'poll_stats': poll_stats
    })

def migrate_database():
    """Add missing columns and tables to existing database."""
    from sqlalchemy import inspect, text
    
    try:
        inspector = inspect(db.engine)
        table_names = inspector.get_table_names()
        
        # Check if student table exists and add missing columns
        if 'student' in table_names:
            student_columns = [col['name'] for col in inspector.get_columns('student')]
            
            if 'preferred_name' not in student_columns:
                try:
                    db.session.execute(text('ALTER TABLE student ADD COLUMN preferred_name VARCHAR(100)'))
                    db.session.commit()
                    print("[OK] Added preferred_name column to student table")
                except Exception as e:
                    db.session.rollback()
                    print(f"[ERROR] Error adding preferred_name column: {e}")
            
            if 'email' not in student_columns:
                try:
                    # SQLite doesn't support DEFAULT in ALTER TABLE, so add as nullable first
                    # We'll update existing records with a placeholder email
                    db.session.execute(text('ALTER TABLE student ADD COLUMN email VARCHAR(120)'))
                    db.session.commit()
                    
                    # Update existing records with a placeholder email if they don't have one
                    # Format: student_number@placeholder.local
                    students_without_email = db.session.execute(
                        text('SELECT id, student_number FROM student WHERE email IS NULL OR email = ""')
                    ).fetchall()
                    
                    for student_id, student_number in students_without_email:
                        placeholder_email = f"{student_number}@placeholder.local"
                        db.session.execute(
                            text('UPDATE student SET email = :email WHERE id = :id'),
                            {'email': placeholder_email, 'id': student_id}
                        )
                    
                    db.session.commit()
                    print("[OK] Added email column to student table")
                    if students_without_email:
                        print(f"  → Updated {len(students_without_email)} existing students with placeholder emails")
                        print("  → Please update student emails via Excel import or manual edit")
                except Exception as e:
                    db.session.rollback()
                    print(f"[ERROR] Error adding email column: {e}")
            
            if 'password_hash' not in student_columns:
                try:
                    db.session.execute(text('ALTER TABLE student ADD COLUMN password_hash VARCHAR(255)'))
                    db.session.commit()
                    print("[OK] Added password_hash column to student table")
                except Exception as e:
                    db.session.rollback()
                    print(f"[ERROR] Error adding password_hash column: {e}")
        
        # Check if attendance table exists and add missing columns
        if 'attendance' in table_names:
            attendance_columns = [col['name'] for col in inspector.get_columns('attendance')]
            
            if 'join_time' not in attendance_columns:
                try:
                    db.session.execute(text('ALTER TABLE attendance ADD COLUMN join_time DATETIME'))
                    db.session.commit()
                    print("[OK] Added join_time column to attendance table")
                except Exception as e:
                    db.session.rollback()
                    print(f"[ERROR] Error adding join_time column: {e}")
            
            if 'leave_time' not in attendance_columns:
                try:
                    db.session.execute(text('ALTER TABLE attendance ADD COLUMN leave_time DATETIME'))
                    db.session.commit()
                    print("[OK] Added leave_time column to attendance table")
                except Exception as e:
                    db.session.rollback()
                    print(f"[ERROR] Error adding leave_time column: {e}")

            attendance_columns = [col['name'] for col in inspector.get_columns('attendance')]
            if 'class_session_id' not in attendance_columns:
                try:
                    db.session.execute(text('ALTER TABLE attendance ADD COLUMN class_session_id INTEGER'))
                    db.session.commit()
                    print("[OK] Added class_session_id column to attendance table")
                except Exception as e:
                    db.session.rollback()
                    print(f"[ERROR] Error adding class_session_id column: {e}")
        
        # Check if poll table exists and add is_graded column
        if 'poll' in table_names:
            poll_columns = [col['name'] for col in inspector.get_columns('poll')]
            if 'is_graded' not in poll_columns:
                try:
                    db.session.execute(text('ALTER TABLE poll ADD COLUMN is_graded BOOLEAN DEFAULT 0'))
                    db.session.commit()
                    print("[OK] Added is_graded column to poll table")
                except Exception as e:
                    db.session.rollback()
                    print(f"[ERROR] Error adding is_graded column: {e}")
            inspector = inspect(db.engine)
            poll_columns = [col['name'] for col in inspector.get_columns('poll')]
            if 'show_results_when_stopped' not in poll_columns:
                try:
                    db.session.execute(
                        text('ALTER TABLE poll ADD COLUMN show_results_when_stopped BOOLEAN DEFAULT 1')
                    )
                    db.session.commit()
                    print("[OK] Added show_results_when_stopped column to poll table")
                except Exception as e:
                    db.session.rollback()
                    print(f"[ERROR] Error adding show_results_when_stopped to poll: {e}")

        # Poll bank: add named sets and link questions to sets
        inspector = inspect(db.engine)
        table_names = inspector.get_table_names()
        if 'poll_bank_set' in table_names and 'poll_bank_question' in table_names:
            pbq_columns = [col['name'] for col in inspector.get_columns('poll_bank_question')]
            if 'set_id' not in pbq_columns:
                try:
                    db.session.execute(text('ALTER TABLE poll_bank_question ADD COLUMN set_id INTEGER'))
                    db.session.commit()
                    print("[OK] Added set_id column to poll_bank_question table")
                except Exception as e:
                    db.session.rollback()
                    print(f"[ERROR] Error adding set_id to poll_bank_question: {e}")
            try:
                legacy_rows = db.session.execute(
                    text('SELECT id, class_id FROM poll_bank_question WHERE set_id IS NULL')
                ).fetchall()
                set_by_class = {}
                for _, class_id in legacy_rows:
                    if class_id in set_by_class:
                        continue
                    existing = PollBankSet.query.filter_by(class_id=class_id, name='Default Set').first()
                    if not existing:
                        existing = PollBankSet(class_id=class_id, name='Default Set')
                        db.session.add(existing)
                        db.session.flush()
                    set_by_class[class_id] = existing.id
                for row_id, class_id in legacy_rows:
                    sid = set_by_class.get(class_id)
                    if sid:
                        db.session.execute(
                            text('UPDATE poll_bank_question SET set_id = :sid WHERE id = :rid'),
                            {'sid': sid, 'rid': row_id}
                        )
                if legacy_rows:
                    db.session.commit()
                    print(f"[OK] Backfilled poll_bank_question.set_id for {len(legacy_rows)} row(s)")
            except Exception as e:
                db.session.rollback()
                print(f"[WARN] poll_bank_question set backfill skipped: {e}")
        
        # Check if class_session table exists and add exclude_from_grading column
        if 'class_session' in table_names:
            session_columns = [col['name'] for col in inspector.get_columns('class_session')]
            if 'exclude_from_grading' not in session_columns:
                try:
                    db.session.execute(text('ALTER TABLE class_session ADD COLUMN exclude_from_grading BOOLEAN DEFAULT 0'))
                    db.session.commit()
                    print("[OK] Added exclude_from_grading column to class_session table")
                except Exception as e:
                    db.session.rollback()
                    print(f"[ERROR] Error adding exclude_from_grading column: {e}")
        
        # Check if enrollment table exists and if is_active column exists
        if 'enrollment' in table_names:
            enrollment_columns = [col['name'] for col in inspector.get_columns('enrollment')]
            if 'is_active' not in enrollment_columns:
                try:
                    db.session.execute(text('ALTER TABLE enrollment ADD COLUMN is_active BOOLEAN DEFAULT 1'))
                    db.session.commit()
                    print("[OK] Added is_active column to enrollment table")
                except Exception as e:
                    db.session.rollback()
                    print(f"[ERROR] Error adding is_active column: {e}")
        
        # Grading weights: participation bucket + instructor/peer split (replaces two top-level participation columns)
        if 'grading_weights' in table_names:
            gw_columns = [col['name'] for col in inspector.get_columns('grading_weights')]
            has_legacy = 'instructor_participation_weight' in gw_columns
            if 'participation_weight' not in gw_columns:
                try:
                    db.session.execute(text('ALTER TABLE grading_weights ADD COLUMN participation_weight FLOAT'))
                    db.session.commit()
                    print("[OK] Added participation_weight column to grading_weights table")
                    gw_columns.append('participation_weight')
                except Exception as e:
                    db.session.rollback()
                    print(f"[ERROR] Error adding participation_weight to grading_weights: {e}")
            if 'participation_instructor_share' not in gw_columns:
                try:
                    db.session.execute(text('ALTER TABLE grading_weights ADD COLUMN participation_instructor_share FLOAT'))
                    db.session.commit()
                    print("[OK] Added participation_instructor_share column to grading_weights table")
                except Exception as e:
                    db.session.rollback()
                    print(f"[ERROR] Error adding participation_instructor_share to grading_weights: {e}")
            inspector = inspect(db.engine)
            gw_columns = [col['name'] for col in inspector.get_columns('grading_weights')]
            has_legacy = 'instructor_participation_weight' in gw_columns
            if has_legacy:
                try:
                    db.session.execute(text("""
                        UPDATE grading_weights SET
                          participation_weight = COALESCE(instructor_participation_weight, 0) + COALESCE(peer_participation_weight, 0),
                          participation_instructor_share = CASE
                            WHEN COALESCE(instructor_participation_weight, 0) + COALESCE(peer_participation_weight, 0) > 0
                            THEN 100.0 * instructor_participation_weight / (instructor_participation_weight + peer_participation_weight)
                            ELSE 50.0 END
                    """))
                    db.session.commit()
                    print("[OK] Backfilled grading_weights participation columns from legacy instructor/peer weights")
                except Exception as e:
                    db.session.rollback()
                    print(f"[ERROR] Error backfilling grading_weights: {e}")
                for legacy_col in ('instructor_participation_weight', 'peer_participation_weight'):
                    if legacy_col in gw_columns:
                        try:
                            db.session.execute(text(f'ALTER TABLE grading_weights DROP COLUMN {legacy_col}'))
                            db.session.commit()
                            print(f"[OK] Dropped legacy column grading_weights.{legacy_col}")
                        except Exception as e:
                            db.session.rollback()
                            print(f"[WARN] Could not drop {legacy_col} (ignored): {e}")
            else:
                try:
                    db.session.execute(text("""
                        UPDATE grading_weights SET
                          participation_weight = 50.0,
                          participation_instructor_share = 50.0
                        WHERE participation_weight IS NULL OR participation_instructor_share IS NULL
                    """))
                    db.session.commit()
                except Exception as e:
                    db.session.rollback()
                    print(f"[ERROR] Error defaulting grading_weights participation columns: {e}")
            inspector = inspect(db.engine)
            gw_columns2 = [col['name'] for col in inspector.get_columns('grading_weights')]
            if 'quiz_weight' not in gw_columns2:
                try:
                    db.session.execute(text('ALTER TABLE grading_weights ADD COLUMN quiz_weight FLOAT DEFAULT 0'))
                    db.session.commit()
                    print("[OK] Added quiz_weight column to grading_weights table")
                except Exception as e:
                    db.session.rollback()
                    print(f"[ERROR] Error adding quiz_weight to grading_weights: {e}")
            inspector = inspect(db.engine)
            gw_columns3 = [col['name'] for col in inspector.get_columns('grading_weights')]
            if 'quiz_count_target' not in gw_columns3:
                try:
                    db.session.execute(text('ALTER TABLE grading_weights ADD COLUMN quiz_count_target INTEGER DEFAULT 0'))
                    db.session.commit()
                    print("[OK] Added quiz_count_target column to grading_weights table")
                except Exception as e:
                    db.session.rollback()
                    print(f"[ERROR] Error adding quiz_count_target to grading_weights: {e}")
        
        # Participation grade rounds: exclude individual rounds from averages
        if 'participation_grade_round' in table_names:
            pgr_columns = [col['name'] for col in inspector.get_columns('participation_grade_round')]
            if 'exclude_from_grading' not in pgr_columns:
                try:
                    db.session.execute(
                        text('ALTER TABLE participation_grade_round ADD COLUMN exclude_from_grading BOOLEAN DEFAULT 0')
                    )
                    db.session.commit()
                    print("[OK] Added exclude_from_grading column to participation_grade_round table")
                except Exception as e:
                    db.session.rollback()
                    print(f"[ERROR] Error adding exclude_from_grading to participation_grade_round: {e}")
            inspector = inspect(db.engine)
            pgr_columns = [col['name'] for col in inspector.get_columns('participation_grade_round')]
            if 'class_session_id' not in pgr_columns:
                try:
                    db.session.execute(
                        text('ALTER TABLE participation_grade_round ADD COLUMN class_session_id INTEGER')
                    )
                    db.session.commit()
                    print("[OK] Added class_session_id column to participation_grade_round table")
                except Exception as e:
                    db.session.rollback()
                    print(f"[ERROR] Error adding class_session_id to participation_grade_round: {e}")
            try:
                rounds = ParticipationGradeRound.query.filter(
                    ParticipationGradeRound.class_session_id.is_(None)
                ).all()
                now = datetime.utcnow()
                updated = 0
                for rnd in rounds:
                    sessions = ClassSession.query.filter_by(class_id=rnd.class_id).all()
                    matches = []
                    for s in sessions:
                        se = s.end_time if s.end_time else now
                        if rnd.created_at and s.start_time <= rnd.created_at <= se:
                            matches.append(s)
                    chosen = None
                    if len(matches) == 1:
                        chosen = matches[0]
                    elif len(matches) > 1:
                        same_date = [s for s in matches if s.start_time.date() == rnd.date]
                        if len(same_date) == 1:
                            chosen = same_date[0]
                    if chosen is not None:
                        rnd.class_session_id = chosen.id
                        updated += 1
                if updated:
                    db.session.commit()
                    print(f"[OK] Backfilled class_session_id for {updated} participation_grade_round row(s)")
            except Exception as e:
                db.session.rollback()
                print(f"[WARN] participation_grade_round class_session_id backfill skipped: {e}")
        
        if 'professor_preferences' in table_names:
            pp_columns = [col['name'] for col in inspector.get_columns('professor_preferences')]
            if 'dark_mode' not in pp_columns:
                try:
                    db.session.execute(
                        text('ALTER TABLE professor_preferences ADD COLUMN dark_mode BOOLEAN DEFAULT 0')
                    )
                    db.session.commit()
                    print("[OK] Added dark_mode column to professor_preferences table")
                except Exception as e:
                    db.session.rollback()
                    print(f"[ERROR] Error adding dark_mode to professor_preferences: {e}")

        # Check if hand_raise table exists
        if 'hand_raise' not in table_names:
            try:
                # Table will be created by create_all
                print("[OK] hand_raise table will be created")
            except Exception as e:
                print(f"[ERROR] Error with hand_raise table: {e}")
        
        # Ensure all tables exist - create_all will handle new tables
        db.create_all()
        print("[OK] Database migration completed")
    except Exception as e:
        print(f"[ERROR] Error during database migration: {e}")

if __name__ == '__main__':
    with app.app_context():
        db.create_all()
        migrate_database()
        
        # Create a default professor for testing
        if not Professor.query.first():
            default_prof = Professor(
                username='professor',
                email='prof@example.com',
                password_hash=generate_password_hash('password')
            )
            db.session.add(default_prof)
            db.session.commit()
    
    # `python app.py` runs the Werkzeug dev server via Flask-SocketIO. That server
    # refuses to bind unless allow_unsafe_werkzeug=True is passed, so we always set it
    # here. Real deploys should set PRODUCTION=1 and run under gunicorn/eventlet
    # behind a reverse proxy instead of invoking this entrypoint.
    socketio.run(
        app,
        debug=_IS_DEBUG,
        host='0.0.0.0',
        port=int(os.environ.get('PORT', '5000')),
        allow_unsafe_werkzeug=True,
    )

