"""Tests for quiz-weighted gradebook aggregation (equal mean over N slots)."""
import json
import os
import sys
from io import BytesIO

from openpyxl import Workbook, load_workbook
import pytest
from datetime import datetime, timedelta
from werkzeug.security import generate_password_hash


@pytest.fixture(scope="module")
def app_module(tmp_path_factory):
    db_path = tmp_path_factory.mktemp("qz") / "quiz_grade.db"
    os.environ["SQLALCHEMY_DATABASE_URI"] = f"sqlite:///{db_path}"
    os.environ["SECRET_KEY"] = "test-secret-quiz-grading"
    os.environ["TESTING"] = "1"
    for name in list(sys.modules.keys()):
        if name == "app" or name.startswith("app."):
            del sys.modules[name]
    import app as app_mod

    with app_mod.app.app_context():
        app_mod.db.create_all()
        app_mod.migrate_database()
    return app_mod


def test_quiz_grade_is_mean_over_n_slots_with_zeros(app_module):
    """N=2, one quiz at 50% student score and no second quiz -> quiz_grade = 25."""
    m = app_module
    with m.app.app_context():
        prof = m.Professor(
            username="qzprof1",
            email="qz1@test.local",
            password_hash=generate_password_hash("x"),
        )
        m.db.session.add(prof)
        m.db.session.commit()
        c = m.Class(professor_id=prof.id, name="Qz Class", class_code="QZ001")
        m.db.session.add(c)
        m.db.session.commit()

        gw = m.GradingWeights(
            class_id=c.id,
            attendance_weight=25.0,
            participation_weight=25.0,
            participation_instructor_share=50.0,
            poll_weight=25.0,
            quiz_weight=25.0,
            quiz_count_target=2,
        )
        m.db.session.add(gw)

        st = m.Student(
            first_name="Q",
            last_name="Student",
            student_number="222333444",
            email="qs@test.local",
        )
        m.db.session.add(st)
        m.db.session.commit()
        m.db.session.add(m.Enrollment(class_id=c.id, student_id=st.id, is_active=True))

        t0 = datetime.utcnow() - timedelta(days=1)
        sess = m.ClassSession(
            class_id=c.id,
            start_time=t0,
            end_time=t0 + timedelta(hours=1),
            exclude_from_grading=False,
        )
        m.db.session.add(sess)
        m.db.session.commit()

        m.db.session.add(
            m.Attendance(
                class_id=c.id,
                student_id=st.id,
                class_session_id=sess.id,
                date=t0.date(),
                present=True,
                join_time=t0,
            )
        )

        qz = m.Quiz(
            class_id=c.id,
            title="Q1",
            time_limit_seconds=300,
            quiz_index=1,
        )
        m.db.session.add(qz)
        m.db.session.flush()
        q1 = m.QuizQuestion(
            quiz_id=qz.id,
            order=1,
            prompt="A?",
            options='["x","y"]',
            correct_index=0,
        )
        q2 = m.QuizQuestion(
            quiz_id=qz.id,
            order=2,
            prompt="B?",
            options='["p","q"]',
            correct_index=1,
        )
        m.db.session.add_all([q1, q2])
        m.db.session.flush()

        run = m.QuizRun(
            quiz_id=qz.id,
            class_id=c.id,
            started_at=t0 + timedelta(minutes=5),
            deadline_at=t0 + timedelta(minutes=30),
            ended_at=t0 + timedelta(minutes=20),
            is_active=False,
        )
        m.db.session.add(run)
        m.db.session.flush()

        m.db.session.add(
            m.QuizAnswer(
                quiz_run_id=run.id,
                student_id=st.id,
                question_id=q1.id,
                selected_index=0,
                is_correct=True,
            )
        )
        m.db.session.add(
            m.QuizAnswer(
                quiz_run_id=run.id,
                student_id=st.id,
                question_id=q2.id,
                selected_index=0,
                is_correct=False,
            )
        )
        m.db.session.commit()

        rows = m._compute_gradebook_rows(c.id, gw)
        assert len(rows) == 1
        r = rows[0]
        assert abs(r["quiz_grade"] - 25.0) < 1e-6
        assert r["quiz_scores_by_index"].get(1) == 50.0
        assert r["quiz_scores_by_index"].get(2) == 0.0


def test_quiz_weight_redistributes_to_attendance_when_no_quiz_sessions(app_module):
    """With quiz_weight>0 but no quiz runs in graded sessions, eff_quiz is 0 and attendance absorbs quiz share."""
    m = app_module
    with m.app.app_context():
        prof = m.Professor(
            username="qzprof2",
            email="qz2@test.local",
            password_hash=generate_password_hash("x"),
        )
        m.db.session.add(prof)
        m.db.session.commit()
        c = m.Class(professor_id=prof.id, name="Qz Class 2", class_code="QZ002")
        m.db.session.add(c)
        m.db.session.commit()

        gw = m.GradingWeights(
            class_id=c.id,
            attendance_weight=50.0,
            participation_weight=0.0,
            participation_instructor_share=50.0,
            poll_weight=0.0,
            quiz_weight=50.0,
            quiz_count_target=1,
        )
        m.db.session.add(gw)

        st = m.Student(
            first_name="Z",
            last_name="Two",
            student_number="333444555",
            email="zt@test.local",
        )
        m.db.session.add(st)
        m.db.session.commit()
        m.db.session.add(m.Enrollment(class_id=c.id, student_id=st.id, is_active=True))

        t0 = datetime.utcnow() - timedelta(days=2)
        sess = m.ClassSession(
            class_id=c.id,
            start_time=t0,
            end_time=t0 + timedelta(hours=1),
            exclude_from_grading=False,
        )
        m.db.session.add(sess)
        m.db.session.commit()
        m.db.session.add(
            m.Attendance(
                class_id=c.id,
                student_id=st.id,
                class_session_id=sess.id,
                date=t0.date(),
                present=True,
                join_time=t0,
            )
        )
        m.db.session.commit()

        rows = m._compute_gradebook_rows(c.id, gw)
        r = rows[0]
        assert abs(r["overall_grade"] - 100.0) < 1e-6


def test_quiz_upload_rejected_when_quiz_count_target_zero(app_module):
    """Upload is blocked until Number of quizzes is set (server-side)."""
    m = app_module
    class_id = None
    with m.app.app_context():
        prof = m.Professor(
            username="qzprof_upload_lock",
            email="qz_upload_lock@test.local",
            password_hash=generate_password_hash("secret12"),
        )
        m.db.session.add(prof)
        m.db.session.commit()
        c = m.Class(professor_id=prof.id, name="Qz Lock Upload", class_code="QZUPL1")
        m.db.session.add(c)
        m.db.session.commit()
        class_id = c.id
        m.db.session.add(
            m.GradingWeights(
                class_id=c.id,
                attendance_weight=100.0,
                participation_weight=0.0,
                participation_instructor_share=50.0,
                poll_weight=0.0,
                quiz_weight=0.0,
                quiz_count_target=0,
            )
        )
        m.db.session.commit()

    with m.app.test_client() as client:
        client.post(
            "/login",
            data={
                "username": "qzprof_upload_lock",
                "password": "secret12",
                "user_type": "professor",
            },
        )
        rv = client.post(
            f"/api/quiz_upload/{class_id}",
            data={
                "file": (BytesIO(b"dummy"), "quiz.xlsx"),
                "quiz_index": "1",
            },
            content_type="multipart/form-data",
        )
    assert rv.status_code == 400
    err = (rv.get_json() or {}).get("error", "")
    assert "number of quizzes" in err.lower()


def test_start_quiz_rejected_when_quiz_count_target_zero(app_module):
    """Starting a run is blocked until Number of quizzes is set (server-side)."""
    m = app_module
    quiz_id = None
    with m.app.app_context():
        prof = m.Professor(
            username="qzprof_start_lock",
            email="qz_start_lock@test.local",
            password_hash=generate_password_hash("secret12"),
        )
        m.db.session.add(prof)
        m.db.session.commit()
        c = m.Class(
            professor_id=prof.id,
            name="Qz Lock Start",
            class_code="QZST1",
            is_active=True,
        )
        m.db.session.add(c)
        m.db.session.commit()
        m.db.session.add(
            m.GradingWeights(
                class_id=c.id,
                attendance_weight=100.0,
                participation_weight=0.0,
                participation_instructor_share=50.0,
                poll_weight=0.0,
                quiz_weight=0.0,
                quiz_count_target=0,
            )
        )
        m.db.session.add(
            m.ClassSession(
                class_id=c.id,
                start_time=datetime.utcnow(),
                end_time=None,
                exclude_from_grading=False,
            )
        )
        qz = m.Quiz(
            class_id=c.id,
            title="Slot 1",
            quiz_index=1,
            time_limit_seconds=120,
        )
        m.db.session.add(qz)
        m.db.session.commit()
        m.db.session.add(
            m.QuizQuestion(
                quiz_id=qz.id,
                order=0,
                prompt="2+2?",
                options=json.dumps(["3", "4", "5"]),
                correct_index=1,
            )
        )
        m.db.session.commit()
        quiz_id = qz.id

    with m.app.test_client() as client:
        client.post(
            "/login",
            data={
                "username": "qzprof_start_lock",
                "password": "secret12",
                "user_type": "professor",
            },
        )
        rv = client.post(
            f"/api/quiz/{quiz_id}/start",
            json={},
            content_type="application/json",
        )
    assert rv.status_code == 400
    err = (rv.get_json() or {}).get("error", "")
    assert "number of quizzes" in err.lower()


def test_parse_quiz_workbook_letter_template(app_module):
    """Option A/B layout + Question Description (matches static quiz_template.xlsx)."""
    m = app_module
    wb = Workbook()
    ws = wb.active
    ws.append(
        [
            'Question #',
            'Question Description',
            '# of Options',
            'Option A',
            'Option B',
            'Option C',
            'Correct Answer',
        ]
    )
    ws.append([1, 'Capitol of France?', 2, 'London', 'Paris', '', 'B'])
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    wb_r = load_workbook(bio, read_only=True, data_only=True)
    questions, err = m._parse_quiz_workbook(wb_r.active)
    wb_r.close()
    assert err == []
    assert len(questions) == 1
    assert questions[0]['prompt'] == 'Capitol of France?'
    assert questions[0]['options'] == ['London', 'Paris']
    assert questions[0]['correct_index'] == 1


def test_parse_quiz_excel_correct_answer(app_module):
    p = app_module._parse_quiz_excel_correct_answer
    assert p(1, 4) == 1
    assert p(4, 4) == 4
    assert p(3.0, 4) == 3
    assert p("B", 4) == 2
    assert p("b", 4) == 2
    assert p("C.", 4) == 3
    assert p("3)", 4) == 3
    assert p("3.0", 4) == 3
    assert p("E", 4) is None
    assert p("Z", 4) is None
    assert p(None, 4) is None
    assert p("", 4) is None
    assert p(2.5, 4) is None
